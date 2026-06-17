"""Local validation for registered manual temp-table workbooks."""

from __future__ import annotations

import re
from collections import defaultdict
from pathlib import Path
from typing import Any, Iterable

from _shared.errors import UsageError

from usql_web_query.manual_table_registry import ManualTableEntry


MAX_SAMPLES = 20


def validate_manual_table(file_path: Path, entry: ManualTableEntry | None) -> dict[str, Any]:
    """Validate a registered workbook and return JSON-serializable issues."""

    if entry is None:
        return {
            "checked": False,
            "ok": True,
            "status": "skipped",
            "message": "No manual-table registry entry matched this file.",
            "issues": [],
        }
    if not file_path.exists():
        return _summary(
            file_path=file_path,
            entry=entry,
            checked=False,
            status="error",
            issues=[
                _issue(
                    "error",
                    "file_exists",
                    f"Manual table file does not exist: {file_path}",
                )
            ],
        )
    if file_path.suffix.lower() == ".xls":
        return {
            "checked": False,
            "ok": True,
            "status": "skipped",
            "message": "Legacy .xls validation is not supported locally; upload validation is left to the platform.",
            "registry_entry": entry.as_summary(),
            "issues": [],
        }
    if file_path.suffix.lower() not in {".xlsx", ".xlsm"}:
        return {
            "checked": False,
            "ok": True,
            "status": "skipped",
            "message": "Manual-table validation currently supports Excel workbooks only.",
            "registry_entry": entry.as_summary(),
            "issues": [],
        }

    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - local dependency guard
        raise UsageError("openpyxl is required for manual-table validation.") from exc

    workbook = load_workbook(file_path, read_only=True, data_only=True)
    try:
        configured_sheet = entry.raw.get("sheet")
        issues: list[dict[str, Any]] = []
        if configured_sheet and configured_sheet in workbook.sheetnames:
            worksheet = workbook[configured_sheet]
        elif configured_sheet:
            worksheet = workbook[workbook.sheetnames[0]]
            issues.append(
                _issue(
                    "error",
                    "sheet_exists",
                    f"Configured sheet not found: {configured_sheet}; inspected {worksheet.title} instead.",
                )
            )
        else:
            worksheet = workbook[workbook.sheetnames[0]]

        rows_iter = worksheet.iter_rows(values_only=True)
        try:
            header_values = next(rows_iter)
        except StopIteration:
            return _summary(
                file_path=file_path,
                entry=entry,
                checked=True,
                status="error",
                sheet=worksheet.title,
                headers=[],
                row_count=0,
                issues=[_issue("error", "empty_workbook", "Workbook sheet has no rows.")],
            )

        headers = [_normalize_header(value) for value in header_values]
        positions = _header_positions(headers)
        data_rows = list(_iter_data_rows(rows_iter))

        for rule in entry.raw.get("validation_rules", []):
            issues.extend(_run_rule(rule, headers, positions, data_rows))

        return _summary(
            file_path=file_path,
            entry=entry,
            checked=True,
            status="passed" if not issues else "issues_found",
            sheet=worksheet.title,
            headers=headers,
            row_count=len(data_rows),
            issues=issues,
        )
    finally:
        workbook.close()


def _run_rule(
    rule: dict[str, Any],
    headers: list[str],
    positions: dict[str, int],
    data_rows: list[tuple[int, tuple[Any, ...]]],
) -> list[dict[str, Any]]:
    rule_type = rule.get("type")
    if rule_type == "required_columns":
        return _required_columns(rule, positions)
    if rule_type == "no_blank_header":
        return _no_blank_header(headers)
    if rule_type == "lowercase_ascii_prefix":
        return _lowercase_ascii_prefix(rule, positions, data_rows)
    if rule_type == "no_blank_or_dash":
        return _no_blank_or_dash(rule, positions, data_rows)
    if rule_type == "disallow_values":
        return _disallow_values(rule, positions, data_rows)
    if rule_type == "unique_key":
        return _unique_key(rule, positions, data_rows)
    if rule_type == "name_variant_consistency":
        return _name_variant_consistency(rule, positions, data_rows)
    if rule_type == "infer_missing_from_mapping":
        return _infer_missing_from_mapping(rule, positions, data_rows)
    if rule_type == "one_to_many_mapping":
        return _one_to_many_mapping(rule, positions, data_rows)
    if rule_type == "external_coverage_check":
        return [
            _issue(
                "info",
                "external_coverage_check",
                f"{', '.join(rule.get('columns', []))} coverage requires SQL-side comparison against {rule.get('scope', 'external data')}.",
                details={"columns": rule.get("columns", []), "scope": rule.get("scope")},
            )
        ]
    return [_issue("warning", "unknown_rule", f"Unknown manual-table validation rule: {rule_type}", details=rule)]


def _required_columns(rule: dict[str, Any], positions: dict[str, int]) -> list[dict[str, Any]]:
    missing = [column for column in rule.get("columns", []) if column not in positions]
    if not missing:
        return []
    return [
        _issue(
            "error",
            "required_columns",
            f"Missing required columns: {', '.join(missing)}",
            details={"missing_columns": missing},
        )
    ]


def _no_blank_header(headers: list[str]) -> list[dict[str, Any]]:
    blank_columns = [index + 1 for index, value in enumerate(headers) if not value]
    if not blank_columns:
        return []
    return [
        _issue(
            "error",
            "no_blank_header",
            f"Blank header cells found at column numbers: {blank_columns}",
            details={"blank_column_numbers": blank_columns},
        )
    ]


def _lowercase_ascii_prefix(
    rule: dict[str, Any],
    positions: dict[str, int],
    data_rows: list[tuple[int, tuple[Any, ...]]],
) -> list[dict[str, Any]]:
    issues = []
    for column in rule.get("columns", []):
        if column not in positions:
            continue
        invalid_rows = []
        for row_number, row in data_rows:
            text = _to_text(_cell(row, positions[column]))
            first_ascii = _first_ascii_letter(text)
            if first_ascii and first_ascii != first_ascii.lower():
                invalid_rows.append({"row": row_number, "value": text})
        if invalid_rows:
            issues.append(
                _issue(
                    "error",
                    "lowercase_ascii_prefix",
                    f"{column} has values whose first ASCII letter is not lowercase.",
                    column=column,
                    rows=[item["row"] for item in invalid_rows[:MAX_SAMPLES]],
                    count=len(invalid_rows),
                    details={"samples": invalid_rows[:MAX_SAMPLES]},
                )
            )
    return issues


def _no_blank_or_dash(
    rule: dict[str, Any],
    positions: dict[str, int],
    data_rows: list[tuple[int, tuple[Any, ...]]],
) -> list[dict[str, Any]]:
    issues = []
    for column in rule.get("columns", []):
        if column not in positions:
            continue
        invalid = [
            row_number
            for row_number, row in data_rows
            if _is_blank_or_dash(_cell(row, positions[column]))
        ]
        if invalid:
            issues.append(
                _issue(
                    "error",
                    "no_blank_or_dash",
                    f"{column} cannot be blank or '-'.",
                    column=column,
                    rows=invalid[:MAX_SAMPLES],
                    count=len(invalid),
                )
            )
    return issues


def _disallow_values(
    rule: dict[str, Any],
    positions: dict[str, int],
    data_rows: list[tuple[int, tuple[Any, ...]]],
) -> list[dict[str, Any]]:
    column = str(rule.get("column", ""))
    if column not in positions:
        return []
    disallowed = {_canonical_value(value) for value in rule.get("values", [])}
    invalid = []
    for row_number, row in data_rows:
        value = _cell(row, positions[column])
        if _canonical_value(value) in disallowed:
            invalid.append({"row": row_number, "value": _to_text(value)})
    if not invalid:
        return []
    return [
        _issue(
            "error",
            "disallow_values",
            f"{column} contains disallowed values: {sorted(disallowed)}",
            column=column,
            rows=[item["row"] for item in invalid[:MAX_SAMPLES]],
            count=len(invalid),
            details={"samples": invalid[:MAX_SAMPLES]},
        )
    ]


def _unique_key(
    rule: dict[str, Any],
    positions: dict[str, int],
    data_rows: list[tuple[int, tuple[Any, ...]]],
) -> list[dict[str, Any]]:
    columns = [column for column in rule.get("columns", []) if column in positions]
    if len(columns) != len(rule.get("columns", [])):
        return []
    groups: dict[tuple[str, ...], list[int]] = defaultdict(list)
    for row_number, row in data_rows:
        key = tuple(_canonical_value(_cell(row, positions[column])) for column in columns)
        if rule.get("ignore_blank") and any(not value for value in key):
            continue
        groups[key].append(row_number)
    duplicates = [
        {"key": dict(zip(columns, key)), "rows": rows[:MAX_SAMPLES], "count": len(rows)}
        for key, rows in groups.items()
        if len(rows) > 1
    ]
    if not duplicates:
        return []
    return [
        _issue(
            "error",
            "unique_key",
            f"Duplicate key values found for {', '.join(columns)}.",
            count=len(duplicates),
            details={"duplicate_groups": duplicates[:MAX_SAMPLES]},
        )
    ]


def _name_variant_consistency(
    rule: dict[str, Any],
    positions: dict[str, int],
    data_rows: list[tuple[int, tuple[Any, ...]]],
) -> list[dict[str, Any]]:
    column = str(rule.get("column", ""))
    if column not in positions:
        return []
    grouped: dict[str, dict[str, list[int]]] = defaultdict(lambda: defaultdict(list))
    for row_number, row in data_rows:
        value = _to_text(_cell(row, positions[column]))
        if not value:
            continue
        base = _strip_numeric_suffix(value)
        grouped[base][value].append(row_number)

    inconsistent = []
    total_rows = 0
    for base, variants in grouped.items():
        if len(variants) < 2:
            continue
        preferred = _preferred_numeric_variant(variants.keys())
        if not preferred:
            continue
        rows_to_change = [
            row_number
            for variant, rows in variants.items()
            if variant != preferred
            for row_number in rows
        ]
        if rows_to_change:
            total_rows += len(rows_to_change)
            inconsistent.append(
                {
                    "base": base,
                    "variants": sorted(variants.keys()),
                    "suggested_value": preferred,
                    "rows_to_review": rows_to_change[:MAX_SAMPLES],
                    "row_count": len(rows_to_change),
                }
            )
    if not inconsistent:
        return []
    return [
        _issue(
            "warning",
            "name_variant_consistency",
            f"{column} has plain and numeric-suffixed name variants.",
            column=column,
            count=total_rows,
            details={"groups": inconsistent[:MAX_SAMPLES]},
        )
    ]


def _infer_missing_from_mapping(
    rule: dict[str, Any],
    positions: dict[str, int],
    data_rows: list[tuple[int, tuple[Any, ...]]],
) -> list[dict[str, Any]]:
    source_column = str(rule.get("source_column", ""))
    target_column = str(rule.get("target_column", ""))
    if source_column not in positions or target_column not in positions:
        return []
    blank_values = {_canonical_value(value) for value in rule.get("blank_values", ["", "-"])}
    mapping: dict[str, set[str]] = defaultdict(set)
    for _, row in data_rows:
        source_value = _canonical_value(_cell(row, positions[source_column]))
        target_value = _canonical_value(_cell(row, positions[target_column]))
        if source_value and target_value not in blank_values:
            mapping[source_value].add(target_value)

    suggestions = []
    unresolved = []
    for row_number, row in data_rows:
        source_value = _canonical_value(_cell(row, positions[source_column]))
        target_value = _canonical_value(_cell(row, positions[target_column]))
        if target_value not in blank_values:
            continue
        candidates = sorted(mapping.get(source_value, set()))
        if len(candidates) == 1:
            suggestions.append({"row": row_number, source_column: source_value, "suggested_value": candidates[0]})
        elif source_value:
            unresolved.append({"row": row_number, source_column: source_value, "candidate_count": len(candidates)})

    issues = []
    if suggestions:
        issues.append(
            _issue(
                "warning",
                "infer_missing_from_mapping",
                f"{target_column} has missing values that can be inferred from {source_column}.",
                column=target_column,
                rows=[item["row"] for item in suggestions[:MAX_SAMPLES]],
                count=len(suggestions),
                details={"suggestions": suggestions[:MAX_SAMPLES]},
            )
        )
    if unresolved:
        issues.append(
            _issue(
                "warning",
                "infer_missing_from_mapping_unresolved",
                f"{target_column} has missing values with no unique {source_column} mapping.",
                column=target_column,
                rows=[item["row"] for item in unresolved[:MAX_SAMPLES]],
                count=len(unresolved),
                details={"samples": unresolved[:MAX_SAMPLES]},
            )
        )
    return issues


def _one_to_many_mapping(
    rule: dict[str, Any],
    positions: dict[str, int],
    data_rows: list[tuple[int, tuple[Any, ...]]],
) -> list[dict[str, Any]]:
    key_columns = [column for column in rule.get("key_columns", []) if column in positions]
    value_column = str(rule.get("value_column", ""))
    if len(key_columns) != len(rule.get("key_columns", [])) or value_column not in positions:
        return []

    grouped: dict[tuple[str, ...], dict[str, list[int]]] = defaultdict(lambda: defaultdict(list))
    for row_number, row in data_rows:
        key = tuple(_canonical_value(_cell(row, positions[column])) for column in key_columns)
        value = _canonical_value(_cell(row, positions[value_column]))
        if not all(key) or not value:
            continue
        grouped[key][value].append(row_number)

    conflicts = []
    for key, value_map in grouped.items():
        if len(value_map) <= 1:
            continue
        conflicts.append(
            {
                "key": dict(zip(key_columns, key)),
                "values": sorted(value_map.keys()),
                "rows": sorted({row for rows in value_map.values() for row in rows})[:MAX_SAMPLES],
            }
        )
    if not conflicts:
        return []
    return [
        _issue(
            "warning",
            "one_to_many_mapping",
            rule.get("message") or f"{', '.join(key_columns)} maps to multiple {value_column} values.",
            column=value_column,
            count=len(conflicts),
            details={"conflicts": conflicts[:MAX_SAMPLES]},
        )
    ]


def _summary(
    *,
    file_path: Path,
    entry: ManualTableEntry,
    checked: bool,
    status: str,
    issues: list[dict[str, Any]],
    sheet: str | None = None,
    headers: list[str] | None = None,
    row_count: int | None = None,
) -> dict[str, Any]:
    return {
        "checked": checked,
        "ok": _count_by_severity(issues, "error") == 0,
        "status": status,
        "file_path": str(file_path),
        "registry_entry": entry.as_summary(),
        "sheet": sheet,
        "row_count": row_count,
        "headers": headers,
        "issue_count": len(issues),
        "error_count": _count_by_severity(issues, "error"),
        "warning_count": _count_by_severity(issues, "warning"),
        "info_count": _count_by_severity(issues, "info"),
        "issues": issues,
    }


def _issue(
    severity: str,
    rule: str,
    message: str,
    *,
    column: str | None = None,
    rows: list[int] | None = None,
    count: int | None = None,
    details: dict[str, Any] | None = None,
) -> dict[str, Any]:
    issue: dict[str, Any] = {"severity": severity, "rule": rule, "message": message}
    if column is not None:
        issue["column"] = column
    if rows is not None:
        issue["rows"] = rows
    if count is not None:
        issue["count"] = count
    if details:
        issue["details"] = details
    return issue


def _iter_data_rows(rows: Iterable[tuple[Any, ...]]) -> Iterable[tuple[int, tuple[Any, ...]]]:
    for offset, row in enumerate(rows, start=2):
        if any(not _is_blank(value) for value in row):
            yield offset, tuple(row)


def _header_positions(headers: list[str]) -> dict[str, int]:
    positions: dict[str, int] = {}
    for index, header in enumerate(headers):
        if header and header not in positions:
            positions[header] = index
    return positions


def _normalize_header(value: Any) -> str:
    return _to_text(value)


def _cell(row: tuple[Any, ...], index: int) -> Any:
    return row[index] if index < len(row) else None


def _to_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value)).strip()
    return str(value).strip()


def _canonical_value(value: Any) -> str:
    return _to_text(value)


def _is_blank(value: Any) -> bool:
    return _to_text(value) == ""


def _is_blank_or_dash(value: Any) -> bool:
    return _to_text(value) in {"", "-"}


def _first_ascii_letter(value: str) -> str:
    for char in value.strip():
        if ("A" <= char <= "Z") or ("a" <= char <= "z"):
            return char
    return ""


def _strip_numeric_suffix(value: str) -> str:
    return re.sub(r"\d+$", "", value)


def _preferred_numeric_variant(values: Iterable[str]) -> str | None:
    suffixed = []
    for value in values:
        match = re.search(r"(\d+)$", value)
        if match:
            suffixed.append((int(match.group(1)), len(match.group(1)), value))
    if not suffixed:
        return None
    suffixed.sort(key=lambda item: (item[0], item[1], item[2]))
    return suffixed[-1][2]


def _count_by_severity(issues: list[dict[str, Any]], severity: str) -> int:
    return sum(1 for issue in issues if issue.get("severity") == severity)
