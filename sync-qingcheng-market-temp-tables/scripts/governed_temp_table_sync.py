from __future__ import annotations

import argparse
import copy
import errno
import hashlib
import json
import os
import posixpath
import re
import shutil
import subprocess
import sys
import time
import zipfile
from collections import Counter, defaultdict
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any, Callable, Iterable
from xml.etree import ElementTree

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from docs_sheet_downloader import download_docs_sheet


SKILL_ROOT = Path(__file__).resolve().parents[1]
SKILLS_ROOT = SKILL_ROOT.parent
CODEX_ROOT = SKILLS_ROOT.parent
DEFAULT_REGISTRY = SKILL_ROOT / "references" / "workflow_registry.json"
DEFAULT_RUNTIME_ROOT = CODEX_ROOT / "runtime" / "sync-qingcheng-market-temp-tables"
DEFAULT_SOURCE_BASELINE_SEED = (
    SKILL_ROOT / "references" / "source_slice_baselines.json"
)
DEFAULT_SOURCE_BASELINE_STATE = (
    DEFAULT_RUNTIME_ROOT / "state" / "source_slice_baselines.json"
)
OPERATOR_ROOT = SKILLS_ROOT / "usql-web-query-operator"
OPERATOR_SCRIPT = OPERATOR_ROOT / "scripts" / "usql_web_query.py"
OPERATOR_SCRIPTS = OPERATOR_ROOT / "scripts"
RECALC_SCRIPT = SKILLS_ROOT / "xlsx" / "scripts" / "recalc.py"
ERROR_VALUES = {"#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#NUM!", "#NULL!"}
SPREADSHEET_MAIN_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
OFFICE_REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
PACKAGE_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
EXTERNAL_LINK_REL_TYPE = f"{OFFICE_REL_NS}/externalLink"
EXTERNAL_LINK_PART_PREFIX = "xl/externalLinks/"
REPLACE_MAX_ATTEMPTS = 8
REPLACE_INITIAL_DELAY_SECONDS = 0.25
REPLACE_MAX_DELAY_SECONDS = 2.0
RETRYABLE_REPLACE_ERRNOS = {errno.EACCES, errno.EPERM}
RETRYABLE_REPLACE_WINERRORS = {5, 32, 33}


class WorkflowError(RuntimeError):
    pass


class AtomicReplaceError(WorkflowError):
    def __init__(self, message: str, details: dict[str, Any]) -> None:
        super().__init__(message)
        self.details = details


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def exception_summary(exc: BaseException) -> dict[str, Any]:
    return {
        "type": type(exc).__name__,
        "message": str(exc),
        "errno": getattr(exc, "errno", None),
        "winerror": getattr(exc, "winerror", None),
    }


def sibling_temp_path(target: Path, label: str) -> Path:
    return target.with_name(
        f".{target.name}.{os.getpid()}.{time.time_ns()}.{label}.tmp"
    )


def replace_file_with_retry(
    source: Path,
    target: Path,
    *,
    max_attempts: int = REPLACE_MAX_ATTEMPTS,
    initial_delay_seconds: float = REPLACE_INITIAL_DELAY_SECONDS,
    replace_func: Callable[[Path, Path], None] | None = None,
    sleep_func: Callable[[float], None] | None = None,
) -> dict[str, Any]:
    if max_attempts < 1:
        raise ValueError("max_attempts must be at least 1")
    replace = replace_func or os.replace
    sleep = sleep_func or time.sleep
    delay = max(initial_delay_seconds, 0.0)
    retry_errors: list[dict[str, Any]] = []
    for attempt in range(1, max_attempts + 1):
        try:
            replace(source, target)
            return {
                "status": "success",
                "attempts": attempt,
                "retries": attempt - 1,
                "retry_errors": retry_errors,
            }
        except OSError as exc:
            details = exception_summary(exc)
            details["attempt"] = attempt
            retry_errors.append(details)
            retryable = (
                isinstance(exc, PermissionError)
                or details["errno"] in RETRYABLE_REPLACE_ERRNOS
                or details["winerror"] in RETRYABLE_REPLACE_WINERRORS
            )
            if not retryable or attempt == max_attempts:
                raise AtomicReplaceError(
                    (
                        "Atomic workbook replacement failed "
                        f"after {attempt} attempt(s) "
                        f"(errno={details['errno']}, "
                        f"winerror={details['winerror']}). "
                        "The target may be open or temporarily locked: "
                        f"{target}"
                    ),
                    {
                        "status": "failed",
                        "attempts": attempt,
                        "retries": attempt - 1,
                        "retry_errors": retry_errors,
                    },
                ) from exc
            if delay:
                sleep(delay)
            delay = min(
                max(delay * 2, REPLACE_INITIAL_DELAY_SECONDS),
                REPLACE_MAX_DELAY_SECONDS,
            )


def canonical_json(value: Any) -> bytes:
    return json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":")).encode("utf-8")


def artifact_hash(value: dict[str, Any], hash_field: str) -> str:
    payload = dict(value)
    payload.pop(hash_field, None)
    return hashlib.sha256(canonical_json(payload)).hexdigest()


def write_artifact(path: Path, value: dict[str, Any], hash_field: str) -> str:
    value[hash_field] = artifact_hash(value, hash_field)
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(value, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    return value[hash_field]


def load_artifact(path: Path, hash_field: str, expected_hash: str | None = None) -> dict[str, Any]:
    value = json.loads(path.read_text(encoding="utf-8"))
    actual_hash = artifact_hash(value, hash_field)
    if value.get(hash_field) != actual_hash:
        raise WorkflowError(f"Artifact hash mismatch: {path}")
    if expected_hash and expected_hash != actual_hash:
        raise WorkflowError(f"Expected {hash_field}={expected_hash}, got {actual_hash}")
    return value


def load_registry(path: Path) -> dict[str, Any]:
    registry = json.loads(path.read_text(encoding="utf-8"))
    domains = registry.get("domains")
    if not isinstance(domains, dict) or not domains:
        raise WorkflowError("Workflow registry domains must be a non-empty object.")
    chat_ids = []
    for domain_id, domain in domains.items():
        if not re.fullmatch(r"[a-z][a-z0-9_]*", domain_id):
            raise WorkflowError(f"Invalid domain id: {domain_id}")
        chat = domain.get("chat") or {}
        chat_id = str(chat.get("expected_chat_id") or "")
        if not re.fullmatch(r"oc_[A-Za-z0-9]+", chat_id):
            raise WorkflowError(
                f"Domain {domain_id} must configure one exact expected_chat_id."
            )
        if not str(chat.get("name") or "").strip():
            raise WorkflowError(f"Domain {domain_id} must configure a chat name.")
        chat_ids.append(chat_id)
    if len(chat_ids) != len(set(chat_ids)):
        raise WorkflowError("Each workflow domain must use a distinct chat id.")
    ids = [family["id"] for family in registry.get("families", [])]
    if len(ids) != len(set(ids)):
        raise WorkflowError("Workflow registry contains duplicate family ids.")
    upload_order = registry.get("upload_order") or []
    if len(upload_order) != len(set(upload_order)) or set(upload_order) != set(ids):
        raise WorkflowError("Workflow registry upload_order must contain every family id exactly once.")
    for family in registry.get("families", []):
        domain_id = family.get("domain")
        if domain_id not in domains:
            raise WorkflowError(
                f"Family {family.get('id')} has unknown domain: {domain_id}"
            )
        source_kind = family.get("source_kind", "file_attachment")
        if source_kind not in {"file_attachment", "link_workbook"}:
            raise WorkflowError(f"Unsupported source_kind for {family['id']}: {source_kind}")
        if source_kind == "file_attachment" and not family.get("source_filename_patterns"):
            raise WorkflowError(f"File source family has no filename patterns: {family['id']}")
        if source_kind == "link_workbook":
            required = (
                "source_sender_open_id",
                "source_url_patterns",
                "source_expected_title_pattern",
                "source_env_file",
                "source_env_section",
                "source_filename",
            )
            missing = [key for key in required if not family.get(key)]
            if missing:
                raise WorkflowError(f"Link source family {family['id']} is missing: {missing}")
        quality = family.get("source_quality")
        if registry.get("require_source_quality_gates") and not quality:
            raise WorkflowError(f"Source quality gate is required for family {family['id']}.")
        if quality:
            _validate_source_quality_config(family, quality)
        merge_mode = family.get("source_merge_mode", "all_source_slices")
        if merge_mode not in {
            "all_source_slices",
            "changed_source_slices",
            "full_source_replace",
        }:
            raise WorkflowError(
                f"Unsupported source_merge_mode for {family['id']}: {merge_mode}"
            )
        if merge_mode == "changed_source_slices":
            baseline_id = str(family.get("source_baseline_id") or "")
            if not baseline_id:
                raise WorkflowError(
                    f"Changed-slice family {family['id']} needs source_baseline_id."
                )
            maximum = family.get("max_changed_slices")
            if not isinstance(maximum, int) or isinstance(maximum, bool) or maximum < 1:
                raise WorkflowError(
                    f"Changed-slice family {family['id']} needs max_changed_slices >= 1."
                )
    _validate_local_temp_table_inventories(registry)
    return registry


def _validate_local_temp_table_inventories(registry: dict[str, Any]) -> None:
    inventories = registry.get("local_temp_table_inventories")
    if inventories is None:
        return
    if not isinstance(inventories, dict) or not inventories:
        raise WorkflowError(
            "Workflow registry local_temp_table_inventories must be a non-empty object."
        )

    domains = registry["domains"]
    families = {
        str(family["id"]): family for family in registry.get("families", [])
    }
    deferred_patterns = list(registry.get("deferred_filename_patterns") or [])
    for domain_id, inventory in inventories.items():
        if domain_id not in domains:
            raise WorkflowError(
                f"Local temp-table inventory has unknown domain: {domain_id}"
            )
        root = str(inventory.get("root") or "")
        if not root or not Path(root).is_absolute():
            raise WorkflowError(
                f"Local temp-table inventory {domain_id} needs an absolute root."
            )
        platform_database = str(inventory.get("platform_database") or "")
        if platform_database != "temp_table":
            raise WorkflowError(
                f"Local temp-table inventory {domain_id} must use platform_database=temp_table."
            )
        verification = inventory.get("verification") or {}
        if (
            verification.get("method") != "zero_row_resolution_probe"
            or verification.get("status") != "success"
            or not re.fullmatch(r"\d+", str(verification.get("query_id") or ""))
        ):
            raise WorkflowError(
                f"Local temp-table inventory {domain_id} lacks successful live verification."
            )
        mappings = inventory.get("mappings")
        if not isinstance(mappings, list) or not mappings:
            raise WorkflowError(
                f"Local temp-table inventory {domain_id} has no mappings."
            )

        local_names: set[str] = set()
        platform_names: set[str] = set()
        managed_family_ids: set[str] = set()
        normalized_root = os.path.normcase(os.path.normpath(root))
        for mapping in mappings:
            local_filename = str(mapping.get("local_filename") or "")
            local_key = local_filename.casefold()
            if (
                not local_filename
                or Path(local_filename).name != local_filename
                or Path(local_filename).suffix.casefold() != ".xlsx"
            ):
                raise WorkflowError(
                    f"Invalid local temp-table filename in {domain_id}: {local_filename}"
                )
            if local_key in local_names:
                raise WorkflowError(
                    f"Duplicate local temp-table filename in {domain_id}: {local_filename}"
                )
            local_names.add(local_key)

            platform_temp_table = str(mapping.get("platform_temp_table") or "")
            platform_key = platform_temp_table.casefold()
            if not re.fullmatch(r"dingxi01_[a-z0-9_]+", platform_temp_table):
                raise WorkflowError(
                    f"Invalid platform temp-table name in {domain_id}: {platform_temp_table}"
                )
            if platform_key in platform_names:
                raise WorkflowError(
                    f"Duplicate platform temp-table target in {domain_id}: "
                    f"{platform_temp_table}"
                )
            platform_names.add(platform_key)
            if mapping.get("mapping_status") != "live_verified_exact_name":
                raise WorkflowError(
                    f"Unverified local/platform mapping in {domain_id}: {local_filename}"
                )

            automation_scope = str(mapping.get("automation_scope") or "")
            family_id = str(mapping.get("workflow_family_id") or "")
            if automation_scope == "managed":
                if not family_id or family_id not in families:
                    raise WorkflowError(
                        f"Managed mapping {local_filename} needs a valid workflow_family_id."
                    )
                family = families[family_id]
                if family.get("domain") != domain_id:
                    raise WorkflowError(
                        f"Managed mapping {local_filename} crosses workflow domains."
                    )
                target_workbook = Path(str(family.get("target_workbook") or ""))
                if target_workbook.name.casefold() != local_key:
                    raise WorkflowError(
                        f"Managed mapping filename drifts from family {family_id}."
                    )
                if (
                    os.path.normcase(os.path.normpath(str(target_workbook.parent)))
                    != normalized_root
                ):
                    raise WorkflowError(
                        f"Managed mapping root drifts from family {family_id}."
                    )
                if family.get("platform_temp_table") != platform_temp_table:
                    raise WorkflowError(
                        f"Managed mapping target drifts from family {family_id}."
                    )
                if family_id in managed_family_ids:
                    raise WorkflowError(
                        f"Workflow family is mapped more than once: {family_id}"
                    )
                managed_family_ids.add(family_id)
            elif automation_scope == "mapping_only":
                if family_id:
                    raise WorkflowError(
                        f"Mapping-only workbook must not bind a workflow family: "
                        f"{local_filename}"
                    )
                if not any(
                    re.fullmatch(pattern, local_filename)
                    for pattern in deferred_patterns
                ):
                    raise WorkflowError(
                        f"Mapping-only workbook is not deferred from automation: "
                        f"{local_filename}"
                    )
            else:
                raise WorkflowError(
                    f"Unsupported automation_scope for {local_filename}: "
                    f"{automation_scope}"
                )

        expected_family_ids = {
            str(family["id"])
            for family in families.values()
            if family.get("domain") == domain_id
            and os.path.normcase(
                os.path.normpath(
                    str(Path(str(family.get("target_workbook") or "")).parent)
                )
            )
            == normalized_root
        }
        if managed_family_ids != expected_family_ids:
            raise WorkflowError(
                f"Managed local temp-table mappings do not cover the workflow families "
                f"for {domain_id}: expected={sorted(expected_family_ids)}, "
                f"actual={sorted(managed_family_ids)}"
            )


def _validate_source_quality_config(
    family: dict[str, Any],
    quality: dict[str, Any],
) -> None:
    family_id = family["id"]
    max_age_hours = quality.get("max_age_hours")
    if not isinstance(max_age_hours, (int, float)) or isinstance(max_age_hours, bool) or max_age_hours <= 0:
        raise WorkflowError(f"Invalid max_age_hours for family {family_id}.")

    row_count = quality.get("row_count") or {}
    minimum = row_count.get("min")
    maximum = row_count.get("max")
    if (
        not isinstance(minimum, int)
        or isinstance(minimum, bool)
        or not isinstance(maximum, int)
        or isinstance(maximum, bool)
        or minimum < 1
        or maximum < minimum
    ):
        raise WorkflowError(f"Invalid row_count bounds for family {family_id}.")

    relative_change = quality.get("relative_change") or {}
    max_ratio = relative_change.get("max_ratio")
    if (
        not isinstance(max_ratio, (int, float))
        or isinstance(max_ratio, bool)
        or max_ratio < 0
    ):
        raise WorkflowError(f"Invalid relative_change.max_ratio for family {family_id}.")
    if relative_change.get("baseline") != "same_slice_or_latest_target":
        raise WorkflowError(f"Unsupported relative_change baseline for family {family_id}.")

    required_columns = quality.get("required_column_null_rate") or {}
    if not required_columns:
        raise WorkflowError(f"required_column_null_rate is empty for family {family_id}.")
    unknown_columns = sorted(set(required_columns) - set(family.get("target_columns", [])))
    if unknown_columns:
        raise WorkflowError(
            f"Unknown required-column null thresholds for family {family_id}: {unknown_columns}"
        )
    for column, threshold in required_columns.items():
        if (
            not isinstance(threshold, (int, float))
            or isinstance(threshold, bool)
            or not 0 <= threshold <= 1
        ):
            raise WorkflowError(
                f"Invalid null-rate threshold for {family_id}.{column}: {threshold}"
            )


def family_map(registry: dict[str, Any]) -> dict[str, dict[str, Any]]:
    return {family["id"]: family for family in registry["families"]}


def domain_map(registry: dict[str, Any]) -> dict[str, dict[str, Any]]:
    return dict(registry["domains"])


def family_domain(
    registry: dict[str, Any], family: dict[str, Any]
) -> dict[str, Any]:
    return domain_map(registry)[str(family["domain"])]


def source_chat_id(registry: dict[str, Any], family: dict[str, Any]) -> str:
    return str(
        family.get("source_chat_id")
        or family_domain(registry, family).get("chat", {}).get("expected_chat_id")
        or ""
    )


def source_chat_name(registry: dict[str, Any], family: dict[str, Any]) -> str:
    return str(
        family.get("source_chat_name")
        or family_domain(registry, family).get("chat", {}).get("name")
        or ""
    )


def _command_argv(executable: str, args: list[str]) -> list[str]:
    if Path(executable).suffix.casefold() in {".cmd", ".bat"}:
        raise WorkflowError(
            "Refusing to launch lark-cli through a Windows batch shim. "
            "Use the package-native lark-cli.exe so JSON content cannot be "
            "reinterpreted by cmd.exe."
        )
    return [executable, *args]


def run_json_command(
    executable: str,
    args: list[str],
    *,
    cwd: Path | None = None,
    timeout: int = 120,
) -> dict[str, Any]:
    env = os.environ.copy()
    env["PYTHONIOENCODING"] = "utf-8"
    env["PYTHONUTF8"] = "1"
    env["LARKSUITE_CLI_NO_UPDATE_NOTIFIER"] = "1"
    env["LARKSUITE_CLI_NO_SKILLS_NOTIFIER"] = "1"
    completed = subprocess.run(
        _command_argv(executable, args),
        cwd=str(cwd) if cwd else None,
        env=env,
        capture_output=True,
        text=True,
        encoding="utf-8",
        errors="replace",
        timeout=timeout,
        check=False,
    )
    output = completed.stdout.strip() or completed.stderr.strip()
    try:
        payload = json.loads(output)
    except json.JSONDecodeError as exc:
        raise WorkflowError(
            f"Command did not return JSON (exit={completed.returncode}): {' '.join(args[:3])}"
        ) from exc
    if completed.returncode != 0 or payload.get("ok") is False:
        error = payload.get("error") or {}
        message = error.get("message") or payload.get("message") or "command failed"
        raise WorkflowError(f"{message} (exit={completed.returncode})")
    return payload


def _windows_lark_cli_launchers() -> list[Path]:
    directories = [Path.cwd()]
    directories.extend(
        Path(value.strip('"'))
        for value in os.environ.get("PATH", "").split(os.pathsep)
        if value.strip('"')
    )
    launchers: list[Path] = []
    seen: set[str] = set()
    for directory in directories:
        for name in ("lark-cli.exe", "lark-cli.cmd", "lark-cli.bat", "lark-cli"):
            candidate = directory / name
            try:
                resolved = candidate.resolve()
            except OSError:
                continue
            key = str(resolved).casefold()
            if key in seen or not resolved.is_file():
                continue
            seen.add(key)
            launchers.append(resolved)
    return launchers


def _native_lark_cli_for_windows_launcher(launcher: Path) -> Path | None:
    if launcher.suffix.casefold() == ".exe":
        return launcher
    candidate = (
        launcher.parent
        / "node_modules"
        / "@larksuite"
        / "cli"
        / "bin"
        / "lark-cli.exe"
    )
    return candidate.resolve() if candidate.is_file() else None


def _native_lark_cli_package_version(native: Path) -> str | None:
    package_json = native.parent.parent / "package.json"
    try:
        payload = json.loads(package_json.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return None
    version = payload.get("version")
    return str(version).strip() if version else None


def resolve_lark_cli() -> str:
    if os.name == "nt":
        launchers = _windows_lark_cli_launchers()
        native_candidates: list[tuple[Path, str | None]] = []
        for launcher in launchers:
            native = _native_lark_cli_for_windows_launcher(launcher)
            if native is not None:
                native_candidates.append(
                    (native, _native_lark_cli_package_version(native))
                )
        versioned_candidates = [
            (native, version)
            for native, version in native_candidates
            if version
        ]
        versions = {version for _, version in versioned_candidates}
        if len(versions) > 1:
            details = ", ".join(
                f"{native}={version}"
                for native, version in versioned_candidates
            )
            raise WorkflowError(
                "lark-cli package version drift detected; refusing to choose "
                f"among native installations: {details}"
            )
        if native_candidates:
            return str(native_candidates[0][0])
        if launchers:
            raise WorkflowError(
                "lark-cli was found, but no package-native lark-cli.exe is "
                "available. Refusing the Windows .cmd/.bat shim because it can "
                "reinterpret JSON reply content containing <, >, |, or &."
            )
        raise WorkflowError("lark-cli is not available on PATH.")

    executable = shutil.which("lark-cli")
    if not executable:
        raise WorkflowError("lark-cli is not available on PATH.")
    return str(Path(executable).resolve())


def source_kind(family: dict[str, Any]) -> str:
    return str(family.get("source_kind") or "file_attachment")


def source_sender_id(registry: dict[str, Any], family: dict[str, Any]) -> str:
    return str(
        family.get("source_sender_open_id")
        or family_domain(registry, family).get("default_sender_open_id")
        or ""
    )


def source_sender_name(registry: dict[str, Any], family: dict[str, Any]) -> str:
    return str(
        family.get("source_sender_name")
        or family_domain(registry, family).get("default_sender_name")
        or ""
    )


def _extract_file_resource(content: str) -> tuple[str, str] | None:
    match = re.search(r'<file\s+key="([^"]+)"\s+name="([^"]+)"\s*/>', content)
    return (match.group(1), match.group(2)) if match else None


def _extract_docs_sheet_url(content: str) -> str | None:
    match = re.search(r"https://docs\.baijia\.com/sheet/[^\s<>\"']+", content)
    return match.group(0).rstrip(".,，。；;") if match else None


def normalize_source_message(message: dict[str, Any]) -> dict[str, Any] | None:
    content = str(message.get("content") or "")
    file_resource = _extract_file_resource(content)
    source_url = _extract_docs_sheet_url(content)
    if file_resource is None and source_url is None:
        return None
    sender = message.get("sender") or {}
    sender_id = sender.get("id") or sender.get("open_id") or message.get("sender_id")
    sender_name = sender.get("name") or message.get("sender_name")
    normalized = {
        "message_id": message.get("message_id"),
        "create_time": message.get("create_time"),
        "message_position": str(message.get("message_position") or ""),
        "message_app_link": message.get("message_app_link"),
        "message_type": message.get("msg_type") or message.get("message_type"),
        "content": content,
        "chat_id": message.get("chat_id"),
        "sender_id": sender_id,
        "sender_name": sender_name,
    }
    if file_resource is not None:
        normalized.update(
            {
                "source_kind": "file_attachment",
                "file_key": file_resource[0],
                "file_name": file_resource[1],
            }
        )
    else:
        normalized.update({"source_kind": "link_workbook", "source_url": source_url})
    return normalized


def message_matches_family(
    registry: dict[str, Any],
    family: dict[str, Any],
    message: dict[str, Any],
) -> bool:
    message_kind = message.get("source_kind")
    if not message_kind and message.get("file_name"):
        message_kind = "file_attachment"
    if message_kind != source_kind(family):
        return False
    expected_chat = source_chat_id(registry, family)
    if expected_chat and message.get("chat_id") != expected_chat:
        return False
    expected_sender = source_sender_id(registry, family)
    if expected_sender and message.get("sender_id") != expected_sender:
        return False
    if source_kind(family) == "file_attachment":
        filename = str(message.get("file_name") or "")
        return any(
            re.fullmatch(pattern, filename)
            for pattern in family.get("source_filename_patterns", [])
        )
    source_url = str(message.get("source_url") or "")
    content = str(message.get("content") or "")
    if not any(
        re.fullmatch(pattern, source_url)
        for pattern in family.get("source_url_patterns", [])
    ):
        return False
    return all(
        re.search(pattern, content)
        for pattern in family.get("source_message_patterns", [])
    )


def classify_source_message(registry: dict[str, Any], message: dict[str, Any]) -> str | None:
    normalized = normalize_source_message(message)
    if normalized is None:
        return None
    matches = [
        family["id"]
        for family in registry["families"]
        if message_matches_family(registry, family, normalized)
    ]
    if len(matches) > 1:
        raise WorkflowError(f"Source message matches multiple families: {matches}")
    return matches[0] if matches else None


def _source_search_profiles(
    registry: dict[str, Any],
    family_ids: list[str],
) -> list[dict[str, str]]:
    families = family_map(registry)
    profiles: dict[tuple[str, str, str, str], dict[str, str]] = {}
    for family_id in family_ids:
        family = families[family_id]
        sender_id = source_sender_id(registry, family)
        if not sender_id:
            raise WorkflowError(f"Source sender open_id is not configured for {family_id}.")
        chat_id = source_chat_id(registry, family)
        if not chat_id:
            raise WorkflowError(f"Source chat id is not configured for {family_id}.")
        kind = source_kind(family)
        query = str(family.get("source_search_query") or "")
        profiles[(chat_id, sender_id, kind, query)] = {
            "chat_id": chat_id,
            "chat_name": source_chat_name(registry, family),
            "sender_id": sender_id,
            "source_kind": kind,
            "query": query,
        }
    return list(profiles.values())


def list_chat_messages_bot(
    cli: str,
    chat_id: str,
    *,
    page_limit: int = 100,
) -> list[dict[str, Any]]:
    page_token = ""
    seen_tokens: set[str] = set()
    messages: list[dict[str, Any]] = []
    for _ in range(page_limit):
        command = [
            "im",
            "+chat-messages-list",
            "--chat-id",
            chat_id,
            "--order",
            "desc",
            "--page-size",
            "50",
            "--no-reactions",
            "--as",
            "bot",
            "--format",
            "json",
        ]
        if page_token:
            command.extend(["--page-token", page_token])
        result = run_json_command(
            cli,
            command,
            timeout=180,
        )
        data = result.get("data", {})
        messages.extend(data.get("messages", []))
        if not data.get("has_more"):
            return messages
        next_token = str(data.get("page_token") or "")
        if not next_token or next_token in seen_tokens:
            raise WorkflowError(
                "Bot chat-message pagination returned an invalid page token "
                f"for {chat_id}."
            )
        seen_tokens.add(next_token)
        page_token = next_token
    raise WorkflowError(
        f"Bot chat-message pagination exceeded {page_limit} pages for {chat_id}."
    )


def discover_live_messages(
    registry: dict[str, Any],
    explicit_message_ids: set[str] | None = None,
    family_ids: list[str] | None = None,
) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    cli = resolve_lark_cli()
    selected_family_ids = list(family_ids or registry["upload_order"])
    profiles = _source_search_profiles(registry, selected_family_ids)
    registered_chats = {
        profile["chat_id"]: {
            "chat_id": profile["chat_id"],
            "name": profile["chat_name"],
        }
        for profile in profiles
    }
    raw_messages: list[dict[str, Any]] = []
    for profile in profiles:
        command = [
            "im",
            "+messages-search",
            "--chat-id",
            profile["chat_id"],
            "--sender",
            profile["sender_id"],
        ]
        if profile["query"]:
            command.extend(["--query", profile["query"]])
        if profile["source_kind"] == "file_attachment":
            command.extend(["--include-attachment-type", "file"])
        command.extend(
            [
                "--page-size",
                "50",
                "--page-all",
                "--no-reactions",
                "--as",
                "user",
                "--format",
                "json",
            ]
        )
        try:
            result = run_json_command(cli, command, timeout=120)
            items = list(result.get("data", {}).get("messages", []))
            raw_messages.extend(items)
        except WorkflowError:
            items = []
        if not items:
            raw_messages.extend(
                list_chat_messages_bot(
                    cli,
                    profile["chat_id"],
                )
            )
    explicit_message_ids = explicit_message_ids or set()
    if explicit_message_ids:
        exact = run_json_command(
            cli,
            [
                "im",
                "+messages-mget",
                "--message-ids",
                ",".join(sorted(explicit_message_ids)),
                "--no-reactions",
                "--as",
                "bot",
                "--format",
                "json",
            ],
            timeout=60,
        )
        raw_messages.extend(exact.get("data", {}).get("messages", []))
    unique_messages = {
        str(message.get("message_id")): message
        for message in raw_messages
        if message.get("message_id")
    }
    normalized = []
    allowed_chat_ids = set(registered_chats)
    for message in unique_messages.values():
        item = normalize_source_message(message)
        if item is None or message.get("deleted"):
            continue
        if item.get("chat_id") not in allowed_chat_ids:
            continue
        normalized.append(item)
    return {"registered_chats": list(registered_chats.values())}, normalized


def classify_messages(
    registry: dict[str, Any], messages: list[dict[str, Any]]
) -> tuple[dict[str, list[dict[str, Any]]], list[dict[str, Any]]]:
    classified: dict[str, list[dict[str, Any]]] = defaultdict(list)
    deferred_patterns = [re.compile(pattern) for pattern in registry.get("deferred_filename_patterns", [])]
    unclassified = []
    for message in messages:
        matches = [
            family
            for family in registry["families"]
            if message_matches_family(registry, family, message)
        ]
        if len(matches) > 1:
            raise WorkflowError(
                f"Source message {message.get('message_id')} matches multiple families: "
                f"{[family['id'] for family in matches]}"
            )
        if matches:
            family = matches[0]
            item = dict(message)
            item["family_id"] = family["id"]
            item["file_name"] = item.get("file_name") or family.get("source_filename")
            item["sender_name"] = item.get("sender_name") or source_sender_name(registry, family)
            classified[family["id"]].append(item)
        else:
            item = dict(message)
            filename = str(message.get("file_name") or "")
            item["classification"] = (
                "deferred"
                if filename and any(pattern.fullmatch(filename) for pattern in deferred_patterns)
                else "excluded"
            )
            unclassified.append(item)
    return classified, unclassified


def message_sort_key(message: dict[str, Any]) -> tuple[str, int, str]:
    try:
        position = int(message.get("message_position") or 0)
    except (TypeError, ValueError):
        position = 0
    return str(message.get("create_time") or ""), position, str(message.get("message_id") or "")


def parse_datetime(value: str) -> datetime:
    text = str(value).strip()
    if not text:
        raise WorkflowError("Datetime value cannot be blank.")
    if re.fullmatch(r"\d+(?:\.\d+)?", text):
        timestamp = float(text)
        if timestamp > 10_000_000_000:
            timestamp /= 1000
        return datetime.fromtimestamp(timestamp, tz=timezone.utc)
    normalized = text[:-1] + "+00:00" if text.endswith("Z") else text
    try:
        parsed = datetime.fromisoformat(normalized)
    except ValueError as exc:
        raise WorkflowError(f"Invalid datetime value: {value}") from exc
    if parsed.tzinfo is None:
        parsed = parsed.astimezone()
    return parsed.astimezone(timezone.utc)


def build_selection_spec(
    registry: dict[str, Any],
    family_ids: list[str] | None = None,
    domains: list[str] | None = None,
    after: str | None = None,
    explicit_message_specs: list[str] | None = None,
) -> dict[str, Any]:
    known = family_map(registry)
    requested_domains = list(domains or [])
    unknown_domains = [
        domain_id for domain_id in requested_domains if domain_id not in domain_map(registry)
    ]
    if unknown_domains:
        raise WorkflowError(f"Unknown workflow domains: {unknown_domains}")
    if family_ids:
        requested = list(family_ids)
    elif requested_domains:
        requested = [
            family_id
            for family_id in registry["upload_order"]
            if known[family_id]["domain"] in set(requested_domains)
        ]
    else:
        requested = list(registry["upload_order"])
    if not requested:
        raise WorkflowError("At least one workbook family must be selected.")
    if len(requested) != len(set(requested)):
        raise WorkflowError(f"Workbook family selection contains duplicates: {requested}")
    unknown = [family_id for family_id in requested if family_id not in known]
    if unknown:
        raise WorkflowError(f"Unknown workbook family ids: {unknown}")
    outside_domains = [
        family_id
        for family_id in requested
        if requested_domains and known[family_id]["domain"] not in set(requested_domains)
    ]
    if outside_domains:
        raise WorkflowError(
            f"Selected families are outside --domain: {outside_domains}"
        )
    requested_set = set(requested)
    ordered = [family_id for family_id in registry["upload_order"] if family_id in requested_set]
    explicit: dict[str, str] = {}
    for raw_spec in explicit_message_specs or []:
        family_id, separator, message_id = raw_spec.partition("=")
        family_id = family_id.strip()
        message_id = message_id.strip()
        if not separator or family_id not in known or not re.fullmatch(r"om_[A-Za-z0-9]+", message_id):
            raise WorkflowError(
                f"Invalid --message-id value {raw_spec!r}; expected <family_id>=<om_message_id>."
            )
        if family_id not in requested_set:
            raise WorkflowError(f"Explicit message family is not selected by --family: {family_id}")
        if family_id in explicit:
            raise WorkflowError(f"Only one explicit message may be bound to family {family_id}.")
        explicit[family_id] = message_id
    after_iso = parse_datetime(after).isoformat() if after else None
    return {
        "family_ids": ordered,
        "domains": sorted({known[family_id]["domain"] for family_id in ordered}),
        "after": after_iso,
        "explicit_message_ids": explicit,
        "selection_modes": {
            family_id: "explicit_message" if family_id in explicit else "latest_matching"
            for family_id in ordered
        },
    }


def select_messages(
    registry: dict[str, Any],
    messages: list[dict[str, Any]],
    selection: dict[str, Any],
) -> tuple[dict[str, dict[str, Any]], dict[str, list[dict[str, Any]]], list[dict[str, Any]]]:
    after = parse_datetime(selection["after"]) if selection.get("after") else None
    eligible = []
    for message in messages:
        if after is not None:
            try:
                created = parse_datetime(str(message.get("create_time") or ""))
            except WorkflowError:
                continue
            if created <= after:
                continue
        eligible.append(message)
    classified, unclassified = classify_messages(registry, eligible)
    selected: dict[str, dict[str, Any]] = {}
    missing = []
    explicit = selection.get("explicit_message_ids") or {}
    for family_id in selection["family_ids"]:
        candidates = classified.get(family_id, [])
        if family_id in explicit:
            candidates = [message for message in candidates if message.get("message_id") == explicit[family_id]]
        if not candidates:
            missing.append(family_id)
        elif len(candidates) > 1 and family_id in explicit:
            raise WorkflowError(f"Explicit message id is not unique in search results: {explicit[family_id]}")
        else:
            selected[family_id] = max(candidates, key=message_sort_key)
    if missing:
        details = {family_id: explicit.get(family_id) for family_id in missing}
        raise WorkflowError(f"No matching source message found for selected families: {details}")
    return selected, classified, unclassified


def select_latest_messages(
    registry: dict[str, Any], messages: list[dict[str, Any]]
) -> tuple[dict[str, dict[str, Any]], dict[str, list[dict[str, Any]]], list[dict[str, Any]]]:
    selection = build_selection_spec(registry)
    return select_messages(registry, messages, selection)


def download_message(
    message: dict[str, Any],
    family: dict[str, Any],
    output_dir: Path,
) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    family_id = family["id"]
    output_name = f"{family_id}__{message['message_id']}.xlsx"
    if source_kind(family) == "link_workbook":
        env_key = str(family.get("source_env_file_env") or "")
        env_file_value = os.environ.get(env_key) if env_key else None
        env_file = Path(env_file_value or family["source_env_file"]).resolve()
        output_path = (output_dir / output_name).resolve()
        download_docs_sheet(
            url=str(message["source_url"]),
            output_path=output_path,
            env_file=env_file,
            credential_section=str(family.get("source_env_section") or "") or None,
            url_patterns=list(family["source_url_patterns"]),
            expected_title_pattern=str(family["source_expected_title_pattern"]),
            browser_channel=str(family.get("source_browser_channel") or "msedge"),
            timeout_seconds=int(family.get("source_download_timeout_seconds") or 120),
        )
        return output_path
    cli = resolve_lark_cli()
    payload = run_json_command(
        cli,
        [
            "im",
            "+messages-resources-download",
            "--message-id",
            message["message_id"],
            "--file-key",
            message["file_key"],
            "--type",
            "file",
            "--output",
            f".\\{output_name}",
            "--as",
            "bot",
        ],
        cwd=output_dir,
        timeout=120,
    )
    saved = Path(payload["data"]["saved_path"]).resolve()
    if saved.parent != output_dir.resolve() or not saved.exists():
        raise WorkflowError(f"Downloaded file escaped the plan directory: {saved}")
    return saved


def normalize_cell(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value


def transform_cell(value: Any, transforms: list[str]) -> Any:
    transformed = normalize_cell(value)
    for transform in transforms:
        if transform == "collapse_whitespace":
            if isinstance(transformed, str):
                transformed = re.sub(r"\s+", " ", transformed).strip()
        elif transform == "to_text":
            if transformed is not None:
                transformed = str(transformed).strip()
        elif transform == "normalize_datetime_text":
            if isinstance(transformed, datetime):
                transformed = transformed.strftime("%Y-%m-%d %H:%M:%S")
            elif transformed is not None:
                text = str(transformed).strip().replace("/", "-")
                try:
                    transformed = datetime.fromisoformat(text).strftime(
                        "%Y-%m-%d %H:%M:%S"
                    )
                except ValueError:
                    transformed = re.sub(r"\s+", " ", text).strip()
        else:
            raise WorkflowError(f"Unsupported source column transform: {transform}")
    return transformed


def normalized_record(record: dict[str, Any], columns: list[str]) -> tuple[Any, ...]:
    values = []
    for column in columns:
        value = normalize_cell(record.get(column))
        if isinstance(value, datetime):
            value = value.isoformat()
        values.append(value)
    return tuple(values)


def read_records(
    path: Path,
    sheet_name: str,
    target_columns: list[str],
    *,
    aliases: dict[str, str] | None = None,
    constants: dict[str, Any] | None = None,
    ignored_columns: list[str] | None = None,
    column_transforms: dict[str, list[str]] | None = None,
    data_only: bool,
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    aliases = aliases or {}
    constants = constants or {}
    ignored = set(ignored_columns or [])
    column_transforms = column_transforms or {}
    workbook = load_workbook(path, read_only=False, data_only=data_only, keep_links=True)
    try:
        requested_sheet = sheet_name
        if sheet_name == "$active":
            sheet_name = workbook.active.title
        if sheet_name not in workbook.sheetnames:
            raise WorkflowError(f"Sheet {sheet_name} not found in {path.name}; found {workbook.sheetnames}")
        sheet = workbook[sheet_name]
        raw_headers = [normalize_cell(cell.value) for cell in sheet[1]]
        while raw_headers and raw_headers[-1] in (None, ""):
            raw_headers.pop()
        headers = [aliases.get(str(header), str(header)) if header not in (None, "") else "" for header in raw_headers]
        nonblank_headers = [header for header in headers if header]
        if len(nonblank_headers) != len(set(nonblank_headers)):
            raise WorkflowError(f"Duplicate headers after alias mapping in {path.name}: {headers}")
        missing = [column for column in target_columns if column not in nonblank_headers and column not in constants]
        extras = [
            header
            for header in nonblank_headers
            if header not in target_columns and header not in ignored
        ]
        if missing or extras:
            raise WorkflowError(f"Schema mismatch in {path.name}: missing={missing}, extras={extras}, headers={headers}")
        positions = {header: index for index, header in enumerate(headers) if header}
        records = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not any(normalize_cell(value) not in (None, "") for value in row):
                continue
            record = {}
            for column in target_columns:
                if column in constants:
                    record[column] = constants[column]
                else:
                    index = positions[column]
                    value = row[index] if index < len(row) else None
                    record[column] = transform_cell(value, column_transforms.get(column, []))
            records.append(record)
        metadata = {
            "sheet": sheet_name,
            "requested_sheet": requested_sheet,
            "headers": headers,
            "ignored_columns": sorted(ignored),
            "column_transforms": column_transforms,
            "row_count": len(records),
            "formula_count": sum(
                1
                for row in sheet.iter_rows(min_row=2)
                for cell in row
                if cell.data_type == "f"
            ),
        }
        return records, metadata
    finally:
        workbook.close()


def _copy_records(records: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [dict(record) for record in records]


def _deduplicate_exact_records(
    records: list[dict[str, Any]], columns: list[str]
) -> tuple[list[dict[str, Any]], int]:
    seen: set[tuple[Any, ...]] = set()
    output = []
    removed = 0
    for record in records:
        key = normalized_record(record, columns)
        if key in seen:
            removed += 1
            continue
        seen.add(key)
        output.append(dict(record))
    return output, removed


def _source_scope_matches(record: dict[str, Any], family: dict[str, Any]) -> bool:
    scope = family.get("source_scope")
    if not scope:
        return True
    value = _text(record.get(scope["column"]))
    if "equals" in scope:
        return value == _text(scope["equals"])
    if "allowed_values" in scope:
        return value in {_text(item) for item in scope["allowed_values"]}
    raise WorkflowError(
        f"Unsupported source_scope for {family['id']}: {scope}"
    )


def _fill_from_reference_mapping(
    records: list[dict[str, Any]],
    reference_records: list[dict[str, Any]],
    transform: dict[str, Any],
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    column = str(transform["column"])
    key_columns = list(transform["key_columns"])
    mapping: dict[tuple[str, ...], set[str]] = defaultdict(set)
    for record in [*reference_records, *records]:
        value = _text(record.get(column))
        if not value:
            continue
        key = tuple(_text(record.get(key_column)) for key_column in key_columns)
        if all(key):
            mapping[key].add(value)
    ambiguous = {
        key: sorted(values)
        for key, values in mapping.items()
        if len(values) > 1
    }
    output = _copy_records(records)
    filled = 0
    unresolved = []
    for row_number, record in enumerate(output, start=2):
        if _text(record.get(column)):
            continue
        key = tuple(_text(record.get(key_column)) for key_column in key_columns)
        values = mapping.get(key, set())
        if len(values) == 1:
            record[column] = next(iter(values))
            filled += 1
        else:
            unresolved.append(
                {
                    "row": row_number,
                    "key": list(key),
                    "candidate_count": len(values),
                }
            )
    if ambiguous and transform.get("block_ambiguous_keys", True):
        used_ambiguous = [
            {
                "key": list(key),
                "values": values,
            }
            for key, values in ambiguous.items()
            if any(
                not _text(record.get(column))
                and tuple(_text(record.get(item)) for item in key_columns) == key
                for record in records
            )
        ]
        if used_ambiguous:
            raise WorkflowError(
                f"Ambiguous reference mapping for {column}: {used_ambiguous[:20]}"
            )
    if unresolved:
        raise WorkflowError(
            f"Cannot fill required source column {column}: {unresolved[:20]}"
        )
    return output, {
        "type": "fill_from_reference_mapping",
        "column": column,
        "key_columns": key_columns,
        "filled": filled,
    }


def _qici_date(value: Any) -> datetime:
    text = _text(value)
    if not re.fullmatch(r"\d{8}期", text):
        raise WorkflowError(f"Cannot order non-standard qici value: {text!r}")
    return datetime.strptime(text[:8], "%Y%m%d")


def _mode_value(values: Iterable[Any]) -> Any | None:
    counts = Counter(value for value in values if _text(value))
    if not counts:
        return None
    maximum = max(counts.values())
    return sorted(
        (value for value, count in counts.items() if count == maximum),
        key=lambda value: _text(value),
    )[0]


def _fill_market_evaluation_grade(
    records: list[dict[str, Any]],
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    output = _copy_records(records)
    filled = 0
    unresolved = []
    for row_number, record in enumerate(output, start=2):
        if _text(record.get("grade")):
            continue
        candidates = [
            candidate
            for candidate in output
            if _text(candidate.get("grade"))
            and _text(candidate.get("employee_email_name"))
            == _text(record.get("employee_email_name"))
            and _text(candidate.get("channel")) == _text(record.get("channel"))
        ]
        if candidates:
            candidates.sort(
                key=lambda candidate: (
                    abs(
                        (
                            _qici_date(candidate.get("qici"))
                            - _qici_date(record.get("qici"))
                        ).days
                    ),
                    _text(candidate.get("qici")),
                    _text(candidate.get("grade")),
                )
            )
            value = candidates[0].get("grade")
        else:
            value = _mode_value(
                candidate.get("grade")
                for candidate in output
                if _text(candidate.get("qici")) == _text(record.get("qici"))
                and _text(candidate.get("department"))
                == _text(record.get("department"))
                and _text(candidate.get("channel"))
                == _text(record.get("channel"))
            )
        if value is None:
            unresolved.append({"row": row_number})
            continue
        record["grade"] = value
        filled += 1
    if unresolved:
        raise WorkflowError(
            "Cannot fill market evaluation grade with registered evidence: "
            f"{unresolved[:20]}"
        )
    return output, {
        "type": "fill_market_evaluation_grade",
        "filled": filled,
    }


def _resequence_market_x_qi_count(
    records: list[dict[str, Any]],
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    output = _copy_records(records)
    original_is_nine = {
        id(record): _text(record.get("x_qi_count")) == "9"
        for record in output
    }
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for record in output:
        grouped[_text(record.get("employee_email_name"))].append(record)
    changes = 0
    for employee_records in grouped.values():
        eligible = [
            record
            for record in employee_records
            if not original_is_nine[id(record)]
        ]
        eligible.sort(
            key=lambda record: (
                _text(record.get("qici")),
                _text(record.get("channel")),
                _text(record.get("grade")),
            )
        )
        for index, record in enumerate(eligible, start=1):
            new_value = index if index <= 4 else 9
            if _text(record.get("x_qi_count")) != str(new_value):
                changes += 1
            record["x_qi_count"] = new_value
        for record in employee_records:
            if original_is_nine[id(record)]:
                record["x_qi_count"] = 9
    return output, {
        "type": "resequence_market_x_qi_count",
        "changed_cells": changes,
    }


def transform_source_records(
    family: dict[str, Any],
    source_records: list[dict[str, Any]],
    target_records: list[dict[str, Any]],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], dict[str, Any]]:
    scoped = [
        dict(record)
        for record in source_records
        if _source_scope_matches(record, family)
    ]
    snapshot_records = _copy_records(scoped)
    output = _copy_records(scoped)
    audit: list[dict[str, Any]] = []
    for transform in family.get("source_record_transforms", []):
        transform_type = transform["type"]
        if transform_type == "deduplicate_exact":
            before = len(output)
            output, removed = _deduplicate_exact_records(
                output, family["target_columns"]
            )
            audit.append(
                {
                    "type": transform_type,
                    "before": before,
                    "after": len(output),
                    "removed": removed,
                }
            )
        elif transform_type == "lowercase_columns":
            columns = list(transform["columns"])
            changed = 0
            for record in output:
                for column in columns:
                    value = record.get(column)
                    if isinstance(value, str) and value != value.lower():
                        record[column] = value.lower()
                        changed += 1
            audit.append(
                {
                    "type": transform_type,
                    "columns": columns,
                    "changed_cells": changed,
                }
            )
        elif transform_type == "fill_from_reference_mapping":
            output, details = _fill_from_reference_mapping(
                output, target_records, transform
            )
            audit.append(details)
        elif transform_type == "fill_market_evaluation_grade":
            output, details = _fill_market_evaluation_grade(output)
            audit.append(details)
        elif transform_type == "resequence_market_x_qi_count":
            output, details = _resequence_market_x_qi_count(output)
            audit.append(details)
        else:
            raise WorkflowError(
                f"Unsupported source record transform for {family['id']}: "
                f"{transform_type}"
            )
    return output, snapshot_records, {
        "source_rows_before_scope": len(source_records),
        "source_rows_after_scope": len(scoped),
        "transforms": audit,
    }


def source_slice_snapshot(
    family: dict[str, Any],
    records: list[dict[str, Any]],
) -> dict[str, Any]:
    slice_column = family["slice_column"]
    columns = family["target_columns"]
    grouped: dict[str, list[tuple[Any, ...]]] = defaultdict(list)
    for record in records:
        grouped[_text(record.get(slice_column))].append(
            normalized_record(record, columns)
        )
    slices = {}
    for slice_value, rows in sorted(
        grouped.items(), key=lambda item: slice_sort_key(item[0])
    ):
        rows.sort(key=repr)
        payload = json.dumps(
            rows,
            ensure_ascii=True,
            separators=(",", ":"),
            default=str,
        ).encode("utf-8")
        slices[slice_value] = {
            "row_count": len(rows),
            "sha256": hashlib.sha256(payload).hexdigest(),
        }
    return {"slices": slices, "row_count": len(records)}


def _load_source_baselines(
    seed_path: Path,
    state_path: Path,
) -> tuple[dict[str, Any], dict[str, Any]]:
    if not seed_path.is_file():
        raise WorkflowError(f"Source baseline seed does not exist: {seed_path}")
    seed = json.loads(seed_path.read_text(encoding="utf-8"))
    if not isinstance(seed.get("families"), dict):
        raise WorkflowError("Source baseline seed has no families object.")
    if state_path.is_file():
        state = json.loads(state_path.read_text(encoding="utf-8"))
        if not isinstance(state.get("families"), dict):
            raise WorkflowError("Source baseline state has no families object.")
        selected = state
        selected_kind = "runtime_state"
        selected_path = state_path
    else:
        selected = seed
        selected_kind = "seed"
        selected_path = seed_path
    return selected, {
        "kind": selected_kind,
        "path": str(selected_path.resolve()),
        "sha256": sha256_file(selected_path),
        "seed_path": str(seed_path.resolve()),
        "seed_sha256": sha256_file(seed_path),
        "state_path": str(state_path.resolve()),
        "state_sha256": sha256_file(state_path) if state_path.is_file() else None,
    }


def assert_source_baseline_state_current(plan: dict[str, Any]) -> None:
    context = plan.get("source_baseline_context") or {}
    state_path_value = context.get("state_path")
    if not state_path_value:
        return
    state_path = Path(state_path_value)
    current_sha256 = sha256_file(state_path) if state_path.is_file() else None
    if current_sha256 != context.get("state_sha256"):
        raise WorkflowError(
            "Source baseline state drifted after planning; create a fresh plan."
        )


def _write_json_atomic(path: Path, value: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_name(f".{path.name}.{os.getpid()}.tmp")
    temporary.write_text(
        json.dumps(value, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    os.replace(temporary, path)


def update_source_baseline_state(
    plan: dict[str, Any],
    uploaded_tables: list[dict[str, Any]],
) -> dict[str, Any] | None:
    eligible = [
        table
        for table in uploaded_tables
        if table.get("source_slice_snapshot")
        and table.get("source_selection", {}).get("mode")
        == "changed_source_slices"
    ]
    if not eligible:
        return None
    context = plan["source_baseline_context"]
    state_path = Path(context["state_path"])
    seed_path = Path(context["seed_path"])
    if state_path.is_file():
        state = json.loads(state_path.read_text(encoding="utf-8"))
    else:
        state = json.loads(seed_path.read_text(encoding="utf-8"))
    families = state.setdefault("families", {})
    for table in eligible:
        baseline_id = table["source_selection"]["baseline_id"]
        families[baseline_id] = {
            "message_id": table["source_message"]["message_id"],
            "create_time": table["source_message"]["create_time"],
            "source_sha256": table["source_message"]["source_sha256"],
            **table["source_slice_snapshot"],
        }
    state["schema_version"] = "1.0.0"
    state["updated_at"] = datetime.now().astimezone().isoformat(
        timespec="seconds"
    )
    _write_json_atomic(state_path, state)
    return {
        "path": str(state_path),
        "sha256": sha256_file(state_path),
        "updated_family_ids": [
            table["family_id"] for table in eligible
        ],
    }


def select_source_records_for_merge(
    family: dict[str, Any],
    transformed_records: list[dict[str, Any]],
    snapshot: dict[str, Any],
    baselines: dict[str, Any],
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    mode = family.get("source_merge_mode", "all_source_slices")
    current_slices = snapshot["slices"]
    if mode in {"all_source_slices", "full_source_replace"}:
        selected = sorted(current_slices, key=slice_sort_key)
        return _copy_records(transformed_records), {
            "mode": mode,
            "selected_slices": selected,
            "changed_slices": selected,
            "added_slices": [],
            "removed_slices": [],
        }
    baseline_id = family["source_baseline_id"]
    baseline = baselines.get("families", {}).get(baseline_id)
    if not baseline or not isinstance(baseline.get("slices"), dict):
        raise WorkflowError(
            f"Source slice baseline is missing for {family['id']}: {baseline_id}"
        )
    baseline_slices = baseline["slices"]
    removed = sorted(
        set(baseline_slices) - set(current_slices), key=slice_sort_key
    )
    if removed:
        raise WorkflowError(
            f"Cumulative source removed registered slices for {family['id']}: {removed}"
        )
    added = sorted(
        set(current_slices) - set(baseline_slices), key=slice_sort_key
    )
    changed = sorted(
        (
            slice_value
            for slice_value in set(current_slices) & set(baseline_slices)
            if current_slices[slice_value].get("sha256")
            != baseline_slices[slice_value].get("sha256")
        ),
        key=slice_sort_key,
    )
    bootstrap = [
        value
        for value in family.get("bootstrap_slices", [])
        if baseline.get("bootstrap_pending") is True
        and value in current_slices
    ]
    selected = sorted(set(added + changed + bootstrap), key=slice_sort_key)
    maximum = int(family["max_changed_slices"])
    if len(selected) > maximum:
        raise WorkflowError(
            f"Source changed too many slices for {family['id']}: "
            f"{len(selected)} > {maximum}; slices={selected}"
        )
    window = int(family.get("recent_slice_window") or 4)
    recent = set(sorted(current_slices, key=slice_sort_key)[-window:])
    reviewed_historical = set(family.get("reviewed_historical_slices", []))
    unreviewed_historical = [
        value
        for value in selected
        if value not in recent
        and value not in reviewed_historical
    ]
    if unreviewed_historical:
        raise WorkflowError(
            f"Source changed unreviewed historical slices for {family['id']}: "
            f"{unreviewed_historical}"
        )
    selected_set = set(selected)
    records = [
        dict(record)
        for record in transformed_records
        if _text(record.get(family["slice_column"])) in selected_set
    ]
    return records, {
        "mode": mode,
        "baseline_id": baseline_id,
        "baseline_message_id": baseline.get("message_id"),
        "selected_slices": selected,
        "changed_slices": changed,
        "added_slices": added,
        "bootstrap_slices": sorted(set(bootstrap), key=slice_sort_key),
        "removed_slices": removed,
        "recent_slice_window": window,
    }


def _text(value: Any) -> str:
    value = normalize_cell(value)
    return "" if value is None else str(value).strip()


def validate_source_records(
    family: dict[str, Any],
    records: list[dict[str, Any]],
    source_metadata: dict[str, Any] | None = None,
) -> list[dict[str, Any]]:
    issues: list[dict[str, Any]] = []
    if not records:
        return [{"severity": "error", "rule": "nonempty_source", "message": "Source workbook has no data rows."}]
    key_columns = family["key_columns"]
    groups: dict[tuple[str, ...], list[int]] = defaultdict(list)
    for row_number, record in enumerate(records, start=2):
        key = tuple(_text(record.get(column)) for column in key_columns)
        if any(not value for value in key):
            issues.append(
                {
                    "severity": "error",
                    "rule": "nonblank_key",
                    "message": f"Blank key value at row {row_number} for {key_columns}.",
                    "row": row_number,
                }
            )
        groups[key].append(row_number)
    duplicates = [{"key": key, "rows": rows} for key, rows in groups.items() if all(key) and len(rows) > 1]
    if duplicates:
        issues.append(
            {
                "severity": "error",
                "rule": "unique_key",
                "message": f"Source has duplicate keys for {key_columns}.",
                "count": len(duplicates),
                "examples": duplicates[:20],
            }
        )
    for rule in family.get("validation_rules", []):
        rule_type = rule["type"]
        if rule_type == "slice_format":
            invalid = [
                {"row": index, "value": _text(record.get(rule["column"]))}
                for index, record in enumerate(records, start=2)
                if not re.fullmatch(rule["pattern"], _text(record.get(rule["column"])))
            ]
        elif rule_type == "month_format":
            invalid = [
                {"row": index, "value": _text(record.get(rule["column"]))}
                for index, record in enumerate(records, start=2)
                if not re.fullmatch(r"\d{6}", _text(record.get(rule["column"])))
            ]
        elif rule_type == "lowercase_ascii_prefix":
            invalid = []
            for index, record in enumerate(records, start=2):
                for column in rule["columns"]:
                    value = _text(record.get(column))
                    first = next((char for char in value if char.isascii() and char.isalpha()), "")
                    if first and first != first.lower():
                        invalid.append({"row": index, "column": column, "value": value})
        elif rule_type == "disallow_values":
            disallowed = {_text(value) for value in rule["values"]}
            invalid = [
                {"row": index, "value": _text(record.get(rule["column"]))}
                for index, record in enumerate(records, start=2)
                if _text(record.get(rule["column"])) in disallowed
            ]
        elif rule_type == "required_value":
            invalid = [
                {"row": index, "value": _text(record.get(rule["column"]))}
                for index, record in enumerate(records, start=2)
                if _text(record.get(rule["column"])) != _text(rule["value"])
            ]
        elif rule_type == "required_nonblank":
            invalid = []
            for index, record in enumerate(records, start=2):
                for column in rule["columns"]:
                    if not _text(record.get(column)):
                        invalid.append({"row": index, "column": column})
        elif rule_type == "allowed_values":
            allowed = {_text(value) for value in rule["values"]}
            invalid = [
                {"row": index, "value": _text(record.get(rule["column"]))}
                for index, record in enumerate(records, start=2)
                if _text(record.get(rule["column"])) not in allowed
            ]
        elif rule_type == "datetime_format":
            invalid = []
            for index, record in enumerate(records, start=2):
                value = record.get(rule["column"])
                if isinstance(value, datetime):
                    continue
                try:
                    datetime.strptime(_text(value), rule["format"])
                except ValueError:
                    invalid.append({"row": index, "value": _text(value)})
        elif rule_type == "sheet_slice_suffix":
            sheet = str((source_metadata or {}).get("sheet") or "")
            invalid = [
                {"row": index, "value": _text(record.get(rule["column"])), "sheet": sheet}
                for index, record in enumerate(records, start=2)
                if not sheet or not _text(record.get(rule["column"])).endswith(sheet)
            ]
        elif rule_type == "formula_count_max":
            formula_count = int((source_metadata or {}).get("formula_count") or 0)
            invalid = (
                [{"formula_count": formula_count, "max": int(rule["max"])}]
                if formula_count > int(rule["max"])
                else []
            )
        elif rule_type == "numeric_nonnegative":
            invalid = []
            for index, record in enumerate(records, start=2):
                value = record.get(rule["column"])
                try:
                    numeric = float(value)
                except (TypeError, ValueError):
                    invalid.append({"row": index, "value": _text(value)})
                    continue
                if numeric < 0:
                    invalid.append({"row": index, "value": _text(value)})
        elif rule_type == "group_name_qici_prefix":
            invalid = []
            prefix_pattern = re.compile(r"^(?P<qici>\d{4}期)-")
            for index, record in enumerate(records, start=2):
                group_name = _text(record.get(rule["name_column"]))
                matched = prefix_pattern.match(group_name)
                if (
                    matched
                    and matched.group("qici")
                    != _text(record.get(rule["qici_column"]))
                ):
                    invalid.append(
                        {
                            "row": index,
                            "qici": _text(
                                record.get(rule["qici_column"])
                            ),
                            "group_name": group_name,
                        }
                    )
        elif rule_type == "market_x_qi_sequence":
            allowed = {"1", "2", "3", "4", "9"}
            invalid = []
            seen_active: dict[tuple[str, str], list[int]] = defaultdict(list)
            for index, record in enumerate(records, start=2):
                employee = _text(record.get("employee_email_name"))
                value = _text(record.get("x_qi_count"))
                if value not in allowed:
                    invalid.append(
                        {"row": index, "value": value, "reason": "invalid_value"}
                    )
                elif value != "9":
                    seen_active[(employee, value)].append(index)
            invalid.extend(
                {
                    "employee_email_name": employee,
                    "x_qi_count": value,
                    "rows": rows,
                    "reason": "duplicate_active_sequence",
                }
                for (employee, value), rows in seen_active.items()
                if len(rows) > 1
            )
        else:
            invalid = [{"rule": rule_type}]
        if invalid:
            issues.append(
                {
                    "severity": "error",
                    "rule": rule_type,
                    "message": f"Source validation failed: {rule_type}.",
                    "count": len(invalid),
                    "examples": invalid[:20],
                }
            )
    return issues


def slice_sort_key(value: Any) -> tuple[int, str]:
    text = _text(value)
    digits = re.sub(r"\D", "", text)
    return (int(digits) if digits else -1, text)


def scope_matches(record: dict[str, Any], family: dict[str, Any]) -> bool:
    scope = family.get("target_scope")
    if not scope:
        return True
    return _text(record.get(scope["column"])) == _text(scope["equals"])


def evaluate_source_quality(
    family: dict[str, Any],
    message: dict[str, Any],
    source_records: list[dict[str, Any]],
    target_records: list[dict[str, Any]],
    *,
    relative_source_records: list[dict[str, Any]] | None = None,
    now: datetime | None = None,
) -> dict[str, Any]:
    quality = family.get("source_quality")
    evaluated_at = (now or datetime.now(timezone.utc)).astimezone(timezone.utc)
    if not quality:
        return {
            "policy_version": None,
            "status": "not_configured",
            "evaluated_at": evaluated_at.isoformat(timespec="seconds"),
            "issues": [],
        }

    issues: list[dict[str, Any]] = []
    source_created_at: datetime | None = None
    source_expires_at: datetime | None = None
    source_age_hours: float | None = None
    try:
        source_created_at = parse_datetime(str(message.get("create_time") or ""))
        source_age_hours = (evaluated_at - source_created_at).total_seconds() / 3600
        source_expires_at = source_created_at + timedelta(hours=float(quality["max_age_hours"]))
        if source_age_hours < -1:
            issues.append(
                {
                    "severity": "error",
                    "rule": "source_time_in_future",
                    "message": "Source message time is more than one hour in the future.",
                    "source_age_hours": round(source_age_hours, 3),
                }
            )
        if evaluated_at > source_expires_at:
            issues.append(
                {
                    "severity": "error",
                    "rule": "source_max_age",
                    "message": "Source message exceeds the configured maximum age.",
                    "source_age_hours": round(source_age_hours, 3),
                    "max_age_hours": quality["max_age_hours"],
                }
            )
    except WorkflowError as exc:
        issues.append(
            {
                "severity": "error",
                "rule": "source_time_parse",
                "message": str(exc),
            }
        )

    row_count = len(source_records)
    row_bounds = quality["row_count"]
    if not row_bounds["min"] <= row_count <= row_bounds["max"]:
        issues.append(
            {
                "severity": "error",
                "rule": "source_row_count",
                "message": "Source row count is outside the configured bounds.",
                "row_count": row_count,
                "min": row_bounds["min"],
                "max": row_bounds["max"],
            }
        )

    required_column_results = []
    for column, threshold in quality["required_column_null_rate"].items():
        null_count = sum(1 for record in source_records if not _text(record.get(column)))
        null_rate = null_count / row_count if row_count else 1.0
        result = {
            "column": column,
            "null_count": null_count,
            "row_count": row_count,
            "null_rate": round(null_rate, 6),
            "max_null_rate": threshold,
            "ok": null_rate <= threshold,
        }
        required_column_results.append(result)
        if not result["ok"]:
            issues.append(
                {
                    "severity": "error",
                    "rule": "required_column_null_rate",
                    "message": f"Required column {column} exceeds its null-rate threshold.",
                    **result,
                }
            )

    slice_column = family["slice_column"]
    relative_records = (
        source_records
        if relative_source_records is None
        else relative_source_records
    )
    source_by_slice: dict[str, int] = Counter(
        _text(record.get(slice_column)) for record in relative_records
    )
    scoped_target_records = [
        record for record in target_records if scope_matches(record, family)
    ]
    target_by_slice: dict[str, int] = Counter(
        _text(record.get(slice_column)) for record in scoped_target_records
    )
    target_slices = sorted(target_by_slice, key=slice_sort_key)
    latest_target_slice = target_slices[-1] if target_slices else None
    relative_results = []
    max_ratio = float(quality["relative_change"]["max_ratio"])
    for slice_value in sorted(source_by_slice, key=slice_sort_key):
        source_count = source_by_slice[slice_value]
        if slice_value in target_by_slice:
            baseline_slice = slice_value
            baseline_kind = "same_slice"
        else:
            baseline_slice = latest_target_slice
            baseline_kind = "latest_target_slice"
        baseline_count = target_by_slice.get(baseline_slice, 0) if baseline_slice else 0
        if baseline_count <= 0:
            result = {
                "slice": slice_value,
                "source_count": source_count,
                "baseline_slice": baseline_slice,
                "baseline_count": baseline_count,
                "baseline_kind": baseline_kind,
                "relative_change": None,
                "max_ratio": max_ratio,
                "ok": False,
            }
            relative_results.append(result)
            issues.append(
                {
                    "severity": "error",
                    "rule": "relative_change_baseline_missing",
                    "message": f"No non-empty target baseline is available for source slice {slice_value}.",
                    **result,
                }
            )
            continue
        relative_change = abs(source_count - baseline_count) / baseline_count
        result = {
            "slice": slice_value,
            "source_count": source_count,
            "baseline_slice": baseline_slice,
            "baseline_count": baseline_count,
            "baseline_kind": baseline_kind,
            "relative_change": round(relative_change, 6),
            "max_ratio": max_ratio,
            "ok": relative_change <= max_ratio,
        }
        relative_results.append(result)
        if not result["ok"]:
            issues.append(
                {
                    "severity": "error",
                    "rule": "source_relative_change",
                    "message": f"Source slice {slice_value} exceeds the configured relative-change threshold.",
                    **result,
                }
            )

    return {
        "policy_version": str(quality.get("policy_version") or "1.0.0"),
        "status": "blocked" if issues else "pass",
        "evaluated_at": evaluated_at.isoformat(timespec="seconds"),
        "source_created_at": (
            source_created_at.isoformat(timespec="seconds") if source_created_at else None
        ),
        "source_expires_at": (
            source_expires_at.isoformat(timespec="seconds") if source_expires_at else None
        ),
        "source_age_hours": (
            round(source_age_hours, 3) if source_age_hours is not None else None
        ),
        "max_age_hours": quality["max_age_hours"],
        "row_count": {
            "actual": row_count,
            "min": row_bounds["min"],
            "max": row_bounds["max"],
            "ok": row_bounds["min"] <= row_count <= row_bounds["max"],
        },
        "required_columns": required_column_results,
        "relative_change": {
            "baseline": quality["relative_change"]["baseline"],
            "max_ratio": max_ratio,
            "slices": relative_results,
        },
        "issues": issues,
    }


def assert_plan_source_quality_current(
    plan: dict[str, Any],
    registry: dict[str, Any],
    *,
    now: datetime | None = None,
) -> None:
    if not registry.get("require_source_quality_gates"):
        return
    current_time = (now or datetime.now(timezone.utc)).astimezone(timezone.utc)
    tables_by_id = {table.get("family_id"): table for table in plan.get("tables", [])}
    for family_id in (plan.get("selection") or {}).get("family_ids", []):
        table = tables_by_id.get(family_id)
        report = (table or {}).get("source_quality") or {}
        if report.get("status") != "pass":
            raise WorkflowError(
                f"Source quality gate is not passing for {family_id}; create a fresh plan."
            )
        expires_at = report.get("source_expires_at")
        if not expires_at or current_time > parse_datetime(str(expires_at)):
            raise WorkflowError(
                f"Source quality gate expired for {family_id}; create a fresh plan."
            )


def records_equal(left: list[dict[str, Any]], right: list[dict[str, Any]], columns: list[str]) -> bool:
    return [normalized_record(record, columns) for record in left] == [
        normalized_record(record, columns) for record in right
    ]


def merge_records(
    family: dict[str, Any],
    target_write: list[dict[str, Any]],
    target_effective: list[dict[str, Any]],
    source_write: list[dict[str, Any]],
    source_effective: list[dict[str, Any]],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], dict[str, Any]]:
    if len(target_write) != len(target_effective) or len(source_write) != len(source_effective):
        raise WorkflowError("Formula and effective record streams are not aligned.")
    slice_column = family["slice_column"]
    columns = family["target_columns"]
    if (
        not source_effective
        and family.get("source_merge_mode") != "full_source_replace"
    ):
        return _copy_records(target_write), _copy_records(target_effective), {
            "changed": False,
            "slice_column": slice_column,
            "source_slices": [],
            "new_slices": [],
            "replaced_slices": [],
            "unchanged_slices": [],
            "deleted_slices": [],
            "target_rows_before": len(target_effective),
            "source_rows": 0,
            "scoped_rows_removed": 0,
            "target_rows_after": len(target_effective),
        }
    target_by_slice: dict[str, list[tuple[dict[str, Any], dict[str, Any]]]] = defaultdict(list)
    source_by_slice: dict[str, list[tuple[dict[str, Any], dict[str, Any]]]] = defaultdict(list)
    for write_record, effective_record in zip(target_write, target_effective):
        target_by_slice[_text(effective_record.get(slice_column))].append((write_record, effective_record))
    for write_record, effective_record in zip(source_write, source_effective):
        source_by_slice[_text(effective_record.get(slice_column))].append((write_record, effective_record))
    source_slices = set(source_by_slice)
    target_slices = set(target_by_slice)
    all_slices = sorted(
        target_slices | source_slices,
        key=slice_sort_key,
        reverse=family.get("slice_order") == "desc",
    )
    merged_write = []
    merged_effective = []
    new_slices = []
    replaced_slices = []
    unchanged_slices = []
    deleted_slices = []
    removed_rows = 0
    replace_full_scope = (
        family.get("source_merge_mode") == "full_source_replace"
    )
    for slice_value in all_slices:
        target_pairs = target_by_slice.get(slice_value, [])
        source_pairs = source_by_slice.get(slice_value)
        if source_pairs is None:
            if replace_full_scope:
                scoped_target = [
                    pair for pair in target_pairs if scope_matches(pair[1], family)
                ]
                output_pairs = [
                    pair
                    for pair in target_pairs
                    if not scope_matches(pair[1], family)
                ]
                removed_rows += len(scoped_target)
                if scoped_target:
                    deleted_slices.append(slice_value)
            else:
                output_pairs = target_pairs
        else:
            scoped_target = [pair for pair in target_pairs if scope_matches(pair[1], family)]
            preserved_target = [pair for pair in target_pairs if not scope_matches(pair[1], family)]
            removed_rows += len(scoped_target)
            source_scope_effective = [pair[1] for pair in source_pairs]
            target_scope_effective = [pair[1] for pair in scoped_target]
            if slice_value not in target_slices:
                new_slices.append(slice_value)
            elif Counter(normalized_record(record, columns) for record in source_scope_effective) == Counter(
                normalized_record(record, columns) for record in target_scope_effective
            ):
                unchanged_slices.append(slice_value)
            else:
                replaced_slices.append(slice_value)
            output_pairs = [*source_pairs, *preserved_target]
        merged_write.extend(pair[0] for pair in output_pairs)
        merged_effective.extend(pair[1] for pair in output_pairs)
    changed = not records_equal(merged_effective, target_effective, columns)
    diff = {
        "changed": changed,
        "slice_column": slice_column,
        "source_slices": sorted(source_slices, key=slice_sort_key),
        "new_slices": new_slices,
        "replaced_slices": replaced_slices,
        "unchanged_slices": unchanged_slices,
        "deleted_slices": deleted_slices,
        "target_rows_before": len(target_effective),
        "source_rows": len(source_effective),
        "scoped_rows_removed": removed_rows,
        "target_rows_after": len(merged_effective),
    }
    return merged_write, merged_effective, diff


def count_formula_values(records: Iterable[dict[str, Any]]) -> int:
    return sum(isinstance(value, str) and value.startswith("=") for record in records for value in record.values())


def inspect_external_link_integrity(path: Path) -> dict[str, Any]:
    with zipfile.ZipFile(path) as archive:
        names = set(archive.namelist())
        external_parts = sorted(
            name for name in names if name.startswith(EXTERNAL_LINK_PART_PREFIX)
        )
        link_xml_parts = sorted(
            name
            for name in external_parts
            if "/_rels/" not in name and name.endswith(".xml")
        )
        workbook_relationships = ElementTree.fromstring(
            archive.read("xl/_rels/workbook.xml.rels")
        )
        workbook_relations = {
            relation.attrib.get("Id"): relation
            for relation in workbook_relationships.findall(
                f"{{{PACKAGE_REL_NS}}}Relationship"
            )
        }
        workbook = ElementTree.fromstring(archive.read("xl/workbook.xml"))
        workbook_targets = []
        unresolved_workbook_relationship_ids = []
        missing_workbook_targets = []
        for reference in workbook.findall(
            f".//{{{SPREADSHEET_MAIN_NS}}}externalReference"
        ):
            relationship_id = reference.attrib.get(f"{{{OFFICE_REL_NS}}}id")
            relation = workbook_relations.get(relationship_id)
            if (
                relation is None
                or relation.attrib.get("Type") != EXTERNAL_LINK_REL_TYPE
            ):
                unresolved_workbook_relationship_ids.append(relationship_id)
                continue
            raw_target = str(relation.attrib.get("Target") or "").replace(
                "\\", "/"
            )
            target = posixpath.normpath(
                raw_target
                if raw_target.startswith("/")
                else posixpath.join("xl", raw_target)
            ).lstrip("/")
            workbook_targets.append(
                {"relationship_id": relationship_id, "target": target}
            )
            if target not in names:
                missing_workbook_targets.append(target)

        unresolved_external_relationships = []
        for part in link_xml_parts:
            external_link = ElementTree.fromstring(archive.read(part))
            referenced_ids = {
                value
                for element in external_link.iter()
                for key, value in element.attrib.items()
                if key == f"{{{OFFICE_REL_NS}}}id"
            }
            rels_path = (
                f"{posixpath.dirname(part)}/_rels/{posixpath.basename(part)}.rels"
            )
            defined_ids: set[str] = set()
            if rels_path in names:
                relations = ElementTree.fromstring(archive.read(rels_path))
                defined_ids = {
                    relation.attrib.get("Id")
                    for relation in relations.findall(
                        f"{{{PACKAGE_REL_NS}}}Relationship"
                    )
                    if relation.attrib.get("Id")
                }
            missing_ids = sorted(referenced_ids - defined_ids)
            if missing_ids:
                unresolved_external_relationships.append(
                    {"part": part, "relationship_ids": missing_ids}
                )

    result = {
        "external_link_count": len(link_xml_parts),
        "external_link_parts": external_parts,
        "workbook_targets": workbook_targets,
        "unresolved_workbook_relationship_ids": sorted(
            value
            for value in unresolved_workbook_relationship_ids
            if value is not None
        ),
        "missing_workbook_targets": sorted(set(missing_workbook_targets)),
        "unresolved_external_relationships": unresolved_external_relationships,
    }
    result["ok"] = not (
        result["unresolved_workbook_relationship_ids"]
        or result["missing_workbook_targets"]
        or result["unresolved_external_relationships"]
    )
    return result


def preserve_external_link_parts(source_path: Path, stage_path: Path) -> dict[str, Any]:
    source_integrity = inspect_external_link_integrity(source_path)
    if not source_integrity["ok"]:
        raise WorkflowError(
            f"Source workbook external links are invalid: {source_integrity}"
        )
    if source_integrity["external_link_count"] == 0:
        return {
            "restored": False,
            "reason": "source_has_no_external_links",
            "source_integrity": source_integrity,
            "stage_integrity": inspect_external_link_integrity(stage_path),
        }

    stage_before = inspect_external_link_integrity(stage_path)
    source_targets = sorted(
        item["target"] for item in source_integrity["workbook_targets"]
    )
    stage_targets = sorted(item["target"] for item in stage_before["workbook_targets"])
    if stage_targets != source_targets:
        raise WorkflowError(
            "Staged workbook external-link targets changed before preservation: "
            f"source={source_targets}, staged={stage_targets}"
        )

    temporary_path = stage_path.with_name(
        f"{stage_path.name}.{os.getpid()}.external-links.tmp"
    )
    try:
        with (
            zipfile.ZipFile(source_path) as source_archive,
            zipfile.ZipFile(stage_path) as stage_archive,
            zipfile.ZipFile(temporary_path, "w") as output_archive,
        ):
            source_parts = {
                info.filename: (info, source_archive.read(info.filename))
                for info in source_archive.infolist()
                if info.filename.startswith(EXTERNAL_LINK_PART_PREFIX)
            }
            for info in stage_archive.infolist():
                if info.filename.startswith(EXTERNAL_LINK_PART_PREFIX):
                    continue
                output_archive.writestr(info, stage_archive.read(info.filename))
            for name in sorted(source_parts):
                info, payload = source_parts[name]
                output_archive.writestr(info, payload)
        replace_file_with_retry(temporary_path, stage_path)
    finally:
        temporary_path.unlink(missing_ok=True)

    stage_after = inspect_external_link_integrity(stage_path)
    if not stage_after["ok"]:
        raise WorkflowError(
            f"Staged workbook external links remain invalid: {stage_after}"
        )
    with zipfile.ZipFile(source_path) as source_archive, zipfile.ZipFile(
        stage_path
    ) as stage_archive:
        source_hashes = {
            name: hashlib.sha256(source_archive.read(name)).hexdigest()
            for name in source_integrity["external_link_parts"]
        }
        stage_hashes = {
            name: hashlib.sha256(stage_archive.read(name)).hexdigest()
            for name in stage_after["external_link_parts"]
        }
    if stage_hashes != source_hashes:
        raise WorkflowError("Staged workbook external-link parts were not preserved byte-for-byte.")
    return {
        "restored": True,
        "source_integrity": source_integrity,
        "stage_before_integrity": stage_before,
        "stage_after_integrity": stage_after,
        "part_sha256": stage_hashes,
    }


def rebuild_workbook(
    target_path: Path,
    stage_path: Path,
    family: dict[str, Any],
    records: list[dict[str, Any]],
) -> dict[str, Any]:
    shutil.copy2(target_path, stage_path)
    workbook = load_workbook(stage_path, data_only=False, keep_links=True)
    sheet = workbook[family["target_sheet"]]
    columns = family["target_columns"]
    template_cells = [copy.copy(sheet.cell(row=2, column=index)._style) for index in range(1, len(columns) + 1)]
    template_height = sheet.row_dimensions[2].height
    existing_filter = sheet.auto_filter.ref
    if sheet.max_row > 1:
        sheet.delete_rows(2, sheet.max_row - 1)
    for record in records:
        sheet.append([record.get(column) for column in columns])
        row_number = sheet.max_row
        if template_height is not None:
            sheet.row_dimensions[row_number].height = template_height
        for index, style in enumerate(template_cells, start=1):
            sheet.cell(row=row_number, column=index)._style = copy.copy(style)
    if existing_filter:
        sheet.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{max(sheet.max_row, 1)}"
    calculation = getattr(workbook, "calculation", None)
    if calculation is not None:
        calculation.fullCalcOnLoad = True
        calculation.forceFullCalc = True
        calculation.calcMode = "auto"
    workbook.save(stage_path)
    workbook.close()
    external_link_preservation = None
    if family.get("preserve_external_links"):
        external_link_preservation = preserve_external_link_parts(
            target_path, stage_path
        )
    formula_count = count_formula_values(records)
    recalc = None
    if formula_count:
        recalc = recalculate_workbook(stage_path)
    external_link_integrity = None
    if family.get("preserve_external_links"):
        external_link_integrity = inspect_external_link_integrity(stage_path)
        if not external_link_integrity["ok"]:
            raise WorkflowError(
                "Excel recalculation left invalid external links: "
                f"{external_link_integrity}"
            )
    return {
        "formula_count": formula_count,
        "recalculation": recalc,
        "external_link_preservation": external_link_preservation,
        "external_link_integrity": external_link_integrity,
    }


def recalculate_workbook(path: Path) -> dict[str, Any]:
    if not RECALC_SCRIPT.exists():
        raise WorkflowError(f"Spreadsheet recalculation script not found: {RECALC_SCRIPT}")
    payload = run_json_command(sys.executable, [str(RECALC_SCRIPT), str(path), "60"], timeout=120)
    if payload.get("status") not in {"success", "ok"} or payload.get("total_errors", 0):
        raise WorkflowError(f"Excel recalculation failed for {path}: {payload}")
    return payload


def operator_validation(file_path: Path, target_path: Path) -> dict[str, Any]:
    if str(OPERATOR_SCRIPTS) not in sys.path:
        sys.path.insert(0, str(OPERATOR_SCRIPTS))
    from usql_web_query.manual_table_registry import ManualTableRegistry  # noqa: PLC0415
    from usql_web_query.manual_table_validation import validate_manual_table  # noqa: PLC0415

    registry = ManualTableRegistry.load()
    entry = registry.resolve_file(target_path)
    if entry is None:
        raise WorkflowError(f"Operator manual-table registry does not match target: {target_path}")
    return validate_manual_table(file_path, entry)


def validation_regressions(before: dict[str, Any], after: dict[str, Any]) -> list[dict[str, Any]]:
    baseline = {}
    for issue in before.get("issues", []):
        if issue.get("severity") != "error":
            continue
        signature = (issue.get("rule"), issue.get("column"), issue.get("message"))
        baseline[signature] = int(issue.get("count", 1))
    regressions = []
    for issue in after.get("issues", []):
        if issue.get("severity") != "error":
            continue
        signature = (issue.get("rule"), issue.get("column"), issue.get("message"))
        after_count = int(issue.get("count", 1))
        if after_count > baseline.get(signature, 0):
            regressions.append(
                {
                    "rule": issue.get("rule"),
                    "column": issue.get("column"),
                    "before_count": baseline.get(signature, 0),
                    "after_count": after_count,
                    "message": issue.get("message"),
                }
            )
    return regressions


def plan_sync(args: argparse.Namespace) -> int:
    registry_path = Path(
        getattr(args, "registry", DEFAULT_REGISTRY)
    ).resolve()
    registry = load_registry(registry_path)
    baseline_seed_path = Path(
        getattr(
            args,
            "source_baseline_seed",
            DEFAULT_SOURCE_BASELINE_SEED,
        )
    ).resolve()
    baseline_state_path = Path(
        getattr(
            args,
            "source_baseline_state",
            DEFAULT_SOURCE_BASELINE_STATE,
        )
    ).resolve()
    source_baselines, source_baseline_context = _load_source_baselines(
        baseline_seed_path,
        baseline_state_path,
    )
    selection = build_selection_spec(
        registry,
        family_ids=args.family,
        domains=getattr(args, "domain", None),
        after=args.after,
        explicit_message_specs=args.message_id,
    )
    explicit_ids = set(selection.get("explicit_message_ids", {}).values())
    chat, messages = discover_live_messages(
        registry,
        explicit_ids,
        selection["family_ids"],
    )
    selected, classified, unclassified = select_messages(registry, messages, selection)
    run_id = datetime.now().strftime("%Y%m%d-%H%M%S") + f"-{os.getpid()}"
    run_dir = args.runtime_root.resolve() / run_id
    downloads_dir = run_dir / "downloads"
    stages_dir = run_dir / "staged"
    stages_dir.mkdir(parents=True, exist_ok=True)
    blockers = []
    tables = []
    families = family_map(registry)
    for family_id in selection["family_ids"]:
        family = families[family_id]
        message = dict(selected[family_id])
        source_quality = None
        try:
            source_path = download_message(message, family, downloads_dir)
            message["download_path"] = str(source_path)
            message["source_sha256"] = sha256_file(source_path)
            target_path = Path(family["target_workbook"]).resolve()
            if not target_path.exists():
                raise WorkflowError(f"Target workbook does not exist: {target_path}")
            source_write, source_meta = read_records(
                source_path,
                family["source_sheet"],
                family["target_columns"],
                aliases=family.get("column_aliases"),
                constants=family.get("constant_columns"),
                ignored_columns=family.get("ignored_source_columns"),
                column_transforms=family.get("column_transforms"),
                data_only=False,
            )
            source_effective, _ = read_records(
                source_path,
                family["source_sheet"],
                family["target_columns"],
                aliases=family.get("column_aliases"),
                constants=family.get("constant_columns"),
                ignored_columns=family.get("ignored_source_columns"),
                column_transforms=family.get("column_transforms"),
                data_only=True,
            )
            if family.get("materialize_source_values"):
                source_write = [dict(record) for record in source_effective]
            target_write, target_meta = read_records(
                target_path,
                family["target_sheet"],
                family["target_columns"],
                data_only=False,
            )
            target_effective, _ = read_records(
                target_path,
                family["target_sheet"],
                family["target_columns"],
                data_only=True,
            )
            requires_materialized_transform = bool(
                family.get("source_scope")
                or family.get("source_record_transforms")
            )
            if requires_materialized_transform:
                transformed_effective, snapshot_records, transform_audit = (
                    transform_source_records(
                        family,
                        source_effective,
                        target_effective,
                    )
                )
                transformed_write = _copy_records(transformed_effective)
            else:
                transformed_write = source_write
                transformed_effective = source_effective
                snapshot_records = _copy_records(source_effective)
                transform_audit = {
                    "source_rows_before_scope": len(source_effective),
                    "source_rows_after_scope": len(source_effective),
                    "transforms": [],
                }
            snapshot = source_slice_snapshot(family, snapshot_records)
            selected_effective, source_selection = (
                select_source_records_for_merge(
                    family,
                    transformed_effective,
                    snapshot,
                    source_baselines,
                )
            )
            selected_slice_values = set(source_selection["selected_slices"])
            selected_write = [
                dict(record)
                for record in transformed_write
                if _text(record.get(family["slice_column"]))
                in selected_slice_values
            ]
            if len(selected_write) != len(selected_effective):
                raise WorkflowError(
                    f"Selected source streams are not aligned: {family_id}"
                )
            validation_records = transformed_effective
            if (
                family.get("source_merge_mode")
                == "changed_source_slices"
            ):
                validation_records = selected_effective
            source_issues = (
                validate_source_records(
                    family,
                    validation_records,
                    source_meta,
                )
                if validation_records
                else []
            )
            if source_issues:
                raise WorkflowError(f"Source validation failed: {source_issues}")
            source_quality = evaluate_source_quality(
                family,
                message,
                transformed_effective,
                target_effective,
                relative_source_records=selected_effective,
            )
            if (
                registry.get("require_source_quality_gates")
                and source_quality["status"] != "pass"
            ):
                raise WorkflowError(
                    f"Source quality gate blocked {family_id}: {source_quality['issues']}"
                )
            merged_write, merged_effective, diff = merge_records(
                family,
                target_write,
                target_effective,
                selected_write,
                selected_effective,
            )
            candidate_record_issues = []
            if family.get("validate_candidate_records"):
                candidate_record_issues = validate_source_records(
                    family,
                    merged_effective,
                    {
                        "sheet": family["target_sheet"],
                        "formula_count": count_formula_values(merged_write),
                    },
                )
                if candidate_record_issues:
                    raise WorkflowError(
                        f"Candidate record validation failed: "
                        f"{candidate_record_issues}"
                    )
            before_validation = operator_validation(target_path, target_path)
            stage_path = None
            stage_meta = None
            candidate_path = target_path
            if diff["changed"]:
                stage_path = stages_dir / target_path.name
                stage_meta = rebuild_workbook(target_path, stage_path, family, merged_write)
                staged_effective, _ = read_records(
                    stage_path,
                    family["target_sheet"],
                    family["target_columns"],
                    data_only=True,
                )
                if not records_equal(staged_effective, merged_effective, family["target_columns"]):
                    raise WorkflowError(f"Staged workbook effective values differ from planned values: {family_id}")
                candidate_path = stage_path
            after_validation = operator_validation(candidate_path, target_path)
            regressions = validation_regressions(before_validation, after_validation)
            if family.get("allow_baseline_target_errors"):
                if regressions:
                    raise WorkflowError(f"Target validation regressed: {regressions}")
            elif after_validation.get("error_count", 0):
                raise WorkflowError(f"Target validation failed: {after_validation.get('issues', [])}")
            tables.append(
                {
                    "family_id": family_id,
                    "business_name": family["business_name"],
                    "source_message": message,
                    "source_metadata": source_meta,
                    "source_transform_audit": transform_audit,
                    "source_slice_snapshot": snapshot,
                    "source_selection": source_selection,
                    "source_validation": {"ok": True, "issues": []},
                    "source_quality": source_quality,
                    "target_path": str(target_path),
                    "target_sheet": family["target_sheet"],
                    "platform_temp_table": family["platform_temp_table"],
                    "target_before_sha256": sha256_file(target_path),
                    "target_after_sha256": sha256_file(candidate_path),
                    "stage_path": str(stage_path) if stage_path else None,
                    "stage_metadata": stage_meta,
                    "diff": diff,
                    "validation_before": before_validation,
                    "validation_after": after_validation,
                    "validation_regressions": regressions,
                    "candidate_record_validation": {
                        "ok": not candidate_record_issues,
                        "issues": candidate_record_issues,
                    },
                    "allow_baseline_target_errors": bool(family.get("allow_baseline_target_errors")),
                }
            )
        except Exception as exc:  # noqa: BLE001
            blocker = {
                "family_id": family_id,
                "message": str(exc),
                "error_type": type(exc).__name__,
            }
            if source_quality is not None:
                blocker["source_quality"] = source_quality
            blockers.append(blocker)
    plan = {
        "schema_version": "1.1.0",
        "artifact_type": "GovernedDepartmentTempTableSyncPlan",
        "generated_at": datetime.now().astimezone().isoformat(timespec="seconds"),
        "status": "ready" if not blockers and len(tables) == len(selection["family_ids"]) else "blocked",
        "registry_path": str(registry_path),
        "registry_sha256": sha256_file(registry_path),
        "source_baseline_context": source_baseline_context,
        "runtime_dir": str(run_dir),
        "source_context": {
            "registered_chats": chat.get("registered_chats", []),
            "registered_sources": [
                {
                    "family_id": family_id,
                    "domain": families[family_id]["domain"],
                    "chat_name": source_chat_name(registry, families[family_id]),
                    "chat_id": source_chat_id(registry, families[family_id]),
                    "source_kind": source_kind(families[family_id]),
                    "sender_name": source_sender_name(registry, families[family_id]),
                    "sender_open_id": source_sender_id(registry, families[family_id]),
                }
                for family_id in selection["family_ids"]
            ],
            "source_message_count": len(messages),
        },
        "selection": selection,
        "selected_message_ids": {family_id: message["message_id"] for family_id, message in selected.items()},
        "history_counts": {family_id: len(items) for family_id, items in classified.items()},
        "unclassified_files": unclassified,
        "tables": tables,
        "blockers": blockers,
        "production_upload_authorized": False,
    }
    plan_path = run_dir / "sync_plan.json"
    plan_sha = write_artifact(plan_path, plan, "plan_sha256")
    summary = {
        "ok": plan["status"] == "ready",
        "status": plan["status"],
        "plan_path": str(plan_path),
        "plan_sha256": plan_sha,
        "table_count": len(tables),
        "blockers": blockers,
        "tables": [
            {
                "family_id": table["family_id"],
                "domain": families[table["family_id"]]["domain"],
                "source_file": table["source_message"]["file_name"],
                "source_time": table["source_message"]["create_time"],
                "source_quality": table["source_quality"],
                "source_selection": table["source_selection"],
                "source_transform_audit": table["source_transform_audit"],
                "target_path": table["target_path"],
                "platform_temp_table": table["platform_temp_table"],
                "diff": table["diff"],
            }
            for table in tables
        ],
    }
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    return 0 if summary["ok"] else 1


def apply_local(args: argparse.Namespace) -> int:
    if not args.confirm_local_write:
        raise WorkflowError("apply-local requires --confirm-local-write.")
    plan = load_artifact(args.plan.resolve(), "plan_sha256", args.expected_plan_sha256)
    if plan.get("status") != "ready":
        raise WorkflowError("Only a ready plan can be applied locally.")
    registry_path = Path(plan["registry_path"])
    if not registry_path.exists() or sha256_file(registry_path) != plan["registry_sha256"]:
        raise WorkflowError("Workflow registry drifted after planning.")
    assert_source_baseline_state_current(plan)
    registry = load_registry(registry_path)
    assert_plan_source_quality_current(plan, registry)
    selection = plan.get("selection") or build_selection_spec(registry)
    explicit_ids = set(selection.get("explicit_message_ids", {}).values())
    _, current_messages = discover_live_messages(
        registry,
        explicit_ids,
        selection["family_ids"],
    )
    current_selected, _, _ = select_messages(registry, current_messages, selection)
    current_ids = {family_id: message["message_id"] for family_id, message in current_selected.items()}
    if current_ids != plan["selected_message_ids"]:
        raise WorkflowError("Newer matching Feishu source messages appeared after planning; create a fresh plan.")
    for table in plan["tables"]:
        target = Path(table["target_path"])
        source = Path(table["source_message"]["download_path"])
        if sha256_file(target) != table["target_before_sha256"]:
            raise WorkflowError(f"Target workbook drifted after planning: {target}")
        if sha256_file(source) != table["source_message"]["source_sha256"]:
            raise WorkflowError(f"Downloaded source drifted after planning: {source}")
        if table["stage_path"]:
            stage = Path(table["stage_path"])
            if sha256_file(stage) != table["target_after_sha256"]:
                raise WorkflowError(f"Staged workbook drifted after planning: {stage}")
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backups = []
    replacements = []
    applied_tables = []
    changed_tables = [table for table in plan["tables"] if table["diff"]["changed"]]
    try:
        for table in changed_tables:
            target = Path(table["target_path"])
            backup = target.with_name(f"{target.stem}.backup_{timestamp}{target.suffix}")
            if backup.exists():
                raise WorkflowError(f"Backup path already exists: {backup}")
            shutil.copy2(target, backup)
            backups.append({"family_id": table["family_id"], "target": str(target), "backup": str(backup)})
        for table in changed_tables:
            target = Path(table["target_path"])
            stage = Path(table["stage_path"])
            temp_target = sibling_temp_path(target, "apply")
            replacement = {
                "family_id": table["family_id"],
                "target": str(target),
                "temporary_path": str(temp_target),
                "status": "copying",
            }
            replacements.append(replacement)
            shutil.copy2(stage, temp_target)
            replacement["status"] = "replacing"
            try:
                replacement.update(replace_file_with_retry(temp_target, target))
            except Exception as exc:
                if isinstance(exc, AtomicReplaceError):
                    replacement.update(exc.details)
                else:
                    replacement["status"] = "failed"
                replacement["error"] = exception_summary(exc)
                raise
            applied_tables.append(table)
            if sha256_file(target) != table["target_after_sha256"]:
                raise WorkflowError(f"Applied target hash mismatch: {target}")
            validation = operator_validation(target, target)
            if table["allow_baseline_target_errors"]:
                regressions = validation_regressions(table["validation_before"], validation)
                if regressions:
                    raise WorkflowError(f"Applied validation regressed for {target}: {regressions}")
            elif validation.get("error_count", 0):
                raise WorkflowError(f"Applied validation failed for {target}: {validation.get('issues', [])}")
    except Exception as exc:
        backup_by_target = {item["target"]: item for item in backups}
        rollback_entries = []
        for table in reversed(applied_tables):
            target = Path(table["target_path"])
            backup = Path(backup_by_target[str(target)]["backup"])
            rollback_temp = sibling_temp_path(target, "rollback")
            rollback_entry = {
                "family_id": table["family_id"],
                "target": str(target),
                "backup": str(backup),
                "temporary_path": str(rollback_temp),
                "status": "copying",
            }
            rollback_entries.append(rollback_entry)
            try:
                shutil.copy2(backup, rollback_temp)
                rollback_entry["status"] = "replacing"
                rollback_entry.update(
                    replace_file_with_retry(rollback_temp, target)
                )
                if sha256_file(target) != table["target_before_sha256"]:
                    raise WorkflowError(
                        f"Rollback target hash mismatch: {target}"
                    )
                rollback_entry["verified"] = True
            except Exception as rollback_exc:
                if isinstance(rollback_exc, AtomicReplaceError):
                    rollback_entry.update(rollback_exc.details)
                else:
                    rollback_entry["status"] = "failed"
                rollback_entry["verified"] = False
                rollback_entry["error"] = exception_summary(rollback_exc)

        rollback_verification = []
        for table in changed_tables:
            target = Path(table["target_path"])
            current_sha256 = sha256_file(target) if target.exists() else None
            rollback_verification.append(
                {
                    "family_id": table["family_id"],
                    "target": str(target),
                    "current_sha256": current_sha256,
                    "expected_sha256": table["target_before_sha256"],
                    "verified": current_sha256
                    == table["target_before_sha256"],
                }
            )
        rollback_verified = all(
            item["verified"] for item in rollback_verification
        )
        failure_receipt = {
            "schema_version": "1.0.0",
            "artifact_type": "GovernedDepartmentTempTableLocalApplyFailureReceipt",
            "created_at": datetime.now().astimezone().isoformat(
                timespec="seconds"
            ),
            "status": (
                "failed_rolled_back"
                if rollback_verified
                else "failed_rollback_incomplete"
            ),
            "plan_path": str(args.plan.resolve()),
            "plan_sha256": plan["plan_sha256"],
            "failure": exception_summary(exc),
            "backups": backups,
            "replacements": replacements,
            "rollback_entries": rollback_entries,
            "rollback_verification": rollback_verification,
        }
        failure_receipt_path = (
            Path(plan["runtime_dir"]) / "local_apply_failure_receipt.json"
        )
        write_artifact(
            failure_receipt_path,
            failure_receipt,
            "receipt_sha256",
        )
        if rollback_verified:
            raise WorkflowError(
                "Local apply failed and all targets remain at their "
                f"pre-plan hashes. Failure receipt: {failure_receipt_path}. "
                f"Cause: {exc}"
            ) from exc
        raise WorkflowError(
            "Local apply failed and rollback could not be fully verified. "
            f"Manual inspection is required. Failure receipt: "
            f"{failure_receipt_path}. Cause: {exc}"
        ) from exc
    receipt = {
        "schema_version": "1.0.0",
        "artifact_type": "GovernedDepartmentTempTableLocalApplyReceipt",
        "created_at": datetime.now().astimezone().isoformat(timespec="seconds"),
        "status": "success" if changed_tables else "success_no_local_changes",
        "plan_path": str(args.plan.resolve()),
        "plan_sha256": plan["plan_sha256"],
        "backups": backups,
        "replacements": replacements,
        "tables": [
            {
                "family_id": table["family_id"],
                "target_path": table["target_path"],
                "changed": table["diff"]["changed"],
                "current_sha256": sha256_file(Path(table["target_path"])),
                "expected_sha256": table["target_after_sha256"],
            }
            for table in plan["tables"]
        ],
    }
    receipt_path = Path(plan["runtime_dir"]) / "local_apply_receipt.json"
    receipt_sha = write_artifact(receipt_path, receipt, "receipt_sha256")
    print(
        json.dumps(
            {
                "ok": True,
                "status": receipt["status"],
                "receipt_path": str(receipt_path),
                "receipt_sha256": receipt_sha,
                "changed_table_count": len(changed_tables),
                "backups": backups,
            },
            ensure_ascii=False,
            indent=2,
        )
    )
    return 0


def upload_production(args: argparse.Namespace) -> int:
    if not args.confirm_production_upload:
        raise WorkflowError("upload requires --confirm-production-upload.")
    receipt = load_artifact(args.local_receipt.resolve(), "receipt_sha256", args.expected_receipt_sha256)
    if not str(receipt.get("status", "")).startswith("success"):
        raise WorkflowError("Local apply receipt is not successful.")
    plan_path = Path(receipt["plan_path"])
    plan = load_artifact(plan_path, "plan_sha256", receipt["plan_sha256"])
    registry_path = Path(plan["registry_path"])
    if sha256_file(registry_path) != plan["registry_sha256"]:
        raise WorkflowError("Workflow registry drifted after planning.")
    assert_source_baseline_state_current(plan)
    registry = load_registry(registry_path)
    assert_plan_source_quality_current(plan, registry)
    selection = plan.get("selection") or build_selection_spec(registry)
    explicit_ids = set(selection.get("explicit_message_ids", {}).values())
    _, current_messages = discover_live_messages(
        registry,
        explicit_ids,
        selection["family_ids"],
    )
    current_selected, _, _ = select_messages(registry, current_messages, selection)
    current_ids = {family_id: message["message_id"] for family_id, message in current_selected.items()}
    if current_ids != plan["selected_message_ids"]:
        raise WorkflowError("Newer matching Feishu source messages appeared after planning; create a fresh plan.")
    tables_by_id = {table["family_id"]: table for table in plan["tables"]}
    for item in receipt["tables"]:
        target = Path(item["target_path"])
        if sha256_file(target) != item["expected_sha256"]:
            raise WorkflowError(f"Local target drifted after apply: {target}")
    uploads = []
    failed = None
    accepted_baseline_tables = []
    processed_family_ids = []
    selected_order = selection["family_ids"]
    for family_id in selected_order:
        table = tables_by_id[family_id]
        if not table["diff"]["changed"]:
            uploads.append(
                {
                    "family_id": family_id,
                    "target_path": table["target_path"],
                    "platform_temp_table": table["platform_temp_table"],
                    "ok": True,
                    "status": "accepted_no_upload",
                    "import_history_row": None,
                    "validation_result": table["validation_after"],
                    "elapsed_seconds": 0,
                }
            )
            accepted_baseline_tables.append(table)
            processed_family_ids.append(family_id)
            continue
        command = [
            str(OPERATOR_SCRIPT),
            "upload-temp-table",
            "--file",
            table["target_path"],
            "--target-table",
            table["platform_temp_table"],
            "--target-mode",
            "reuse",
            "--import-mode",
            "overwrite",
        ]
        if args.headed:
            command.append("--headed")
        if not table["allow_baseline_target_errors"]:
            command.append("--strict-validation")
        try:
            result = run_json_command(sys.executable, command, cwd=OPERATOR_ROOT, timeout=args.timeout_seconds)
            if not result.get("ok"):
                raise WorkflowError(
                    f"Operator upload did not succeed for {family_id}: "
                    f"{result.get('status')}"
                )
            uploads.append(
                {
                    "family_id": family_id,
                    "target_path": table["target_path"],
                    "platform_temp_table": table["platform_temp_table"],
                    "ok": bool(result.get("ok")),
                    "status": result.get("status"),
                    "import_history_row": result.get("import_history_row"),
                    "validation_result": result.get("validation_result"),
                    "elapsed_seconds": result.get("elapsed_seconds"),
                }
            )
            accepted_baseline_tables.append(table)
            processed_family_ids.append(family_id)
        except Exception as exc:  # noqa: BLE001
            failed = {"family_id": family_id, "message": str(exc), "error_type": type(exc).__name__}
            break
    baseline_update = update_source_baseline_state(
        plan,
        accepted_baseline_tables,
    )
    pending_families = [
        family_id
        for family_id in selected_order
        if family_id not in set(processed_family_ids)
    ]
    upload_receipt = {
        "schema_version": "1.0.0",
        "artifact_type": "GovernedDepartmentTempTableUploadReceipt",
        "created_at": datetime.now().astimezone().isoformat(timespec="seconds"),
        "status": "success" if failed is None and len(uploads) == len(selected_order) else "partial_failure",
        "local_receipt_path": str(args.local_receipt.resolve()),
        "local_receipt_sha256": receipt["receipt_sha256"],
        "uploads": uploads,
        "source_baseline_update": baseline_update,
        "failure": failed,
        "pending_families": pending_families if failed else [],
    }
    upload_receipt_path = Path(plan["runtime_dir"]) / "upload_receipt.json"
    upload_sha = write_artifact(upload_receipt_path, upload_receipt, "receipt_sha256")
    summary = {
        "ok": upload_receipt["status"] == "success",
        "status": upload_receipt["status"],
        "receipt_path": str(upload_receipt_path),
        "receipt_sha256": upload_sha,
        "uploads": uploads,
        "failure": failed,
        "pending_families": upload_receipt["pending_families"],
    }
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    return 0 if summary["ok"] else 1


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description=(
            "Plan and execute the governed Qingcheng and market-consultant "
            "Feishu-to-temp-table workflow."
        )
    )
    subparsers = parser.add_subparsers(dest="command", required=True)

    plan = subparsers.add_parser("plan", help="Discover and download current group files, then build a local dry-run plan.")
    plan.add_argument("--registry", type=Path, default=DEFAULT_REGISTRY)
    plan.add_argument("--runtime-root", type=Path, default=DEFAULT_RUNTIME_ROOT)
    plan.add_argument(
        "--source-baseline-seed",
        type=Path,
        default=DEFAULT_SOURCE_BASELINE_SEED,
    )
    plan.add_argument(
        "--source-baseline-state",
        type=Path,
        default=DEFAULT_SOURCE_BASELINE_STATE,
    )
    plan.add_argument(
        "--family",
        action="append",
        help="Limit the plan to a registered workbook family id; repeat for multiple families.",
    )
    plan.add_argument(
        "--domain",
        action="append",
        help="Limit the plan to a registered domain id; repeat for multiple domains.",
    )
    plan.add_argument(
        "--after",
        help="Only consider messages strictly after this ISO datetime or Unix timestamp.",
    )
    plan.add_argument(
        "--message-id",
        action="append",
        help="Bind one selected family to an exact Feishu message: <family_id>=<om_message_id>.",
    )
    plan.set_defaults(func=plan_sync)

    local = subparsers.add_parser("apply-local", help="Apply a reviewed plan to local maintenance workbooks only.")
    local.add_argument("--plan", type=Path, required=True)
    local.add_argument("--expected-plan-sha256", required=True)
    local.add_argument("--confirm-local-write", action="store_true")
    local.set_defaults(func=apply_local)

    upload = subparsers.add_parser("upload", help="Upload verified local maintenance workbooks to existing temp tables.")
    upload.add_argument("--local-receipt", type=Path, required=True)
    upload.add_argument("--expected-receipt-sha256", required=True)
    upload.add_argument("--confirm-production-upload", action="store_true")
    upload.add_argument("--headed", action="store_true")
    upload.add_argument("--timeout-seconds", type=int, default=600)
    upload.set_defaults(func=upload_production)
    return parser


def main(argv: list[str] | None = None) -> int:
    parser = build_parser()
    args = parser.parse_args(argv)
    try:
        return int(args.func(args))
    except WorkflowError as exc:
        print(json.dumps({"ok": False, "error": {"type": "workflow", "message": str(exc)}}, ensure_ascii=False, indent=2))
        return 1
    except Exception as exc:  # noqa: BLE001
        print(
            json.dumps(
                {
                    "ok": False,
                    "error": {
                        "type": "internal",
                        "subtype": type(exc).__name__,
                        "message": (
                            "Unexpected workflow failure "
                            f"({type(exc).__name__}): {exc}"
                        ),
                    },
                },
                ensure_ascii=False,
                indent=2,
            )
        )
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
