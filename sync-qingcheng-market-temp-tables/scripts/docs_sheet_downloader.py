from __future__ import annotations

import argparse
import hashlib
import json
import re
from pathlib import Path
from typing import Any
from urllib.parse import urlparse

from openpyxl import load_workbook


class DocsSheetDownloadError(RuntimeError):
    pass


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def load_env_file(path: Path) -> dict[str, str]:
    if not path.is_file():
        raise DocsSheetDownloadError("Configured credential environment file does not exist.")
    values: dict[str, str] = {}
    for line in path.read_text(encoding="utf-8-sig").splitlines():
        match = re.match(
            r"^\s*(?:export\s+)?([A-Za-z_][A-Za-z0-9_]*)\s*=\s*(.*)$",
            line,
        )
        if not match:
            continue
        value = match.group(2).strip()
        if len(value) >= 2 and value[0] == value[-1] and value[0] in "\"'":
            value = value[1:-1]
        values[match.group(1)] = value
    return values


def validate_source_url(url: str, patterns: list[str]) -> None:
    parsed = urlparse(url)
    if parsed.scheme != "https" or parsed.netloc.casefold() != "docs.baijia.com":
        raise DocsSheetDownloadError("Source URL is not an approved docs.baijia.com HTTPS link.")
    if not any(re.fullmatch(pattern, url) for pattern in patterns):
        raise DocsSheetDownloadError("Source URL does not match the registered workbook URL patterns.")


def inspect_workbook(path: Path) -> dict[str, Any]:
    if not path.is_file() or path.stat().st_size == 0:
        raise DocsSheetDownloadError("Downloaded workbook is missing or empty.")
    try:
        workbook = load_workbook(path, read_only=True, data_only=False)
    except Exception as exc:  # noqa: BLE001
        raise DocsSheetDownloadError("Downloaded file is not a readable XLSX workbook.") from exc
    try:
        if not workbook.sheetnames:
            raise DocsSheetDownloadError("Downloaded workbook has no worksheets.")
        active = workbook.active
        return {
            "sheet_count": len(workbook.sheetnames),
            "active_sheet": active.title,
            "active_sheet_index": workbook.index(active),
            "sheet_names": list(workbook.sheetnames),
        }
    finally:
        workbook.close()


def download_docs_sheet(
    *,
    url: str,
    output_path: Path,
    env_file: Path,
    url_patterns: list[str],
    expected_title_pattern: str,
    browser_channel: str = "msedge",
    timeout_seconds: int = 120,
    headed: bool = False,
) -> dict[str, Any]:
    validate_source_url(url, url_patterns)
    if timeout_seconds <= 0:
        raise DocsSheetDownloadError("Browser timeout must be positive.")
    if output_path.exists():
        raise DocsSheetDownloadError("Refusing to overwrite an existing downloaded workbook.")

    values = load_env_file(env_file)
    username = values.get("BAIJIA_USERNAME", "")
    password = values.get("BAIJIA_PASSWORD", "")
    if not username or not password:
        raise DocsSheetDownloadError(
            "Credential environment file must define non-empty BAIJIA_USERNAME and BAIJIA_PASSWORD."
        )
    if "@" in username:
        username = username.split("@", 1)[0]

    try:
        from playwright.sync_api import sync_playwright
    except ImportError as exc:
        raise DocsSheetDownloadError(
            "Playwright is unavailable in the configured Python runtime."
        ) from exc

    output_path.parent.mkdir(parents=True, exist_ok=True)
    timeout_ms = timeout_seconds * 1000
    page_title = ""
    with sync_playwright() as playwright:
        browser = playwright.chromium.launch(channel=browser_channel, headless=not headed)
        try:
            context = browser.new_context(accept_downloads=True)
            page = context.new_page()
            page.set_default_timeout(timeout_ms)
            page.goto(url, wait_until="domcontentloaded", timeout=timeout_ms)
            if urlparse(page.url).netloc.casefold() == "cas.baijia.com":
                page.get_by_placeholder("请输入邮箱前缀").fill(username)
                page.get_by_placeholder("请输入密码").fill(password)
                page.get_by_role("button", name="登录", exact=True).click()
            page.wait_for_url(
                re.compile(r"^https://docs\.baijia\.com/sheet/"),
                wait_until="domcontentloaded",
                timeout=timeout_ms,
            )
            page_title = page.title().strip()
            if not re.fullmatch(expected_title_pattern, page_title):
                raise DocsSheetDownloadError("Opened document title does not match the registered source.")
            page.get_by_label("file").click()
            with page.expect_download(timeout=timeout_ms) as download_info:
                page.get_by_text("下载", exact=True).click()
            download_info.value.save_as(str(output_path))
        except DocsSheetDownloadError:
            raise
        except Exception as exc:  # noqa: BLE001
            raise DocsSheetDownloadError(
                "Could not authenticate, open, or download the registered document."
            ) from exc
        finally:
            browser.close()

    workbook_info = inspect_workbook(output_path)
    return {
        "path": str(output_path.resolve()),
        "sha256": sha256_file(output_path),
        "size_bytes": output_path.stat().st_size,
        "page_title": page_title,
        **workbook_info,
    }


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Download one registered docs.baijia.com sheet.")
    parser.add_argument("--url", required=True)
    parser.add_argument("--output", type=Path, required=True)
    parser.add_argument("--env-file", type=Path, required=True)
    parser.add_argument("--url-pattern", action="append", required=True)
    parser.add_argument("--expected-title-pattern", required=True)
    parser.add_argument("--browser-channel", default="msedge")
    parser.add_argument("--timeout-seconds", type=int, default=120)
    parser.add_argument("--headed", action="store_true")
    return parser


def main(argv: list[str] | None = None) -> int:
    args = build_parser().parse_args(argv)
    try:
        summary = download_docs_sheet(
            url=args.url,
            output_path=args.output.resolve(),
            env_file=args.env_file.resolve(),
            url_patterns=args.url_pattern,
            expected_title_pattern=args.expected_title_pattern,
            browser_channel=args.browser_channel,
            timeout_seconds=args.timeout_seconds,
            headed=args.headed,
        )
        print(json.dumps({"ok": True, "data": summary}, ensure_ascii=False, indent=2))
        return 0
    except DocsSheetDownloadError as exc:
        print(
            json.dumps(
                {"ok": False, "error": {"type": "docs_sheet_download", "message": str(exc)}},
                ensure_ascii=False,
                indent=2,
            )
        )
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
