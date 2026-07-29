from __future__ import annotations

import errno
import io
import json
import os
import sys
import tempfile
import unittest
import zipfile
from contextlib import redirect_stdout
from datetime import datetime, timedelta, timezone
from pathlib import Path
from types import SimpleNamespace
from unittest import mock
from xml.etree import ElementTree

from openpyxl import Workbook, load_workbook


SKILL_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(SKILL_ROOT / "scripts"))

from governed_temp_table_sync import (  # noqa: E402
    WorkflowError,
    apply_local,
    assert_plan_source_quality_current,
    build_selection_spec,
    classify_source_message,
    evaluate_source_quality,
    inspect_external_link_integrity,
    list_chat_messages_bot,
    main,
    merge_records,
    load_registry,
    plan_sync,
    read_records,
    replace_file_with_retry,
    resolve_lark_cli,
    select_messages,
    select_source_records_for_merge,
    sha256_file,
    source_slice_snapshot,
    transform_source_records,
    validate_source_records,
    write_artifact,
)


SPREADSHEET_MAIN_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
OFFICE_REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
PACKAGE_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
CONTENT_TYPES_NS = "http://schemas.openxmlformats.org/package/2006/content-types"
MARKUP_COMPATIBILITY_NS = "http://schemas.openxmlformats.org/markup-compatibility/2006"
EXTERNAL_LINK_2021_NS = "http://schemas.microsoft.com/office/spreadsheetml/2021/extlinks2021"


def _xml_bytes(element: ElementTree.Element) -> bytes:
    return ElementTree.tostring(element, encoding="utf-8", xml_declaration=True)


def _inject_dual_external_link(workbook_path: Path, external_path: Path) -> None:
    ElementTree.register_namespace("", SPREADSHEET_MAIN_NS)
    ElementTree.register_namespace("r", OFFICE_REL_NS)
    ElementTree.register_namespace("mc", MARKUP_COMPATIBILITY_NS)
    ElementTree.register_namespace("xxl21", EXTERNAL_LINK_2021_NS)

    with zipfile.ZipFile(workbook_path) as archive:
        workbook = ElementTree.fromstring(archive.read("xl/workbook.xml"))
        sheets = workbook.find(f"{{{SPREADSHEET_MAIN_NS}}}sheets")
        external_references = ElementTree.Element(
            f"{{{SPREADSHEET_MAIN_NS}}}externalReferences"
        )
        ElementTree.SubElement(
            external_references,
            f"{{{SPREADSHEET_MAIN_NS}}}externalReference",
            {f"{{{OFFICE_REL_NS}}}id": "rId99"},
        )
        workbook.insert(list(workbook).index(sheets) + 1, external_references)

        workbook_relationships = ElementTree.fromstring(
            archive.read("xl/_rels/workbook.xml.rels")
        )
        ElementTree.SubElement(
            workbook_relationships,
            f"{{{PACKAGE_REL_NS}}}Relationship",
            {
                "Id": "rId99",
                "Type": f"{OFFICE_REL_NS}/externalLink",
                "Target": "externalLinks/externalLink1.xml",
            },
        )

        content_types = ElementTree.fromstring(archive.read("[Content_Types].xml"))
        ElementTree.SubElement(
            content_types,
            f"{{{CONTENT_TYPES_NS}}}Override",
            {
                "PartName": "/xl/externalLinks/externalLink1.xml",
                "ContentType": (
                    "application/vnd.openxmlformats-officedocument."
                    "spreadsheetml.externalLink+xml"
                ),
            },
        )

        worksheet = ElementTree.fromstring(
            archive.read("xl/worksheets/sheet1.xml")
        )
        formula_cell = worksheet.find(
            f".//{{{SPREADSHEET_MAIN_NS}}}c[@r='A2']"
        )
        formula_value = formula_cell.find(f"{{{SPREADSHEET_MAIN_NS}}}v")
        formula_value.text = "7"

        replacements = {
            "xl/workbook.xml": _xml_bytes(workbook),
            "xl/_rels/workbook.xml.rels": _xml_bytes(workbook_relationships),
            "[Content_Types].xml": _xml_bytes(content_types),
            "xl/worksheets/sheet1.xml": _xml_bytes(worksheet),
        }
        entries = [
            (info, replacements.get(info.filename, archive.read(info.filename)))
            for info in archive.infolist()
        ]

    external_link = ElementTree.Element(
        f"{{{SPREADSHEET_MAIN_NS}}}externalLink",
        {f"{{{MARKUP_COMPATIBILITY_NS}}}Ignorable": "xxl21"},
    )
    external_book = ElementTree.SubElement(
        external_link,
        f"{{{SPREADSHEET_MAIN_NS}}}externalBook",
        {f"{{{OFFICE_REL_NS}}}id": "rId1"},
    )
    alternate_urls = ElementTree.SubElement(
        external_book, f"{{{EXTERNAL_LINK_2021_NS}}}alternateUrls"
    )
    ElementTree.SubElement(
        alternate_urls,
        f"{{{EXTERNAL_LINK_2021_NS}}}absoluteUrl",
        {f"{{{OFFICE_REL_NS}}}id": "rId2"},
    )
    sheet_names = ElementTree.SubElement(
        external_book, f"{{{SPREADSHEET_MAIN_NS}}}sheetNames"
    )
    ElementTree.SubElement(
        sheet_names,
        f"{{{SPREADSHEET_MAIN_NS}}}sheetName",
        {"val": "Sheet1"},
    )
    sheet_data_set = ElementTree.SubElement(
        external_book, f"{{{SPREADSHEET_MAIN_NS}}}sheetDataSet"
    )
    sheet_data = ElementTree.SubElement(
        sheet_data_set,
        f"{{{SPREADSHEET_MAIN_NS}}}sheetData",
        {"sheetId": "0"},
    )
    row = ElementTree.SubElement(
        sheet_data, f"{{{SPREADSHEET_MAIN_NS}}}row", {"r": "1"}
    )
    cell = ElementTree.SubElement(
        row,
        f"{{{SPREADSHEET_MAIN_NS}}}cell",
        {"r": "A1", "t": "n"},
    )
    ElementTree.SubElement(cell, f"{{{SPREADSHEET_MAIN_NS}}}v").text = "7"

    external_relationships = ElementTree.Element(
        f"{{{PACKAGE_REL_NS}}}Relationships"
    )
    for relationship_id in ("rId2", "rId1"):
        ElementTree.SubElement(
            external_relationships,
            f"{{{PACKAGE_REL_NS}}}Relationship",
            {
                "Id": relationship_id,
                "Type": f"{OFFICE_REL_NS}/externalLinkPath",
                "Target": external_path.resolve().as_uri(),
                "TargetMode": "External",
            },
        )

    temporary_path = workbook_path.with_name(f"{workbook_path.name}.fixture.tmp")
    with zipfile.ZipFile(temporary_path, "w") as output:
        for info, payload in entries:
            output.writestr(info, payload)
        output.writestr(
            "xl/externalLinks/externalLink1.xml", _xml_bytes(external_link)
        )
        output.writestr(
            "xl/externalLinks/_rels/externalLink1.xml.rels",
            _xml_bytes(external_relationships),
        )
    temporary_path.replace(workbook_path)


class MergeWorkflowTests(unittest.TestCase):
    def test_replace_file_with_retry_recovers_from_transient_permission_error(
        self,
    ) -> None:
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            source = root / "candidate.xlsx"
            target = root / "target.xlsx"
            source.write_bytes(b"new")
            target.write_bytes(b"old")
            attempts = []
            original_replace = os.replace

            def flaky_replace(source_path: Path, target_path: Path) -> None:
                attempts.append((source_path, target_path))
                if len(attempts) < 3:
                    raise PermissionError(
                        errno.EACCES,
                        "simulated sharing violation",
                        str(target_path),
                    )
                original_replace(source_path, target_path)

            result = replace_file_with_retry(
                source,
                target,
                max_attempts=3,
                initial_delay_seconds=0,
                replace_func=flaky_replace,
                sleep_func=lambda _: None,
            )

            self.assertEqual(result["attempts"], 3)
            self.assertEqual(result["retries"], 2)
            self.assertEqual(len(result["retry_errors"]), 2)
            self.assertEqual(target.read_bytes(), b"new")
            self.assertFalse(source.exists())

    def test_replace_file_with_retry_reports_persistent_lock(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            source = root / "candidate.xlsx"
            target = root / "target.xlsx"
            source.write_bytes(b"new")
            target.write_bytes(b"old")
            attempts = []

            def locked_replace(source_path: Path, target_path: Path) -> None:
                attempts.append((source_path, target_path))
                raise PermissionError(
                    errno.EACCES,
                    "simulated sharing violation",
                    str(target_path),
                )

            with self.assertRaisesRegex(
                WorkflowError,
                "Atomic workbook replacement failed after 3 attempt",
            ):
                replace_file_with_retry(
                    source,
                    target,
                    max_attempts=3,
                    initial_delay_seconds=0,
                    replace_func=locked_replace,
                    sleep_func=lambda _: None,
                )

            self.assertEqual(len(attempts), 3)
            self.assertEqual(target.read_bytes(), b"old")
            self.assertEqual(source.read_bytes(), b"new")

    def test_apply_local_persistent_lock_writes_verified_failure_receipt(
        self,
    ) -> None:
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            run_dir = root / "run"
            run_dir.mkdir()
            target = root / "target.xlsx"
            source = run_dir / "source.xlsx"
            stage = run_dir / "stage.xlsx"
            target.write_bytes(b"old")
            source.write_bytes(b"source")
            stage.write_bytes(b"new")

            registry = {
                "domains": {
                    "test": {
                        "business_name": "Test",
                        "chat": {
                            "name": "Test chat",
                            "expected_chat_id": "oc_test",
                        },
                        "default_sender_name": "Source",
                        "default_sender_open_id": "ou_source",
                    }
                },
                "families": [
                    {
                        "id": "test_family",
                        "domain": "test",
                        "source_filename_patterns": ["^source\\.xlsx$"],
                    }
                ],
                "upload_order": ["test_family"],
            }
            registry_path = root / "registry.json"
            registry_path.write_text(
                json.dumps(registry, ensure_ascii=False),
                encoding="utf-8",
            )
            plan = {
                "status": "ready",
                "registry_path": str(registry_path),
                "registry_sha256": sha256_file(registry_path),
                "runtime_dir": str(run_dir),
                "selection": {
                    "family_ids": ["test_family"],
                    "after": None,
                    "explicit_message_ids": {
                        "test_family": "om_test",
                    },
                    "selection_modes": {
                        "test_family": "explicit_message_id",
                    },
                },
                "selected_message_ids": {
                    "test_family": "om_test",
                },
                "tables": [
                    {
                        "family_id": "test_family",
                        "target_path": str(target),
                        "target_before_sha256": sha256_file(target),
                        "target_after_sha256": sha256_file(stage),
                        "source_message": {
                            "download_path": str(source),
                            "source_sha256": sha256_file(source),
                        },
                        "stage_path": str(stage),
                        "diff": {"changed": True},
                        "allow_baseline_target_errors": False,
                        "validation_before": {
                            "error_count": 0,
                            "issues": [],
                        },
                    }
                ],
            }
            plan_path = run_dir / "sync_plan.json"
            plan_sha256 = write_artifact(
                plan_path,
                plan,
                "plan_sha256",
            )
            args = SimpleNamespace(
                plan=plan_path,
                expected_plan_sha256=plan_sha256,
                confirm_local_write=True,
            )

            def persistent_lock(source_path: Path, target_path: Path) -> None:
                raise PermissionError(
                    errno.EACCES,
                    "simulated persistent target lock",
                    str(target_path),
                )

            with (
                mock.patch(
                    "governed_temp_table_sync.discover_live_messages",
                    return_value=(
                        {"name": "test", "chat_id": "oc_test"},
                        [],
                    ),
                ),
                mock.patch(
                    "governed_temp_table_sync.select_messages",
                    return_value=(
                        {
                            "test_family": {
                                "message_id": "om_test",
                            }
                        },
                        {},
                        [],
                    ),
                ),
                mock.patch(
                    "governed_temp_table_sync.os.replace",
                    side_effect=persistent_lock,
                ),
                mock.patch("governed_temp_table_sync.time.sleep"),
                self.assertRaisesRegex(
                    WorkflowError,
                    "all targets remain at their pre-plan hashes",
                ),
            ):
                apply_local(args)

            self.assertEqual(target.read_bytes(), b"old")
            self.assertFalse(
                (run_dir / "local_apply_receipt.json").exists()
            )
            self.assertFalse((run_dir / "upload_receipt.json").exists())
            failure_receipt = json.loads(
                (
                    run_dir / "local_apply_failure_receipt.json"
                ).read_text(encoding="utf-8")
            )
            self.assertEqual(
                failure_receipt["status"],
                "failed_rolled_back",
            )
            replacement = failure_receipt["replacements"][0]
            self.assertEqual(replacement["status"], "failed")
            self.assertEqual(replacement["attempts"], 8)
            self.assertEqual(replacement["retries"], 7)
            self.assertEqual(len(replacement["retry_errors"]), 8)
            self.assertTrue(
                Path(replacement["temporary_path"]).exists()
            )
            self.assertTrue(
                failure_receipt["rollback_verification"][0]["verified"]
            )

    def test_main_serializes_unexpected_errors(self) -> None:
        captured_stdout = io.StringIO()
        with (
            mock.patch(
                "governed_temp_table_sync.plan_sync",
                side_effect=PermissionError(
                    errno.EACCES,
                    "simulated internal failure",
                ),
            ),
            redirect_stdout(captured_stdout),
        ):
            exit_code = main(["plan"])

        payload = json.loads(captured_stdout.getvalue())
        self.assertEqual(exit_code, 1)
        self.assertFalse(payload["ok"])
        self.assertEqual(payload["error"]["type"], "internal")
        self.assertEqual(payload["error"]["subtype"], "PermissionError")

    def test_registry_upload_order_must_cover_each_family_once(self) -> None:
        registry = {
            "domains": {
                "test": {
                    "business_name": "Test",
                    "chat": {
                        "name": "Test chat",
                        "expected_chat_id": "oc_test",
                    },
                }
            },
            "families": [
                {
                    "id": "a",
                    "domain": "test",
                    "source_filename_patterns": ["^a\\.xlsx$"],
                },
                {
                    "id": "b",
                    "domain": "test",
                    "source_filename_patterns": ["^b\\.xlsx$"],
                },
            ],
            "upload_order": ["a", "a"],
        }
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "registry.json"
            path.write_text(json.dumps(registry), encoding="utf-8")
            with self.assertRaisesRegex(RuntimeError, "upload_order"):
                load_registry(path)

    def test_registered_families_require_complete_source_quality_gates(self) -> None:
        registry = load_registry(SKILL_ROOT / "references" / "workflow_registry.json")
        self.assertTrue(registry["require_source_quality_gates"])
        self.assertEqual(len(registry["families"]), 12)
        self.assertEqual(
            {
                domain: sum(
                    family["domain"] == domain
                    for family in registry["families"]
                )
                for domain in registry["domains"]
            },
            {"qingcheng": 6, "market_consultant": 6},
        )
        for family in registry["families"]:
            quality = family["source_quality"]
            self.assertGreater(quality["max_age_hours"], 0)
            self.assertGreaterEqual(quality["row_count"]["max"], quality["row_count"]["min"])
            self.assertEqual(
                quality["relative_change"]["baseline"],
                "same_slice_or_latest_target",
            )
            self.assertTrue(quality["required_column_null_rate"])

    def test_market_local_temp_table_inventory_is_one_to_one(self) -> None:
        registry = load_registry(SKILL_ROOT / "references" / "workflow_registry.json")
        inventory = registry["local_temp_table_inventories"]["market_consultant"]
        mappings = inventory["mappings"]

        self.assertEqual(inventory["platform_database"], "temp_table")
        self.assertEqual(inventory["verification"]["query_id"], "1506786567")
        self.assertEqual(len(mappings), 9)
        self.assertEqual(
            {
                mapping["local_filename"]: mapping["platform_temp_table"]
                for mapping in mappings
            },
            {
                "ceshiqudao_pingyou.xlsx": "dingxi01_ceshiqudao_pingyou",
                "cost.xlsx": "dingxi01_cost",
                "daoke_1_6_t.xlsx": "dingxi01_daoke_1_6_t",
                "jiagou_db.xlsx": "dingxi01_jiagou_db",
                "jiagou_xinren.xlsx": "dingxi01_jiagou_xinren",
                "jiagou_zx.xlsx": "dingxi01_jiagou_zx",
                "jinliang_goal.xlsx": "dingxi01_jinliang_goal",
                "pingyou_jg.xlsx": "dingxi01_pingyou_jg",
                "plan_id.xlsx": "dingxi01_plan_id",
            },
        )
        self.assertEqual(
            sum(mapping["automation_scope"] == "managed" for mapping in mappings),
            6,
        )
        self.assertEqual(
            sum(mapping["automation_scope"] == "mapping_only" for mapping in mappings),
            3,
        )

    def test_registry_rejects_duplicate_platform_temp_table_mapping(self) -> None:
        registry_path = SKILL_ROOT / "references" / "workflow_registry.json"
        registry = json.loads(registry_path.read_text(encoding="utf-8"))
        mappings = registry["local_temp_table_inventories"]["market_consultant"][
            "mappings"
        ]
        mappings[1]["platform_temp_table"] = mappings[0]["platform_temp_table"]

        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "registry.json"
            path.write_text(
                json.dumps(registry, ensure_ascii=False),
                encoding="utf-8",
            )
            with self.assertRaisesRegex(
                WorkflowError,
                "Duplicate platform temp-table target",
            ):
                load_registry(path)

    def test_registry_rejects_missing_required_source_quality_gate(self) -> None:
        registry = {
            "require_source_quality_gates": True,
            "domains": {
                "test": {
                    "business_name": "Test",
                    "chat": {
                        "name": "Test chat",
                        "expected_chat_id": "oc_test",
                    },
                }
            },
            "families": [
                {
                    "id": "a",
                    "domain": "test",
                    "source_filename_patterns": ["^a\\.xlsx$"],
                }
            ],
            "upload_order": ["a"],
        }
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "registry.json"
            path.write_text(json.dumps(registry), encoding="utf-8")
            with self.assertRaisesRegex(WorkflowError, "Source quality gate is required"):
                load_registry(path)

    def test_lark_cli_is_resolved_before_download_cwd_changes(self) -> None:
        with mock.patch("governed_temp_table_sync.shutil.which", return_value=r".\lark-cli.cmd"):
            resolved = resolve_lark_cli()

        self.assertTrue(Path(resolved).is_absolute())
        self.assertEqual(Path(resolved).name, "lark-cli.cmd")

    def test_bot_chat_message_listing_uses_explicit_pagination(self) -> None:
        responses = [
            {
                "data": {
                    "has_more": True,
                    "page_token": "next-page",
                    "messages": [{"message_id": "m1"}],
                }
            },
            {
                "data": {
                    "has_more": False,
                    "page_token": "",
                    "messages": [{"message_id": "m2"}],
                }
            },
        ]
        with mock.patch(
            "governed_temp_table_sync.run_json_command",
            side_effect=responses,
        ) as run:
            messages = list_chat_messages_bot("lark-cli", "oc_test")

        self.assertEqual(
            [message["message_id"] for message in messages],
            ["m1", "m2"],
        )
        commands = [call.args[1] for call in run.call_args_list]
        self.assertTrue(all("--page-all" not in command for command in commands))
        self.assertNotIn("--page-token", commands[0])
        self.assertEqual(
            commands[1][commands[1].index("--page-token") + 1],
            "next-page",
        )

    def test_overlapping_slice_is_replaced_instead_of_appended(self) -> None:
        family = {
            "target_columns": ["qici", "name", "goal"],
            "slice_column": "qici",
            "slice_order": "asc",
            "key_columns": ["qici", "name"],
        }
        target = [
            {"qici": "20260710期", "name": "A", "goal": 10},
            {"qici": "20260716期", "name": "A", "goal": 20},
        ]
        source = [{"qici": "20260716期", "name": "A", "goal": 25}]

        merged, effective, diff = merge_records(family, target, target, source, source)

        self.assertEqual(merged, effective)
        self.assertEqual([row["goal"] for row in effective], [10, 25])
        self.assertEqual(diff["replaced_slices"], ["20260716期"])
        self.assertEqual(diff["target_rows_after"], 2)

    def test_qingcheng_scope_replacement_preserves_market_rows(self) -> None:
        family = {
            "target_columns": ["qici", "employee_email_name", "dept_1"],
            "slice_column": "qici",
            "slice_order": "desc",
            "key_columns": ["qici", "employee_email_name"],
            "target_scope": {"column": "dept_1", "equals": "青橙项目部"},
        }
        target = [
            {"qici": "20260722期", "employee_email_name": "old", "dept_1": "青橙项目部"},
            {"qici": "20260722期", "employee_email_name": "market", "dept_1": "市场顾问部"},
        ]
        source = [{"qici": "20260722期", "employee_email_name": "new", "dept_1": "青橙项目部"}]

        _, effective, diff = merge_records(family, target, target, source, source)

        self.assertEqual([row["employee_email_name"] for row in effective], ["new", "market"])
        self.assertEqual(diff["scoped_rows_removed"], 1)
        self.assertEqual(diff["replaced_slices"], ["20260722期"])

    def test_same_effective_values_are_idempotent_even_if_formula_stream_differs(self) -> None:
        family = {
            "target_columns": ["qici", "name", "goal"],
            "slice_column": "qici",
            "slice_order": "asc",
            "key_columns": ["qici", "name"],
        }
        target_write = [{"qici": "20260716期", "name": "A", "goal": "=10+10"}]
        target_effective = [{"qici": "20260716期", "name": "A", "goal": 20}]
        source_write = [{"qici": "20260716期", "name": "A", "goal": 20}]
        source_effective = [{"qici": "20260716期", "name": "A", "goal": 20}]

        _, _, diff = merge_records(
            family, target_write, target_effective, source_write, source_effective
        )

        self.assertFalse(diff["changed"])
        self.assertEqual(diff["unchanged_slices"], ["20260716期"])

    def test_changed_slice_noop_preserves_original_row_order(self) -> None:
        family = {
            "target_columns": ["qici", "name"],
            "slice_column": "qici",
            "slice_order": "desc",
            "source_merge_mode": "changed_source_slices",
        }
        target = [
            {"qici": "20260101期", "name": "oldest"},
            {"qici": "20260728期", "name": "latest"},
        ]

        merged_write, merged_effective, diff = merge_records(
            family,
            target,
            target,
            [],
            [],
        )

        self.assertFalse(diff["changed"])
        self.assertEqual(merged_write, target)
        self.assertEqual(merged_effective, target)

    def test_bootstrap_slices_are_one_time_pending_work(self) -> None:
        family = {
            "id": "market_plan_id",
            "target_columns": ["qici", "group_id"],
            "slice_column": "qici",
            "source_merge_mode": "changed_source_slices",
            "source_baseline_id": "market_plan_id",
            "max_changed_slices": 2,
            "recent_slice_window": 1,
            "bootstrap_slices": ["0529期"],
            "reviewed_historical_slices": ["0529期"],
        }
        records = [{"qici": "0529期", "group_id": "1"}]
        snapshot = source_slice_snapshot(family, records)
        baseline = {
            "families": {
                "market_plan_id": {
                    "message_id": "om_seed",
                    **snapshot,
                    "bootstrap_pending": True,
                }
            }
        }

        selected, selection = select_source_records_for_merge(
            family,
            records,
            snapshot,
            baseline,
        )
        self.assertEqual(selected, records)
        self.assertEqual(selection["bootstrap_slices"], ["0529期"])

        baseline["families"]["market_plan_id"].pop(
            "bootstrap_pending"
        )
        selected, selection = select_source_records_for_merge(
            family,
            records,
            snapshot,
            baseline,
        )
        self.assertEqual(selected, [])
        self.assertEqual(selection["bootstrap_slices"], [])

    def test_unreviewed_added_historical_slice_is_blocking(self) -> None:
        family = {
            "id": "market_period_architecture",
            "target_columns": ["qici", "name"],
            "slice_column": "qici",
            "source_merge_mode": "changed_source_slices",
            "source_baseline_id": "market_period_architecture",
            "max_changed_slices": 2,
            "recent_slice_window": 1,
        }
        baseline_records = [
            {"qici": "20260728期", "name": "A"},
        ]
        baseline_snapshot = source_slice_snapshot(
            family,
            baseline_records,
        )
        current_records = [
            {"qici": "20260101期", "name": "legacy"},
            *baseline_records,
        ]
        current_snapshot = source_slice_snapshot(
            family,
            current_records,
        )
        baselines = {
            "families": {
                "market_period_architecture": baseline_snapshot,
            }
        }

        with self.assertRaisesRegex(
            WorkflowError,
            "unreviewed historical slices",
        ):
            select_source_records_for_merge(
                family,
                current_records,
                current_snapshot,
                baselines,
            )

    def test_plan_group_name_prefix_checks_only_period_prefixed_names(
        self,
    ) -> None:
        family = {
            "target_columns": [
                "year",
                "qici",
                "group_id",
                "group_name",
            ],
            "key_columns": ["year", "qici", "group_id"],
            "validation_rules": [
                {
                    "type": "group_name_qici_prefix",
                    "qici_column": "qici",
                    "name_column": "group_name",
                }
            ],
        }
        accepted = [
            {
                "year": 2026,
                "qici": "0529期",
                "group_id": "1",
                "group_name": "0529期-市场-高中年级",
            },
            {
                "year": 2026,
                "qici": "0529期",
                "group_id": "2",
                "group_name": "2026年-短期班-抖音私信",
            },
        ]
        self.assertEqual(validate_source_records(family, accepted), [])

        rejected = [
            dict(
                accepted[0],
                group_name="0522期-市场-高中年级",
            )
        ]
        issues = validate_source_records(family, rejected)
        self.assertEqual(issues[0]["rule"], "group_name_qici_prefix")

    def test_required_reference_mapping_blocks_unknown_channel(
        self,
    ) -> None:
        family = {
            "id": "market_attendance_schedule",
            "target_columns": [
                "qici",
                "qudao",
                "grade",
                "begin_time",
                "dow",
                "ke_1",
                "channel",
            ],
            "source_record_transforms": [
                {
                    "type": "fill_from_reference_mapping",
                    "column": "channel",
                    "key_columns": [
                        "qici",
                        "qudao",
                        "grade",
                        "begin_time",
                        "dow",
                        "ke_1",
                    ],
                    "block_ambiguous_keys": True,
                }
            ],
        }
        source = [
            {
                "qici": "20260728期",
                "qudao": "A",
                "grade": "初一",
                "begin_time": "2026-07-28 10:00:00",
                "dow": "1",
                "ke_1": "1",
                "channel": "",
            }
        ]

        with self.assertRaisesRegex(
            WorkflowError,
            "Cannot fill required source column channel",
        ):
            transform_source_records(family, source, [])

        reference = [dict(source[0], channel="paid")]
        transformed, _, audit = transform_source_records(
            family,
            source,
            reference,
        )
        self.assertEqual(transformed[0]["channel"], "paid")
        self.assertEqual(audit["transforms"][0]["filled"], 1)

    def test_duplicate_source_key_is_blocking(self) -> None:
        family = {
            "key_columns": ["qici", "name"],
            "validation_rules": [{"type": "slice_format", "column": "qici", "pattern": "^\\d{8}期$"}],
        }
        rows = [
            {"qici": "20260716期", "name": "A"},
            {"qici": "20260716期", "name": "A"},
        ]

        issues = validate_source_records(family, rows)

        self.assertTrue(any(issue["rule"] == "unique_key" for issue in issues))

    def test_source_quality_passes_all_four_gate_types(self) -> None:
        family = {
            "id": "quality",
            "slice_column": "qici",
            "target_columns": ["qici", "name", "goal"],
            "source_quality": {
                "policy_version": "1.0.0",
                "max_age_hours": 48,
                "row_count": {"min": 2, "max": 4},
                "relative_change": {
                    "baseline": "same_slice_or_latest_target",
                    "max_ratio": 0.5,
                },
                "required_column_null_rate": {
                    "qici": 0,
                    "name": 0,
                    "goal": 0,
                },
            },
        }
        now = datetime(2026, 7, 29, 0, 0, tzinfo=timezone.utc)
        source = [
            {"qici": "20260729期", "name": "A", "goal": 1},
            {"qici": "20260729期", "name": "B", "goal": 2},
        ]
        target = [
            {"qici": "20260722期", "name": "A", "goal": 1},
            {"qici": "20260722期", "name": "B", "goal": 2},
        ]

        report = evaluate_source_quality(
            family,
            {"create_time": (now - timedelta(hours=2)).isoformat()},
            source,
            target,
            now=now,
        )

        self.assertEqual(report["status"], "pass")
        self.assertTrue(report["row_count"]["ok"])
        self.assertEqual(
            report["relative_change"]["slices"][0]["baseline_kind"],
            "latest_target_slice",
        )
        self.assertFalse(report["issues"])

    def test_source_quality_blocks_age_rows_relative_change_and_null_rate(self) -> None:
        family = {
            "id": "quality",
            "slice_column": "qici",
            "target_columns": ["qici", "name", "goal"],
            "source_quality": {
                "policy_version": "1.0.0",
                "max_age_hours": 24,
                "row_count": {"min": 3, "max": 4},
                "relative_change": {
                    "baseline": "same_slice_or_latest_target",
                    "max_ratio": 0.25,
                },
                "required_column_null_rate": {
                    "qici": 0,
                    "name": 0,
                    "goal": 0,
                },
            },
        }
        now = datetime(2026, 7, 29, 0, 0, tzinfo=timezone.utc)
        source = [
            {"qici": "20260722期", "name": "A", "goal": None},
            {"qici": "20260722期", "name": "B", "goal": 2},
        ]
        target = [
            {"qici": "20260722期", "name": "A", "goal": 1},
            {"qici": "20260722期", "name": "B", "goal": 2},
            {"qici": "20260722期", "name": "C", "goal": 3},
            {"qici": "20260722期", "name": "D", "goal": 4},
        ]

        report = evaluate_source_quality(
            family,
            {"create_time": (now - timedelta(hours=25)).isoformat()},
            source,
            target,
            now=now,
        )

        self.assertEqual(report["status"], "blocked")
        self.assertEqual(
            {issue["rule"] for issue in report["issues"]},
            {
                "source_max_age",
                "source_row_count",
                "required_column_null_rate",
                "source_relative_change",
            },
        )

    def test_apply_and_upload_gate_rejects_expired_quality_report(self) -> None:
        now = datetime(2026, 7, 29, 0, 0, tzinfo=timezone.utc)
        registry = {"require_source_quality_gates": True}
        plan = {
            "selection": {"family_ids": ["a"]},
            "tables": [
                {
                    "family_id": "a",
                    "source_quality": {
                        "status": "pass",
                        "source_expires_at": (now - timedelta(seconds=1)).isoformat(),
                    },
                }
            ],
        }

        with self.assertRaisesRegex(WorkflowError, "expired"):
            assert_plan_source_quality_current(plan, registry, now=now)

    def test_subset_selection_uses_only_requested_families(self) -> None:
        registry = {
            "domains": {
                "test": {
                    "business_name": "Test",
                    "chat": {
                        "name": "Test chat",
                        "expected_chat_id": "oc_test",
                    },
                    "default_sender_open_id": "ou_source",
                }
            },
            "families": [
                {
                    "id": "a",
                    "domain": "test",
                    "source_filename_patterns": ["^a\\.xlsx$"],
                },
                {
                    "id": "b",
                    "domain": "test",
                    "source_filename_patterns": ["^b\\.xlsx$"],
                },
            ],
            "upload_order": ["a", "b"],
        }
        selection = build_selection_spec(registry, family_ids=["b"])
        messages = [
            {
                "message_id": "om_a",
                "file_name": "a.xlsx",
                "create_time": "1000",
                "chat_id": "oc_test",
                "sender_id": "ou_source",
            },
            {
                "message_id": "om_b",
                "file_name": "b.xlsx",
                "create_time": "2000",
                "chat_id": "oc_test",
                "sender_id": "ou_source",
            },
        ]

        selected, _, _ = select_messages(registry, messages, selection)

        self.assertEqual(selection["family_ids"], ["b"])
        self.assertEqual(selected["b"]["message_id"], "om_b")

    def test_explicit_message_binding_does_not_float_to_a_newer_message(self) -> None:
        registry = {
            "domains": {
                "test": {
                    "business_name": "Test",
                    "chat": {
                        "name": "Test chat",
                        "expected_chat_id": "oc_test",
                    },
                    "default_sender_open_id": "ou_source",
                }
            },
            "families": [
                {
                    "id": "a",
                    "domain": "test",
                    "source_filename_patterns": ["^a\\.xlsx$"],
                }
            ],
            "upload_order": ["a"],
        }
        selection = build_selection_spec(
            registry,
            family_ids=["a"],
            explicit_message_specs=["a=om_old"],
        )
        messages = [
            {
                "message_id": "om_old",
                "file_name": "a.xlsx",
                "create_time": "1000",
                "chat_id": "oc_test",
                "sender_id": "ou_source",
            },
            {
                "message_id": "om_new",
                "file_name": "a.xlsx",
                "create_time": "2000",
                "chat_id": "oc_test",
                "sender_id": "ou_source",
            },
        ]

        selected, _, _ = select_messages(registry, messages, selection)

        self.assertEqual(selected["a"]["message_id"], "om_old")

    def test_after_cutoff_is_strict(self) -> None:
        registry = {
            "domains": {
                "test": {
                    "business_name": "Test",
                    "chat": {
                        "name": "Test chat",
                        "expected_chat_id": "oc_test",
                    },
                    "default_sender_open_id": "ou_source",
                }
            },
            "families": [
                {
                    "id": "a",
                    "domain": "test",
                    "source_filename_patterns": ["^a\\.xlsx$"],
                }
            ],
            "upload_order": ["a"],
        }
        selection = build_selection_spec(registry, family_ids=["a"], after="1970-01-01T00:00:01Z")
        messages = [
            {
                "message_id": "om_equal",
                "file_name": "a.xlsx",
                "create_time": "1000",
                "chat_id": "oc_test",
                "sender_id": "ou_source",
            },
            {
                "message_id": "om_after",
                "file_name": "a.xlsx",
                "create_time": "2000",
                "chat_id": "oc_test",
                "sender_id": "ou_source",
            },
        ]

        selected, _, _ = select_messages(registry, messages, selection)

        self.assertEqual(selected["a"]["message_id"], "om_after")

    def test_course_schedule_link_is_bound_to_registered_sender(self) -> None:
        registry = load_registry(SKILL_ROOT / "references" / "workflow_registry.json")
        message = {
            "message_id": "om_course",
            "chat_id": "oc_e604e064976c022ab4289fc2fb979332",
            "sender": {
                "id": "ou_3168c83ffe93b49a192755c8e31e2bc5",
                "name": "李怡青",
            },
            "content": (
                "【青橙行课--开课时间(1-6节课)】\n"
                "https://docs.baijia.com/sheet/DQUZrYU50dkZPa1FtZWJSU1JE?tab=vhzqxm"
            ),
        }

        self.assertEqual(classify_source_message(registry, message), "course_schedule")
        message["sender"] = {"id": "ou_unregistered", "name": "其他人"}
        self.assertIsNone(classify_source_message(registry, message))

    def test_active_course_sheet_maps_and_normalizes_registered_columns(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "course.xlsx"
            workbook = Workbook()
            old = workbook.active
            old.title = "0722期"
            current = workbook.create_sheet("0728期")
            workbook.active = workbook.index(current)
            current.append(["渠道", "开课时间", "期次", "工作日-1", "年级", "课节数(第几节课)", "工作日"])
            current.append(["抖音", "2026-07-28  13:30:00", "20260728期", 2, "初一", 1, "周二"])
            workbook.save(path)
            workbook.close()

            records, metadata = read_records(
                path,
                "$active",
                ["qudao", "begin_time", "qici", "dow", "grade", "ke_1"],
                aliases={
                    "渠道": "qudao",
                    "开课时间": "begin_time",
                    "期次": "qici",
                    "工作日-1": "dow",
                    "年级": "grade",
                    "课节数(第几节课)": "ke_1",
                },
                ignored_columns=["工作日"],
                column_transforms={
                    "begin_time": ["collapse_whitespace"],
                    "ke_1": ["to_text"],
                },
                data_only=True,
            )

        self.assertEqual(metadata["sheet"], "0728期")
        self.assertEqual(records[0]["begin_time"], "2026-07-28 13:30:00")
        self.assertEqual(records[0]["ke_1"], "1")
        self.assertNotIn("工作日", records[0])

    def test_course_sheet_must_match_slice_suffix(self) -> None:
        family = {
            "key_columns": ["qici", "qudao"],
            "validation_rules": [{"type": "sheet_slice_suffix", "column": "qici"}],
        }

        issues = validate_source_records(
            family,
            [{"qici": "20260722期", "qudao": "抖音"}],
            {"sheet": "0728期"},
        )

        self.assertTrue(any(issue["rule"] == "sheet_slice_suffix" for issue in issues))

    @unittest.skipUnless(sys.platform == "win32", "requires Windows Excel COM")
    def test_result_architecture_windows_candidate_preserves_links_and_appends_117_rows(
        self,
    ) -> None:
        try:
            import pythoncom  # noqa: F401, PLC0415
            import win32com.client  # noqa: F401, PLC0415
        except ImportError:
            self.skipTest("pywin32 is unavailable")

        columns = [
            "employee_name",
            "employee_email_name",
            "email_prefix",
            "leader_employee_email_name",
            "leader_email_prefix",
            "dazu",
            "jingli",
            "xuebu",
            "qici",
        ]
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            external_path = root / "external.xlsx"
            external_workbook = Workbook()
            external_workbook.active["A1"] = 7
            external_workbook.save(external_path)
            external_workbook.close()

            target_path = root / "qing_team_jg.xlsx"
            target_workbook = Workbook()
            target_sheet = target_workbook.active
            target_sheet.title = "data"
            target_sheet.append(columns)
            target_sheet.append(
                [
                    "=[1]Sheet1!$A$1",
                    "history01",
                    "history01",
                    "leader01",
                    "leader01",
                    "历史组",
                    "历史经理",
                    "一部",
                    "20260710期",
                ]
            )
            target_sheet.auto_filter.ref = "A1:I2"
            target_workbook.save(target_path)
            target_workbook.close()
            _inject_dual_external_link(target_path, external_path)
            self.assertTrue(inspect_external_link_integrity(target_path)["ok"])
            target_sha256 = sha256_file(target_path)

            source_path = root / "全员结果数据架构.xlsx"
            source_workbook = Workbook()
            source_sheet = source_workbook.active
            source_sheet.title = "data"
            source_sheet.append(columns)
            for index in range(1, 118):
                source_sheet.append(
                    [
                        f"员工{index:03}",
                        f"employee{index:03}",
                        f"employee{index:03}",
                        "主管001",
                        "leader001",
                        "一组",
                        "经理001",
                        "一部",
                        "20260728期",
                    ]
                )
            source_workbook.save(source_path)
            source_workbook.close()

            family = {
                "id": "result_architecture",
                "domain": "qingcheng",
                "business_name": "全员结果数据架构",
                "source_filename_patterns": ["^全员结果数据架构\\.xlsx$"],
                "source_sheet": "data",
                "target_workbook": str(target_path),
                "target_sheet": "data",
                "platform_temp_table": "dingxi01_qing_team_jg",
                "target_columns": columns,
                "column_aliases": {},
                "constant_columns": {},
                "preserve_external_links": True,
                "slice_column": "qici",
                "slice_order": "asc",
                "key_columns": ["qici", "employee_email_name"],
                "validation_rules": [
                    {
                        "type": "slice_format",
                        "column": "qici",
                        "pattern": "^\\d{8}期$",
                    },
                    {
                        "type": "lowercase_ascii_prefix",
                        "columns": ["email_prefix", "leader_email_prefix"],
                    },
                    {
                        "type": "disallow_values",
                        "column": "dazu",
                        "values": ["0"],
                    },
                ],
            }
            registry = {
                "version": 2,
                "domains": {
                    "qingcheng": {
                        "business_name": "青橙项目部",
                        "chat": {
                            "name": "青橙数据对接",
                            "expected_chat_id": "oc_test",
                        },
                        "default_sender_name": "郅玲玉",
                        "default_sender_open_id": "ou_source",
                    }
                },
                "families": [family],
                "upload_order": ["result_architecture"],
                "deferred_filename_patterns": [],
            }
            registry_path = root / "registry.json"
            registry_path.write_text(
                json.dumps(registry, ensure_ascii=False, indent=2),
                encoding="utf-8",
            )
            runtime_root = root / "runtime"
            args = SimpleNamespace(
                registry=registry_path,
                runtime_root=runtime_root,
                family=["result_architecture"],
                after=None,
                message_id=[],
            )
            message = {
                "message_id": "om_result",
                "create_time": "2026-07-27 22:40",
                "message_position": "1517",
                "message_app_link": None,
                "message_type": "file",
                "content": (
                    '<file key="file_test" name="全员结果数据架构.xlsx"/>'
                ),
                "chat_id": "oc_test",
                "sender_id": "ou_source",
                "sender_name": "郅玲玉",
                "source_kind": "file_attachment",
                "file_key": "file_test",
                "file_name": "全员结果数据架构.xlsx",
            }
            validation = {"error_count": 0, "issues": []}
            captured_stdout = io.StringIO()

            with (
                mock.patch(
                    "governed_temp_table_sync.discover_live_messages",
                    return_value=(
                        {"name": "青橙数据对接", "chat_id": "oc_test"},
                        [message],
                    ),
                ),
                mock.patch(
                    "governed_temp_table_sync.download_message",
                    return_value=source_path,
                ),
                mock.patch(
                    "governed_temp_table_sync.operator_validation",
                    return_value=validation,
                ),
                mock.patch(
                    "governed_temp_table_sync.upload_production"
                ) as upload_production,
                redirect_stdout(captured_stdout),
            ):
                exit_code = plan_sync(args)

            self.assertEqual(exit_code, 0, captured_stdout.getvalue())
            upload_production.assert_not_called()
            self.assertEqual(sha256_file(target_path), target_sha256)

            plan_paths = list(runtime_root.glob("*/sync_plan.json"))
            self.assertEqual(len(plan_paths), 1)
            plan = json.loads(plan_paths[0].read_text(encoding="utf-8"))
            self.assertEqual(plan["status"], "ready")
            self.assertFalse(plan["production_upload_authorized"])
            self.assertFalse(list(runtime_root.glob("*/upload_receipt.json")))

            table = plan["tables"][0]
            self.assertEqual(table["diff"]["source_rows"], 117)
            self.assertEqual(table["diff"]["new_slices"], ["20260728期"])
            self.assertEqual(table["diff"]["target_rows_after"], 118)
            stage_metadata = table["stage_metadata"]
            self.assertTrue(
                stage_metadata["external_link_preservation"]["restored"]
            )
            self.assertFalse(
                stage_metadata["external_link_preservation"][
                    "stage_before_integrity"
                ]["ok"]
            )
            self.assertTrue(stage_metadata["external_link_integrity"]["ok"])
            self.assertEqual(
                stage_metadata["recalculation"]["backend"], "excel_com"
            )
            self.assertEqual(stage_metadata["recalculation"]["total_errors"], 0)

            stage_path = Path(table["stage_path"])
            self.assertTrue(inspect_external_link_integrity(stage_path)["ok"])
            staged_workbook = load_workbook(stage_path, data_only=True)
            staged_rows = list(
                staged_workbook["data"].iter_rows(min_row=2, values_only=True)
            )
            staged_workbook.close()
            self.assertEqual(
                sum(row[8] == "20260728期" for row in staged_rows), 117
            )
            self.assertEqual(len(staged_rows), 118)

            original_replace = os.replace
            target_replace_attempts = []

            def transient_target_lock(
                source_candidate: Path,
                destination: Path,
            ) -> None:
                if Path(destination) == target_path:
                    target_replace_attempts.append(
                        (Path(source_candidate), Path(destination))
                    )
                    if len(target_replace_attempts) == 1:
                        raise PermissionError(
                            errno.EACCES,
                            "simulated transient target lock",
                            str(destination),
                        )
                original_replace(source_candidate, destination)

            apply_args = SimpleNamespace(
                plan=plan_paths[0],
                expected_plan_sha256=plan["plan_sha256"],
                confirm_local_write=True,
            )
            apply_stdout = io.StringIO()
            with (
                mock.patch(
                    "governed_temp_table_sync.discover_live_messages",
                    return_value=(
                        {"name": "青橙数据对接", "chat_id": "oc_test"},
                        [message],
                    ),
                ),
                mock.patch(
                    "governed_temp_table_sync.operator_validation",
                    return_value=validation,
                ),
                mock.patch(
                    "governed_temp_table_sync.upload_production"
                ) as upload_production_after_apply,
                mock.patch(
                    "governed_temp_table_sync.os.replace",
                    side_effect=transient_target_lock,
                ),
                mock.patch("governed_temp_table_sync.time.sleep"),
                redirect_stdout(apply_stdout),
            ):
                apply_exit_code = apply_local(apply_args)

            self.assertEqual(
                apply_exit_code,
                0,
                apply_stdout.getvalue(),
            )
            upload_production_after_apply.assert_not_called()
            self.assertEqual(len(target_replace_attempts), 2)
            self.assertEqual(
                sha256_file(target_path),
                table["target_after_sha256"],
            )
            self.assertTrue(inspect_external_link_integrity(target_path)["ok"])

            local_receipt_path = plan_paths[0].parent / "local_apply_receipt.json"
            self.assertTrue(local_receipt_path.exists())
            self.assertFalse(
                (plan_paths[0].parent / "local_apply_failure_receipt.json").exists()
            )
            self.assertFalse(
                (plan_paths[0].parent / "upload_receipt.json").exists()
            )
            local_receipt = json.loads(
                local_receipt_path.read_text(encoding="utf-8")
            )
            self.assertEqual(local_receipt["status"], "success")
            self.assertEqual(local_receipt["replacements"][0]["attempts"], 2)
            self.assertEqual(local_receipt["replacements"][0]["retries"], 1)


if __name__ == "__main__":
    unittest.main()
