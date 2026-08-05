"""
Excel beautification functions for the xlsx skill.
Covers Phase 1–3 of the xlsx beautification plan.

All functions take an openpyxl Worksheet as first argument and return it
for chainable usage. Import style definitions from style_palette.py.

Usage:
    from style_palette import PALETTE, FONTS, FILLS, BORDERS, ALIGNMENTS
    from style_apply import (
        apply_title_banner, apply_header_style, apply_banded_rows,
        apply_number_format, apply_data_bars, apply_color_scale,
        apply_pivot_style, apply_sort_indicator, apply_category_colors,
        apply_border_grid, apply_section_header, apply_subtotal_row,
        apply_grand_total_row, apply_kpi_card, auto_style_sheet,
        detect_data_range, detect_header_row,
    )
"""

from copy import copy as shallow_copy
from openpyxl.formatting.rule import DataBarRule, ColorScaleRule, FormulaRule, CellIsRule
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter, column_index_from_string
from style_palette import (
    PALETTE, CATEGORY_PALETTE,
    FONTS, FILLS, BORDERS, ALIGNMENTS, NUMBER_FORMATS,
    NUMBER_FORMAT_RULES,
    _fill, _make_font, make_border, make_alignment,
)


# ═══════════════════════════════════════════════════════════════════════
# Phase 1 — Core auto-beautification (applied on every output)
# ═══════════════════════════════════════════════════════════════════════

def apply_title_banner(ws, title, subtitle=None, max_col=None, title_row=1):
    """
    Create a styled title banner row.

    - Merges cells across full width
    - Dark blue background with white bold title (16pt)
    - Optional gray subtitle row below

    Args:
        ws: openpyxl Worksheet
        title: Main title text
        subtitle: Optional subtitle / data range description
        max_col: Number of columns to merge across (auto-detected if None)
        title_row: Row number for the title (default 1)

    Returns:
        ws for chaining
    """
    max_c = max_col or _detect_max_col(ws)
    col_letter = get_column_letter(max_c)

    # Title row
    if max_c > 1:
        ws.merge_cells(
            start_row=title_row, start_column=1,
            end_row=title_row, end_column=max_c,
        )
    cell = ws.cell(row=title_row, column=1)
    cell.value = title
    cell.font = FONTS['title']
    cell.fill = FILLS['title']
    cell.alignment = ALIGNMENTS['left']
    ws.row_dimensions[title_row].height = 36

    # Subtitle row
    if subtitle:
        sub_row = title_row + 1
        if max_c > 1:
            ws.merge_cells(
                start_row=sub_row, start_column=1,
                end_row=sub_row, end_column=max_c,
            )
        cell = ws.cell(row=sub_row, column=1)
        cell.value = subtitle
        cell.font = FONTS['subtitle']
        cell.fill = FILLS['metadata']
        cell.alignment = ALIGNMENTS['left']
        ws.row_dimensions[sub_row].height = 22

    return ws


def apply_header_style(ws, row=1, max_col=None, freeze=True):
    """
    Apply standardized header-row styling: dark blue fill, white bold text,
    centered alignment, medium bottom border, and frozen panes.

    Args:
        ws: openpyxl Worksheet
        row: Header row number (default 1)
        max_col: Number of header columns (auto-detected if None)
        freeze: Whether to freeze panes below this header row

    Returns:
        ws for chaining
    """
    max_c = max_col or _detect_max_col(ws)
    ws.row_dimensions[row].height = 28

    for col_idx in range(1, max_c + 1):
        cell = ws.cell(row=row, column=col_idx)
        cell.font = FONTS['header']
        cell.fill = FILLS['header']
        cell.alignment = ALIGNMENTS['center_wrap']
        cell.border = BORDERS['header_row']

    if freeze:
        ws.freeze_panes = ws.cell(row=row + 1, column=1)

    return ws


def apply_banded_rows(ws, start_row, end_row, max_col=None):
    """
    Apply alternating white/light-gray row colors to a data range.
    Skips rows that already have a non-default fill.

    Args:
        ws: openpyxl Worksheet
        start_row: First data row (inclusive)
        end_row: Last data row (inclusive)
        max_col: Number of columns to apply to (auto-detected if None)

    Returns:
        ws for chaining
    """
    max_c = max_col or _detect_max_col(ws)
    colors = [FILLS['alt_1'], FILLS['alt_2']]

    for r in range(start_row, end_row + 1):
        band_idx = (r - start_row) % 2
        target_fill = colors[band_idx]
        for c in range(1, max_c + 1):
            cell = ws.cell(row=r, column=c)
            if _fill_is_empty(cell.fill):
                cell.fill = target_fill

    return ws


def apply_number_format(ws, col_map=None, start_row=2, end_row=None, header_row=1):
    """
    Apply number formatting based on column header keywords.
    Detects headers from the specified header row and auto-assigns formats.

    Args:
        ws: openpyxl Worksheet
        col_map: Optional dict of {column_letter: format_key} for manual control.
                 If None, auto-detects from header text in the header_row.
        start_row: First data row (default 2, below header)
        end_row: Last data row (auto-detected if None)
        header_row: Row containing column headers for auto-detection (default 1)

    Returns:
        dict of {column_letter: format_key} applied
    """
    applied = {}
    end_r = end_row or ws.max_row

    # If col_map is provided, apply directly
    if col_map:
        for col_letter, fmt_key in col_map.items():
            nf = NUMBER_FORMATS.get(fmt_key, fmt_key)
            col_idx = column_index_from_string(col_letter)
            _apply_format_to_column(ws, col_idx, nf, start_row, end_r)
            applied[col_letter] = fmt_key
        return applied

    # Auto-detect from headers
    max_c = _detect_max_col(ws)
    for col_idx in range(1, max_c + 1):
        header_cell = ws.cell(row=header_row, column=col_idx)
        if header_cell.value is None:
            continue
        header_text = str(header_cell.value).strip()
        fmt_key = _detect_number_format(header_text)
        if fmt_key:
            nf = NUMBER_FORMATS[fmt_key]
            _apply_format_to_column(ws, col_idx, nf, start_row, end_r)
            col_letter = get_column_letter(col_idx)
            applied[col_letter] = fmt_key

    return applied


def apply_auto_fit_columns(ws, max_width=42, min_width=8, max_col=None, sample_rows=100):
    """
    Auto-fit column widths based on content length.
    Chinese characters count as ~2 units.

    Args:
        ws: openpyxl Worksheet
        max_width: Maximum column width
        min_width: Minimum column width
        max_col: Number of columns to process (auto-detected if None)
        sample_rows: Max rows to sample for width calculation

    Returns:
        ws for chaining
    """
    max_c = max_col or _detect_max_col(ws)
    max_r = min(ws.max_row, sample_rows)

    for col_idx in range(1, max_c + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0

        for r in range(1, max_r + 1):
            cell = ws.cell(row=r, column=col_idx)
            if cell.value is not None:
                text = str(cell.value)
                # Handle multi-line
                lines = text.split('\n')
                for line in lines:
                    length = sum(2 if ord(ch) > 127 else 1 for ch in line)
                    max_len = max(max_len, length)

        # Adjust for font size: larger fonts need more width
        width = min(max_len + 3, max_width)
        width = max(width, min_width)
        ws.column_dimensions[col_letter].width = width

    return ws


# ═══════════════════════════════════════════════════════════════════════
# Phase 2 — Conditional formatting & visual indicators
# ═══════════════════════════════════════════════════════════════════════

def apply_data_bars(ws, col_range, color=None, show_value=True):
    """
    Add in-cell data bars (horizontal bar proportional to cell value).

    Args:
        ws: openpyxl Worksheet
        col_range: Column range like 'D2:D50' or a list of ranges
        color: Bar color hex (default: accent blue '5B9BD5')
        show_value: Whether to display the numeric value alongside the bar

    Returns:
        ws for chaining
    """
    bar_color = color or PALETTE['accent_blue']
    rule = DataBarRule(
        start_type='min', end_type='max',
        color=bar_color, showValue=show_value,
    )

    ranges = [col_range] if isinstance(col_range, str) else col_range
    for rng in ranges:
        ws.conditional_formatting.add(rng, rule)

    return ws


def apply_color_scale(ws, col_range, scheme='red_white_green'):
    """
    Apply 2-color or 3-color gradient scale to a column range.

    Predefined schemes:
        'red_white_green' — Red (low) → White (mid) → Green (high) — for rates/ratios
        'green_white'     — White (low) → Green (high) — for positive-only metrics
        'red_white'       — White (low) → Red (high) — for negative/downside metrics
        'blue_white'      — White (low) → Blue (high) — for neutral metrics
        'blue_white_red'  — Blue (low) → White (mid) → Red (high) — diverging

    Args:
        ws: openpyxl Worksheet
        col_range: Column range string like 'E2:E100'
        scheme: One of the predefined scheme names

    Returns:
        ws for chaining
    """
    SCHEMES = {
        'red_white_green': [
            {'value': None, 'type': 'min', 'color': PALETTE['accent_red']},
            {'value': None, 'type': 'percentile', 'mid': 50, 'color': 'FFFFFF'},
            {'value': None, 'type': 'max', 'color': PALETTE['accent_green']},
        ],
        'green_white': [
            {'value': None, 'type': 'min', 'color': 'FFFFFF'},
            {'value': None, 'type': 'max', 'color': PALETTE['accent_green']},
        ],
        'red_white': [
            {'value': None, 'type': 'min', 'color': 'FFFFFF'},
            {'value': None, 'type': 'max', 'color': PALETTE['accent_red']},
        ],
        'blue_white': [
            {'value': None, 'type': 'min', 'color': 'FFFFFF'},
            {'value': None, 'type': 'max', 'color': PALETTE['accent_blue']},
        ],
        'blue_white_red': [
            {'value': None, 'type': 'min', 'color': PALETTE['accent_red']},
            {'value': None, 'type': 'percentile', 'mid': 50, 'color': 'FFFFFF'},
            {'value': None, 'type': 'max', 'color': PALETTE['accent_blue']},
        ],
    }

    scheme_data = SCHEMES.get(scheme, SCHEMES['red_white_green'])
    colors = [d['color'] for d in scheme_data]

    if len(scheme_data) == 2:
        rule = ColorScaleRule(
            start_type=scheme_data[0]['type'],
            start_value=scheme_data[0].get('value'),
            start_color=colors[0],
            end_type=scheme_data[1]['type'],
            end_value=scheme_data[1].get('value'),
            end_color=colors[1],
        )
    else:
        rule = ColorScaleRule(
            start_type=scheme_data[0]['type'],
            start_value=scheme_data[0].get('value'),
            start_color=colors[0],
            mid_type=scheme_data[1]['type'],
            mid_value=scheme_data[1].get('value'),
            mid_color=colors[1],
            end_type=scheme_data[2]['type'],
            end_value=scheme_data[2].get('value'),
            end_color=colors[2],
        )

    ranges = [col_range] if isinstance(col_range, str) else col_range
    for rng in ranges:
        ws.conditional_formatting.add(rng, rule)

    return ws


# ═══════════════════════════════════════════════════════════════════════
# Phase 3 — Pivot table & category-aware beautification
# ═══════════════════════════════════════════════════════════════════════

def apply_pivot_style(
    ws,
    data_start_row,
    data_end_row,
    max_col,
    row_label_cols=None,
    value_cols=None,
    has_total_row=True,
):
    """
    Apply pivot-table-specific beautification:
    - Row-label columns: bold + left-aligned
    - Value columns: right-aligned + optional data bars
    - Category grouping: auto-detect repeated labels and add subtle separators
    - Total/subtotal rows: distinct background + top border

    Args:
        ws: openpyxl Worksheet
        data_start_row: First data row of the pivot body
        data_end_row: Last data row of the pivot body
        max_col: Total column count
        row_label_cols: List of column indices (1-based) that are row labels
        value_cols: List of column indices (1-based) that are numeric values
        has_total_row: Whether the last row is a grand total

    Returns:
        ws for chaining
    """
    # Auto-detect row label vs value columns if not specified
    if row_label_cols is None or value_cols is None:
        _labels, _values = _classify_pivot_columns(ws, 1, data_start_row, max_col)
        if row_label_cols is None:
            row_label_cols = _labels
        if value_cols is None:
            value_cols = _values

    # Style row labels: left-aligned, slightly bold
    for r in range(data_start_row, data_end_row + 1):
        for c in row_label_cols:
            cell = ws.cell(row=r, column=c)
            if cell.value is not None:
                cell.alignment = ALIGNMENTS['left']
                cell.font = _make_font('body_bold')

    # Style value columns: right-aligned
    for r in range(data_start_row, data_end_row + 1):
        for c in value_cols:
            cell = ws.cell(row=r, column=c)
            if cell.value is not None:
                cell.alignment = ALIGNMENTS['right']

    # Detect and style subtotal rows (rows where first label col has "(小计)" or "Subtotal")
    subtotal_keywords = ('小计', '合计', 'Subtotal', 'subtotal', 'Sub Total', 'sub total')
    for r in range(data_start_row, data_end_row):
        first_label_val = ws.cell(row=r, column=row_label_cols[0]).value
        if first_label_val and any(kw in str(first_label_val) for kw in subtotal_keywords):
            for c in range(1, max_col + 1):
                cell = ws.cell(row=r, column=c)
                cell.font = _make_font('body_bold')
                cell.fill = FILLS['subtotal']
                cell.border = BORDERS['total_top']

    # Style grand total row
    if has_total_row and data_end_row > data_start_row:
        apply_grand_total_row(ws, data_end_row, max_col)

    # Add light grid borders to the pivot body
    apply_border_grid(ws, data_start_row, data_end_row, max_col, border_name='grid')

    return ws


def apply_sort_indicator(ws, col_letter, direction='desc', header_row=1, data_start_row=2, data_end_row=None):
    """
    Add a sort-direction indicator arrow (▲/▼) to a column header and
    lightly highlight the sorted column's data cells.

    Args:
        ws: openpyxl Worksheet
        col_letter: Column letter of the sorted column (e.g., 'D')
        direction: 'desc' (▼) or 'asc' (▲)
        header_row: Row number of the header
        data_start_row: First data row
        data_end_row: Last data row (auto-detected if None)

    Returns:
        ws for chaining
    """
    arrow = ' ▼' if direction == 'desc' else ' ▲'
    header_cell = ws.cell(row=header_row, column=column_index_from_string(col_letter))
    current_val = str(header_cell.value or '')
    if not current_val.endswith(arrow):
        header_cell.value = current_val + arrow

    # Light highlight on the sorted column's data cells
    end_r = data_end_row or ws.max_row
    col_idx = column_index_from_string(col_letter)
    for r in range(data_start_row, end_r + 1):
        cell = ws.cell(row=r, column=col_idx)
        if cell.value is not None and _fill_is_empty(cell.fill):
            cell.fill = FILLS['sort_highlight']

    return ws


def apply_category_colors(ws, category_col, category_config, start_row=2, end_row=None, max_col=None):
    """
    Apply distinct background colors to rows based on category values,
    using conditional formatting rules so they survive data changes.

    Args:
        ws: openpyxl Worksheet
        category_col: Column letter (e.g., 'A') containing category values
        category_config: Either:
            - dict: {'CategoryA': 'E3F2FD', 'CategoryB': 'FFF3E0', ...}
            - list: ['CatA', 'CatB', ...] → auto-assign from CATEGORY_PALETTE
        start_row: First data row
        end_row: Last data row (auto-detected if None)
        max_col: How many columns the color fills across (auto-detected if None)

    Returns:
        ws for chaining
    """
    end_r = end_row or ws.max_row
    max_c = max_col or _detect_max_col(ws)
    end_col = get_column_letter(max_c)
    apply_range = f'${category_col}${start_row}:${end_col}${end_r}'
    col_abs = f'${category_col.upper()}'

    # Build color mapping
    if isinstance(category_config, list):
        color_map = {}
        for i, cat_val in enumerate(category_config):
            color_map[str(cat_val)] = CATEGORY_PALETTE[i % len(CATEGORY_PALETTE)]
    else:
        color_map = category_config

    # Create conditional formatting rules: one FormulaRule per category value
    for cat_value, color_hex in color_map.items():
        formula = [f'{col_abs}{start_row}="{cat_value}"']
        fill = _fill(color_hex)
        rule = FormulaRule(formula=formula, fill=fill)
        ws.conditional_formatting.add(apply_range, rule)

    return ws


# ═══════════════════════════════════════════════════════════════════════
# Border, section, subtotal, and KPI helpers
# ═══════════════════════════════════════════════════════════════════════

def apply_border_grid(ws, start_row, end_row, max_col=None, border_name='grid'):
    """
    Apply grid borders to a rectangular range.

    Args:
        ws: openpyxl Worksheet
        start_row: First row (inclusive)
        end_row: Last row (inclusive)
        max_col: Number of columns (auto-detected if None)
        border_name: Key in BORDERS dict ('grid', 'outline', etc.)

    Returns:
        ws for chaining
    """
    max_c = max_col or _detect_max_col(ws)
    border = BORDERS.get(border_name, BORDERS['grid'])

    for r in range(start_row, end_row + 1):
        for c in range(1, max_c + 1):
            cell = ws.cell(row=r, column=c)
            if cell.value is not None:
                cell.border = border

    return ws


def apply_section_header(ws, row, max_col, text, color='accent_blue'):
    """
    Insert a colored section-divider row with bold white text.

    Args:
        ws: openpyxl Worksheet
        row: Row number for the section header
        max_col: Columns to merge across
        text: Section header text
        color: 'accent_blue' (default), 'accent_green', or hex color string

    Returns:
        ws for chaining
    """
    if color in PALETTE:
        hex_color = PALETTE[color]
    elif color in ('accent_blue',):
        hex_color = PALETTE['accent_blue']
    else:
        hex_color = color

    fill = _fill(hex_color)

    if max_col > 1:
        ws.merge_cells(
            start_row=row, start_column=1,
            end_row=row, end_column=max_col,
        )
    cell = ws.cell(row=row, column=1)
    cell.value = text
    cell.font = FONTS['section']
    cell.fill = fill
    cell.alignment = ALIGNMENTS['left']
    ws.row_dimensions[row].height = 26

    return ws


def apply_subtotal_row(ws, row, max_col):
    """
    Style a row as a subtotal: bold font, light blue bg, medium top border.

    Args:
        ws: openpyxl Worksheet
        row: The subtotal row number
        max_col: Number of columns to apply to

    Returns:
        ws for chaining
    """
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = _make_font('body_bold')
        cell.fill = FILLS['subtotal']
        cell.border = BORDERS['total_top']
    ws.row_dimensions[row].height = 22
    return ws


def apply_grand_total_row(ws, row, max_col):
    """
    Style a row as a grand total: bold font, light blue bg, medium top
    border and double bottom border.

    Args:
        ws: openpyxl Worksheet
        row: The grand total row number
        max_col: Number of columns to apply to

    Returns:
        ws for chaining
    """
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = _make_font('body_bold', size=11)
        cell.fill = FILLS['grand_total']
        cell.border = BORDERS['total_double_bottom']
    ws.row_dimensions[row].height = 24
    return ws


def apply_kpi_card(ws, row, col, value, label, width=2, height=None):
    """
    Create a single KPI card: large value + small label, light blue background.

    Args:
        ws: openpyxl Worksheet
        row: Starting row
        col: Starting column (int, 1-based)
        value: The KPI value (number or string)
        label: Description label below the value
        width: How many columns to span for the card
        height: Row height override (auto if None)

    Returns:
        ws for chaining
    """
    # Value cell
    val_cell = ws.cell(row=row, column=col)
    val_cell.value = value
    val_cell.font = FONTS['kpi_value']
    val_cell.fill = FILLS['kpi']
    val_cell.alignment = ALIGNMENTS['center']
    if width > 1:
        ws.merge_cells(
            start_row=row, start_column=col,
            end_row=row, end_column=col + width - 1,
        )

    ws.row_dimensions[row].height = height or 36

    # Label cell
    if label:
        lbl_row = row + 1
        lbl_cell = ws.cell(row=lbl_row, column=col)
        lbl_cell.value = label
        lbl_cell.font = FONTS['kpi_label']
        lbl_cell.fill = FILLS['kpi']
        lbl_cell.alignment = ALIGNMENTS['center']
        if width > 1:
            ws.merge_cells(
                start_row=lbl_row, start_column=col,
                end_row=lbl_row, end_column=col + width - 1,
            )
        ws.row_dimensions[lbl_row].height = 20

    return ws


def apply_kpi_cards_row(ws, row, metrics, start_col=1, card_width=3):
    """
    Create a horizontal row of KPI cards.

    Args:
        ws: openpyxl Worksheet
        row: Row for the KPI values
        metrics: List of (value, label) tuples
        start_col: First column index (default 1)
        card_width: Column span per card

    Returns:
        ws for chaining
    """
    for i, (value, label) in enumerate(metrics):
        col_start = start_col + i * card_width
        apply_kpi_card(ws, row, col_start, value, label, card_width)

    return ws


# ═══════════════════════════════════════════════════════════════════════
# One-shot auto-style
# ═══════════════════════════════════════════════════════════════════════

def auto_style_sheet(
    ws,
    title=None,
    subtitle=None,
    header_row=1,
    data_start_row=None,
    data_end_row=None,
    max_col=None,
    value_cols=None,
    category_col=None,
    category_map=None,
    freeze_header=True,
    add_borders=True,
    add_banded_rows=True,
    add_number_formats=True,
    add_data_bars=False,
    add_color_scale=False,
    color_scale_scheme='red_white_green',
):
    """
    One-call full-sheet beautification. Applies a sensible sequence of
    styling operations based on the detected structure.

    Usage for a typical data sheet:
        auto_style_sheet(ws, title='Sales Report', subtitle='2026 Q1')

    Usage for a pivot table:
        auto_style_sheet(
            ws, title='Channel Conversion Pivot',
            value_cols=[3, 4, 5], category_col='A',
            category_map={'KOC': 'E3F2FD', 'Douyin': 'FFF3E0'},
            add_color_scale=True,
        )

    Args:
        ws: openpyxl Worksheet
        title: Sheet title (applied as banner at row 1)
        subtitle: Subtitle text
        header_row: Row number of column headers (default 1)
        data_start_row: First data row (default: header_row + 1)
        data_end_row: Last data row (default: ws.max_row)
        max_col: Column count (auto-detected if None)
        value_cols: List of 1-based column indices for numeric values
        category_col: Column letter for category-based coloring
        category_map: Dict of {category_value: hex_color} or list of values
        freeze_header: Freeze panes below header
        add_borders: Apply grid borders
        add_banded_rows: Apply alternating row colors
        add_number_formats: Auto-detect and apply number formats
        add_data_bars: Add data bars to value columns
        add_color_scale: Add color scales to value columns
        color_scale_scheme: Which scheme for color scales

    Returns:
        ws for chaining
    """
    max_c = max_col or _detect_max_col(ws)
    d_start = data_start_row or (header_row + 1)
    d_end = data_end_row or ws.max_row

    # 1. Title banner
    title_banner_row = None
    if title:
        # If header is at row 1 and we have a title, shift header down
        if header_row == 1:
            ws.insert_rows(1, amount=2)
            # Adjust all row references down by 2
            header_row += 2
            d_start += 2
            d_end += 2
            title_banner_row = 1
        elif header_row > 3:
            title_banner_row = 1
        else:
            title_banner_row = 1
        apply_title_banner(ws, title, subtitle, max_c, title_banner_row)

    # 2. Header styling
    apply_header_style(ws, header_row, max_c, freeze=freeze_header)

    # 3. Auto-fit columns
    apply_auto_fit_columns(ws, max_width=42, max_col=max_c)

    # 4. Number formats
    if add_number_formats:
        apply_number_format(ws, start_row=d_start, end_row=d_end, header_row=header_row)

    # 5. Banded rows
    if add_banded_rows:
        apply_banded_rows(ws, d_start, d_end, max_c)

    # 6. Grid borders
    if add_borders:
        apply_border_grid(ws, d_start, d_end, max_c)

    # 7. Category coloring
    if category_col and category_map:
        apply_category_colors(
            ws, category_col, category_map,
            start_row=d_start, end_row=d_end, max_col=max_c,
        )

    # 8. Data bars on value columns
    if add_data_bars and value_cols:
        for ci in value_cols:
            col_l = get_column_letter(ci)
            apply_data_bars(ws, f'{col_l}{d_start}:{col_l}{d_end}')

    # 9. Color scales on value columns
    if add_color_scale and value_cols:
        for ci in value_cols:
            col_l = get_column_letter(ci)
            apply_color_scale(
                ws, f'{col_l}{d_start}:{col_l}{d_end}',
                scheme=color_scale_scheme,
            )

    return ws


# ═══════════════════════════════════════════════════════════════════════
# Utility / detection helpers
# ═══════════════════════════════════════════════════════════════════════

def detect_data_range(ws, header_row=1, min_rows=0):
    """
    Detect the data range of a worksheet.

    Returns:
        (start_row, end_row, max_col) tuple
    """
    max_c = _detect_max_col(ws)
    max_r = ws.max_row
    start_r = header_row + 1

    # Trim trailing empty rows
    while max_r > start_r:
        row_empty = all(
            ws.cell(row=max_r, column=c).value is None
            for c in range(1, max_c + 1)
        )
        if row_empty:
            max_r -= 1
        else:
            break

    if max_r < start_r:
        max_r = start_r

    return start_r, max_r, max_c


def detect_header_row(ws, scan_rows=10):
    """
    Detect the most likely header row by scanning for bold cells or
    filled backgrounds in the first N rows.

    Returns:
        Row index (1-based) of the detected header
    """
    best_row = 1
    best_score = 0

    for r in range(1, min(ws.max_row + 1, scan_rows + 1)):
        score = 0
        for c in range(1, min(ws.max_column + 1, 50)):
            cell = ws.cell(row=r, column=c)
            if cell.value is None:
                continue
            if cell.font and cell.font.bold:
                score += 3
            if cell.fill and not _fill_is_empty(cell.fill):
                score += 2
            if cell.font and cell.font.color:
                score += 1
        if score > best_score:
            best_score = score
            best_row = r

    return best_row


# ═══════════════════════════════════════════════════════════════════════
# Internal helpers
# ═══════════════════════════════════════════════════════════════════════

def _detect_max_col(ws):
    """Find the rightmost non-empty column in row 1, or fall back to ws.max_column."""
    if ws.max_column and ws.max_column < 200:
        return ws.max_column
    # Scan row 1 for the last non-empty cell
    max_c = 0
    for cell in ws[1]:
        if cell.value is not None:
            max_c = max(max_c, cell.column)
    return max_c if max_c > 0 else (ws.max_column or 10)


def _fill_is_empty(fill):
    """Check if a fill is effectively empty (no background color)."""
    if fill is None:
        return True
    if fill.fill_type is None:
        return True
    if fill.start_color is None:
        return True
    rgb = str(fill.start_color.rgb) if fill.start_color.rgb else ''
    return rgb in ('00000000', '0', '000000', 'None', '')


def _detect_number_format(header_text):
    """Match header text keywords to a NUMBER_FORMATS key. Returns key or None."""
    text = header_text.lower()
    for keywords, fmt_key in NUMBER_FORMAT_RULES:
        for kw in keywords:
            if kw in text or kw in header_text:  # check both lower and original
                return fmt_key
    return None


def _apply_format_to_column(ws, col_idx, number_format, start_row, end_row):
    """Apply a number format string to all data cells in a column."""
    for r in range(start_row, end_row + 1):
        cell = ws.cell(row=r, column=col_idx)
        if cell.value is not None:
            try:
                cell.number_format = number_format
            except (ValueError, TypeError):
                pass


def _classify_pivot_columns(ws, header_row, data_start_row, max_col):
    """
    Classify columns as row labels (text-heavy) or values (numeric).
    Samples first 30 data rows.
    """
    labels = []
    values = []
    sample_end = min(data_start_row + 30, ws.max_row)

    for c in range(1, max_col + 1):
        text_count = 0
        num_count = 0
        total = 0

        for r in range(data_start_row, sample_end + 1):
            val = ws.cell(row=r, column=c).value
            if val is None:
                continue
            total += 1
            if isinstance(val, (int, float)):
                num_count += 1
            else:
                text_count += 1

        if total == 0:
            continue

        # If > 60% numeric, treat as value column
        if num_count / total > 0.6:
            values.append(c)
        else:
            labels.append(c)

    return labels, values


# ═══════════════════════════════════════════════════════════════════════
# Template detection & safety
# ═══════════════════════════════════════════════════════════════════════

def detect_existing_template(ws, sample_rows=20):
    """
    Detect whether a worksheet already has substantial hand-crafted styling
    that should be preserved. Returns True if the sheet appears pre-styled.

    Checks for:
    - Non-palette fill colors (anything beyond our standard palette)
    - Custom fonts (not Arial / Microsoft YaHei)
    - Borders on data cells
    - Complex merged cell patterns (beyond a simple title banner)

    Args:
        ws: openpyxl Worksheet
        sample_rows: How many rows to scan

    Returns:
        bool: True if the sheet has significant existing styling
    """
    fills_seen = set()
    fonts_seen = set()
    border_count = 0
    merged_count = len(ws.merged_cells.ranges)
    cells_checked = 0

    # Colors that are part of our beautification palette (or defaults)
    palette_colors = {
        'FF1F4E78', 'FF294E73', 'FF1F3A5F', 'FF203B61', 'FF17324D',
        'FF5B9BD5', 'FF0F766E', 'FFDCEAF5', 'FFD9EAF7', 'FFF3F6F8',
        'FFF6F9FB', 'FFFFFFFF', 'FFFFFF', 'FFEBF3FA',
        '00000000', '0', '000000', 'None', '',
    }

    for r in range(1, min(ws.max_row + 1, sample_rows + 1)):
        for c in range(1, min(ws.max_column + 1, 30)):
            cell = ws.cell(row=r, column=c)
            if cell.value is None:
                continue
            cells_checked += 1

            # Track fills that are NOT in our palette
            if cell.fill and cell.fill.start_color and cell.fill.start_color.rgb:
                rgb = str(cell.fill.start_color.rgb)
                fills_seen.add(rgb)

            # Track fonts
            if cell.font and cell.font.name:
                fonts_seen.add(cell.font.name)

            # Track borders
            b = cell.border
            if b:
                for side in (b.left, b.right, b.top, b.bottom):
                    if side and side.style:
                        border_count += 1
                        break

    non_palette_fills = fills_seen - palette_colors
    standard_fonts = {'Arial', 'Calibri', 'Microsoft YaHei', '微软雅黑', None, '等线', '等线 Light'}
    has_custom_fonts = bool(fonts_seen - standard_fonts)
    has_non_palette_fills = len(non_palette_fills) > 0
    has_many_fills = len(fills_seen) > 4
    has_heavy_borders = border_count > (cells_checked * 0.3)
    has_complex_merges = merged_count > 2

    return has_custom_fonts or has_non_palette_fills or has_many_fills or has_heavy_borders or has_complex_merges


# ═══════════════════════════════════════════════════════════════════════
# Quick-style shorthand (for the most common use case)
# ═══════════════════════════════════════════════════════════════════════

def quick_style(ws, title=None, value_cols=None, category_col=None):
    """
    Minimal one-call styling for quick data exports. Applies header style,
    auto-fit, banded rows, number formats, and borders in one shot.

    This is the recommended default for any new data sheet — use
    auto_style_sheet() when you need full control over every option.

    Args:
        ws: openpyxl Worksheet
        title: Optional sheet title (banner at row 1)
        value_cols: Optional list of 1-based column indices for data bars + color scales
        category_col: Optional column letter for category color coding

    Returns:
        ws for chaining

    Example:
        quick_style(ws, title='月度销售汇总', value_cols=[3, 4, 5], category_col='A')
    """
    # Detect whether headers are already styled
    if detect_existing_template(ws):
        # Sheet has existing styling — only apply non-destructive enhancements
        apply_auto_fit_columns(ws)
        return ws

    # Detect structure
    header_r = detect_header_row(ws)
    d_start, d_end, max_c = detect_data_range(ws, header_row=header_r)

    # Guard: skip if no data
    if d_end <= header_r:
        return ws

    # Title banner
    title_r = None
    if title and header_r <= 2:
        ws.insert_rows(1, amount=2)
        header_r += 2
        d_start += 2
        d_end += 2
        title_r = 1
        apply_title_banner(ws, title, max_col=max_c, title_row=title_r)

    # Core styling
    apply_header_style(ws, row=header_r, max_col=max_c, freeze=True)
    apply_auto_fit_columns(ws, max_width=42, max_col=max_c)
    apply_number_format(ws, start_row=d_start, end_row=d_end, header_row=header_r)
    apply_banded_rows(ws, d_start, d_end, max_c)
    apply_border_grid(ws, d_start, d_end, max_c)

    # Conditional formatting on value columns
    if value_cols:
        for ci in value_cols:
            cl = get_column_letter(ci)
            apply_data_bars(ws, f'{cl}{d_start}:{cl}{d_end}')

    # Category coloring: auto-detect unique values if no explicit map
    if category_col and d_end > d_start:
        cat_values = set()
        col_idx = column_index_from_string(category_col)
        for r in range(d_start, d_end + 1):
            v = ws.cell(row=r, column=col_idx).value
            if v is not None:
                cat_values.add(str(v))
        if len(cat_values) <= 10:
            apply_category_colors(
                ws, category_col, sorted(cat_values),
                start_row=d_start, end_row=d_end, max_col=max_c,
            )

    return ws
