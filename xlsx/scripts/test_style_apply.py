"""
Integration test for xlsx style_apply.py and style_palette.py.
Tests all Phase 1-3 beautification functions against realistic data patterns.
"""
import os, sys, shutil

# Add scripts directory to path
scripts_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'scripts')
sys.path.insert(0, scripts_dir)

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import DataBarRule, ColorScaleRule, FormulaRule

from style_palette import (
    PALETTE, CATEGORY_PALETTE, FONTS, FILLS, BORDERS, ALIGNMENTS,
    NUMBER_FORMATS, NUMBER_FORMAT_RULES,
)
from style_apply import (
    apply_title_banner, apply_header_style, apply_banded_rows,
    apply_number_format, apply_data_bars, apply_color_scale,
    apply_pivot_style, apply_sort_indicator, apply_category_colors,
    apply_border_grid, apply_section_header, apply_subtotal_row,
    apply_grand_total_row, apply_kpi_card, apply_kpi_cards_row,
    apply_auto_fit_columns, auto_style_sheet,
    detect_data_range, detect_header_row,
)

_BASE = os.path.abspath(__file__)
for _ in range(4):
    _BASE = os.path.dirname(_BASE)
OUTPUT_DIR = os.path.join(_BASE, 'outputs')
os.makedirs(OUTPUT_DIR, exist_ok=True)
TEST_FILE = os.path.join(OUTPUT_DIR, 'xlsx_style_integration_test.xlsx')


def make_test_workbook():
    """Build a realistic multi-sheet workbook covering common patterns."""
    wb = Workbook()

    # ── Sheet 1: Pivot table (refund analysis) ──
    ws1 = wb.active
    ws1.title = '透视分析'

    # Header row
    headers = ['期次', '年级', '渠道', '退费金额', '退费率', '净收金额', '触达人数', '转化率']
    for c, h in enumerate(headers, 1):
        ws1.cell(row=3, column=c, value=h)

    # Data rows
    data = [
        ['20260501期', '初中', 'KOC-抖音数学', 90500, 0.062, 1458000, 3420, 0.158],
        ['20260501期', '初中', 'B站信息流-非标', 24200, 0.089, 271910, 1200, 0.112],
        ['20260501期', '初中', '私域IC', 14400, 0.031, 464516, 800, 0.195],
        ['20260501期', '高中', 'KOC-抖音数学', 55800, 0.045, 1240000, 2100, 0.142],
        ['20260501期', '高中', 'B站信息流-非标', 63900, 0.078, 819231, 1800, 0.098],
        ['20260501期', '高中', '私域IC', 75900, 0.052, 1459615, 1500, 0.173],
        ['20260501期', '小学', 'KOC-抖音数学', 12300, 0.021, 585714, 900, 0.231],
        ['20260501期', '小学', 'B站信息流-非标', 45000, 0.115, 391304, 700, 0.087],
        ['20260501期', '小学', '私域IC', 27100, 0.044, 615909, 600, 0.156],
        ['20260526期', '初中', 'KOC-抖音数学', 82300, 0.055, 1496364, 3800, 0.169],
        ['20260526期', '初中', 'B站信息流-非标', 31100, 0.072, 431944, 1350, 0.121],
        ['20260526期', '初中', '私域IC', 17800, 0.028, 635714, 920, 0.203],
        ['20260526期', '高中', 'KOC-抖音数学', 62100, 0.039, 1592308, 2400, 0.151],
        ['20260526期', '高中', 'B站信息流-非标', 55400, 0.064, 865625, 1950, 0.105],
        ['20260526期', '高中', '私域IC', 89200, 0.048, 1858333, 1680, 0.188],
    ]
    for i, row_data in enumerate(data):
        r = 4 + i
        for c, val in enumerate(row_data, 1):
            ws1.cell(row=r, column=c, value=val)

    # Grand total row
    total_row = 4 + len(data)
    ws1.cell(row=total_row, column=1, value='合计')
    ws1.cell(row=total_row, column=4, value='=SUM(D4:D{})'.format(total_row - 1))
    ws1.cell(row=total_row, column=5, value='=IFERROR(D{}/F{},0)'.format(total_row, total_row))
    ws1.cell(row=total_row, column=6, value='=SUM(F4:F{})'.format(total_row - 1))
    ws1.cell(row=total_row, column=7, value='=SUM(G4:G{})'.format(total_row - 1))
    ws1.cell(row=total_row, column=8, value='=IFERROR(G{}/H{},0)'.format(total_row, total_row))

    # ── Sheet 2: KPI dashboard ──
    ws2 = wb.create_sheet('KPI仪表盘')
    kpi_metrics = [
        ('¥12.8M', '总净收'),
        ('5.6%', '综合退费率'),
        ('18,420', '触达总人数'),
        ('15.8%', '平均转化率'),
        ('9', '活跃渠道数'),
    ]
    for i, (val, label) in enumerate(kpi_metrics):
        apply_kpi_card(ws2, row=3, col=2 + i * 3, value=val, label=label, width=3)

    # ── Sheet 3: Category-color coded channel breakdown ──
    ws3 = wb.create_sheet('渠道分类对比')
    ch_headers = ['渠道分类', '渠道', '净收金额', '退费金额', '退费率']
    for c, h in enumerate(ch_headers, 1):
        ws3.cell(row=1, column=c, value=h)

    ch_data = [
        ['抖音系', 'KOC-抖音数学', 5250000, 241400, 0.046],
        ['抖音系', '抖音-信息流', 3200000, 192000, 0.060],
        ['抖音系', '抖音-私域', 1800000, 72000, 0.040],
        ['B站系', 'B站信息流-非标', 2100000, 168000, 0.080],
        ['B站系', 'B站-品牌专区', 950000, 65550, 0.069],
        ['B站系', 'B站-私域IC', 780000, 54600, 0.070],
        ['私域系', '私域IC-学科', 1650000, 85800, 0.052],
        ['私域系', '私域-社群', 890000, 53400, 0.060],
        ['私域系', '私域-公众号', 430000, 30100, 0.070],
    ]
    for i, row_data in enumerate(ch_data):
        r = 2 + i
        for c, val in enumerate(row_data, 1):
            ws3.cell(row=r, column=c, value=val)

    # ── Sheet 4: Monthly trend (perfect for data bars & sparklines) ──
    ws4 = wb.create_sheet('月度趋势')
    mh = ['月份', '净收(万元)', '退费(万元)', '退费率', '触达人数', '转化人数', '转化率']
    for c, h in enumerate(mh, 1):
        ws4.cell(row=1, column=c, value=h)

    months_data = [
        ['2026-01', 1050, 63, 0.060, 4500, 680, 0.151],
        ['2026-02', 980, 72, 0.073, 4200, 590, 0.140],
        ['2026-03', 1200, 58, 0.048, 5100, 780, 0.153],
        ['2026-04', 1350, 65, 0.048, 5600, 850, 0.152],
        ['2026-05', 1420, 71, 0.050, 5800, 910, 0.157],
        ['2026-06', 1580, 68, 0.043, 6200, 1020, 0.165],
        ['2026-07', 1280, 85, 0.066, 5400, 800, 0.148],
    ]
    for i, row_data in enumerate(months_data):
        r = 2 + i
        for c, val in enumerate(row_data, 1):
            ws4.cell(row=r, column=c, value=val)

    return wb


def test_all_functions():
    wb = make_test_workbook()
    ws1 = wb['透视分析']
    ws2 = wb['KPI仪表盘']
    ws3 = wb['渠道分类对比']
    ws4 = wb['月度趋势']

    errors = []
    passes = []

    def check(condition, name):
        if condition:
            passes.append(name)
        else:
            errors.append(name)

    # ── Phase 1 tests ──
    print('=== Phase 1: Core Auto-Beautification ===')

    # Test: apply_title_banner (on a fresh sheet to avoid overwriting data)
    ws_title = wb.create_sheet('Title Test')
    ws_title.cell(row=3, column=1, value='Col1')
    ws_title.cell(row=3, column=2, value='Col2')
    apply_title_banner(ws_title, 'Monthly Refund & Conversion Trend',
                       'Period: 2026-01 ~ 2026-07 | Unit: 万元 | Source: CRM Full-link',
                       max_col=7, title_row=1)
    check(ws_title.cell(row=1, column=1).font.bold, 'title_banner: title bold')
    check(ws_title.cell(row=1, column=1).font.size >= 14, 'title_banner: large font')
    title_fill = str(ws_title.cell(row=1, column=1).fill.start_color.rgb or '')
    check('1F4E78' in title_fill, f'title_banner: dark blue fill (got {title_fill})')

    # Test: apply_header_style (ws4 headers at row 1, data from row 2)
    apply_header_style(ws4, row=1, max_col=7, freeze=True)
    check(ws4.cell(row=1, column=1).font.bold, 'header_style: bold')
    check('FFFFFF' in str(ws4.cell(row=1, column=1).font.color.rgb or '').upper(),
          'header_style: white text')
    check(ws4.freeze_panes is not None, 'header_style: freeze panes set')
    print(f'  Passes: {len(passes)}, Errors: {len(errors)}')

    # Test: apply_auto_fit_columns
    apply_auto_fit_columns(ws4, max_col=7)
    check(ws4.column_dimensions['A'].width >= 8, 'auto_fit: column A has width')
    check(ws4.column_dimensions['B'].width >= 8, 'auto_fit: column B has width')

    # Test: apply_banded_rows (data rows 2-8, headers at row 1)
    apply_banded_rows(ws4, start_row=2, end_row=8, max_col=7)
    r2_fill = str(ws4.cell(row=2, column=1).fill.start_color.rgb)
    r3_fill = str(ws4.cell(row=3, column=1).fill.start_color.rgb)
    check(r2_fill != r3_fill, f'banded_rows: alternating ({r2_fill} vs {r3_fill})')

    # Test: apply_number_format (headers at row 1)
    fmt_applied = apply_number_format(ws4, start_row=2, end_row=8, header_row=1)
    check(len(fmt_applied) > 0, f'number_format: applied to {len(fmt_applied)} cols: {fmt_applied}')

    # Test: apply_border_grid
    apply_border_grid(ws4, start_row=2, end_row=8, max_col=7)
    b = ws4.cell(row=3, column=3).border
    has_border = (b.left.style is not None or b.bottom.style is not None
                  or b.right.style is not None or b.top.style is not None)
    check(has_border, 'border_grid: cells have borders')

    # ── Phase 2 tests ──
    print(f'\n=== Phase 2: Conditional Formatting ===')
    phase2_start = len(passes)

    # Test: apply_data_bars
    apply_data_bars(ws4, 'B2:B8', color='5B9BD5')
    check(len(ws4.conditional_formatting) > 0, 'data_bars: CF rules added')
    rule_type = None
    for cf in ws4.conditional_formatting:
        for rule in cf.rules:
            if rule.dataBar:
                rule_type = 'dataBar'
    check(rule_type == 'dataBar', f'data_bars: rule type = {rule_type}')

    # Test: apply_color_scale on percentages
    apply_color_scale(ws4, 'D2:D8', scheme='red_white_green')
    has_cs = False
    for cf in ws4.conditional_formatting:
        for rule in cf.rules:
            if rule.colorScale:
                has_cs = True
    check(has_cs, 'color_scale: colorScale rule present')
    print(f'  Passes since Phase 1: {len(passes) - phase2_start}')

    # ── Phase 3 tests: Pivot & Category ──
    print(f'\n=== Phase 3: Pivot Table & Category Beautification ===')
    phase3_start = len(passes)

    # Test: apply_pivot_style
    apply_pivot_style(
        ws1, data_start_row=4, data_end_row=19, max_col=8,
        row_label_cols=[1, 2, 3],
        value_cols=[4, 5, 6, 7, 8],
        has_total_row=True,
    )
    # Check row labels are left-aligned
    check(ws1.cell(row=5, column=1).alignment.horizontal == 'left',
          'pivot_style: row labels left-aligned')
    # Check value columns are right-aligned
    check(ws1.cell(row=5, column=4).alignment.horizontal == 'right',
          'pivot_style: value columns right-aligned')

    # Test: apply_sort_indicator
    apply_sort_indicator(ws1, 'D', direction='desc', header_row=3)
    header_val = str(ws1.cell(row=3, column=4).value or '')
    check('▼' in header_val, f'sort_indicator: arrow present in "{header_val}"')

    # Test: apply_category_colors
    apply_category_colors(
        ws3, 'A',
        {'抖音系': 'E3F2FD', 'B站系': 'FCE4EC', '私域系': 'E8F5E9'},
        start_row=2, end_row=10, max_col=5,
    )
    total_rules = sum(len(cf.rules) for cf in ws3.conditional_formatting)
    check(total_rules >= 3,
          f'category_colors: at least 3 CF rules across all ranges (got {total_rules})')

    # Test: apply_section_header
    apply_section_header(ws3, row=12, max_col=5, text='按渠道汇总', color='accent_green')
    check(ws3.cell(row=12, column=1).font.bold, 'section_header: bold')
    check('0F766E' in str(ws3.cell(row=12, column=1).fill.start_color.rgb),
          'section_header: green fill')

    # Test: apply_subtotal_row + apply_grand_total_row
    # Add a subtotal row and grand total
    ws3.cell(row=20, column=1, value='合计')
    ws3.cell(row=20, column=3, value=12800000)
    apply_grand_total_row(ws3, row=20, max_col=5)
    check(ws3.cell(row=20, column=1).font.bold, 'grand_total: bold')
    check('DCEAF5' in str(ws3.cell(row=20, column=1).fill.start_color.rgb).upper(),
          'grand_total: light blue bg')

    # Test: detect helpers
    start_r, end_r, max_c = detect_data_range(ws4, header_row=3)
    check(start_r == 4, f'detect_data_range: start={start_r}')
    check(end_r >= 8, f'detect_data_range: end={end_r}')
    check(max_c == 7, f'detect_data_range: max_col={max_c}')

    hdr = detect_header_row(ws3)
    check(hdr == 1, f'detect_header_row: found row {hdr}')

    print(f'  Passes since Phase 2: {len(passes) - phase3_start}')

    # ── Test: auto_style_sheet (comprehensive one-shot) ──
    print(f'\n=== auto_style_sheet ===')
    auto_start = len(passes)

    ws5 = wb.create_sheet('Auto Styled')
    h5 = ['渠道', '月份', '触达', '成交', '转化率', '净收']
    for c, h in enumerate(h5, 1):
        ws5.cell(row=1, column=c, value=h)
    auto_data = [
        ['KOC', '2026-07', 1200, 180, 0.15, 450000],
        ['信息流', '2026-07', 800, 72, 0.09, 180000],
        ['私域', '2026-07', 500, 95, 0.19, 237500],
    ]
    for i, rd in enumerate(auto_data):
        for c, v in enumerate(rd, 1):
            ws5.cell(row=2 + i, column=c, value=v)

    auto_style_sheet(
        ws5,
        title='渠道转化对比',
        subtitle='2026年7月 | 按渠道维度汇总',
        header_row=1,
        value_cols=[3, 4, 5, 6],
        category_col='A',
        category_map=['KOC', '信息流', '私域'],
        add_color_scale=True,
        add_data_bars=True,
    )
    # After auto_style_sheet with title, rows are shifted by 2
    check(ws5.cell(row=1, column=1).font.bold, 'auto_style: title exists')
    check(ws5.cell(row=3, column=1).font.bold, 'auto_style: header styled')
    check(ws5.freeze_panes is not None, 'auto_style: frozen panes')

    print(f'  Passes: {len(passes) - auto_start}')

    # ── Summary ──
    print(f'\n{"="*60}')
    print(f'TOTAL: {len(passes)} passed, {len(errors)} failed')
    print(f'{"="*60}')

    if errors:
        print('\nFAILURES:')
        for e in errors:
            print(f'  [FAIL] {e}')
    else:
        print('\n[OK] All tests passed!')

    # Save test file
    wb.save(TEST_FILE)
    print(f'\nTest workbook saved to: {TEST_FILE}')
    return len(errors) == 0


if __name__ == '__main__':
    success = test_all_functions()
    sys.exit(0 if success else 1)
