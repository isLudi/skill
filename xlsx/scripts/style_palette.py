"""
Unified color palette, font, border, and alignment presets for xlsx beautification.
All functions in style_apply.py reference these named presets rather than hardcoded values.

Usage:
    from style_palette import PALETTE, FONTS, FILLS, BORDERS, ALIGNMENTS, NUMBER_FORMATS
"""

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers

# ── Core Color Palette ─────────────────────────────────────────────
PALETTE = {
    # Primary — headers, titles, main structural elements
    'primary_dark': '1F4E78',
    'primary_text': 'FFFFFF',
    'secondary_dark': '294E73',
    'secondary_text': 'FFFFFF',

    # Accent — section dividers, interactive highlights
    'accent_blue': '5B9BD5',
    'accent_green': '0F766E',
    'accent_orange': 'ED7D31',
    'accent_red': 'C00000',
    'accent_teal': '0F766E',

    # Semantic
    'positive': '0F766E',
    'negative': 'C00000',
    'warning': 'ED7D31',
    'info': '5B9BD5',

    # Neutral scale (light → dark)
    'neutral_50': 'FAFBFC',
    'neutral_100': 'F3F6F8',
    'neutral_200': 'E5E7EB',
    'neutral_300': 'D1D5DB',
    'neutral_400': '9CA3AF',
    'neutral_500': '5B6573',
    'neutral_600': '374151',
    'neutral_700': '1F2937',
    'neutral_800': '111827',

    # Backgrounds
    'kpi_bg': 'DCEAF5',
    'metadata_bg': 'F3F6F8',
    'alt_row_1': 'FFFFFF',
    'alt_row_2': 'F6F9FB',
    'section_bg': '5B9BD5',
    'subtotal_bg': 'EBF3FA',
    'grand_total_bg': 'DCEAF5',
    'sort_highlight_bg': 'F0F4FA',

    # Borders
    'border_light': 'D1D5DB',
    'border_medium': '9CA3AF',
    'border_header': '1F4E78',
}

# ── Category Color Sequence ─────────────────────────────────────────
# Used by apply_category_colors() for auto-assignment when no explicit mapping given
CATEGORY_PALETTE = [
    'E3F2FD',  # light blue
    'E8F5E9',  # light green
    'FFF3E0',  # light orange
    'FCE4EC',  # light pink
    'F3E5F5',  # light purple
    'E0F7FA',  # light cyan
    'FFF9C4',  # light yellow
    'EFEBE9',  # light brown
    'E8EAF6',  # light indigo
    'F1F8E9',  # light lime
]

# ── Font Presets ────────────────────────────────────────────────────
FONTS = {
    'title': Font(
        name='Arial', size=16, bold=True,
        color=PALETTE['primary_text'],
    ),
    'subtitle': Font(
        name='Arial', size=10, bold=False,
        color=PALETTE['neutral_500'],
    ),
    'header': Font(
        name='Arial', size=10, bold=True,
        color=PALETTE['primary_text'],
    ),
    'body': Font(
        name='Arial', size=10, bold=False,
        color=PALETTE['neutral_800'],
    ),
    'body_bold': Font(
        name='Arial', size=10, bold=True,
        color=PALETTE['neutral_800'],
    ),
    'kpi_value': Font(
        name='Arial', size=17, bold=True,
        color=PALETTE['neutral_800'],
    ),
    'kpi_label': Font(
        name='Arial', size=9, bold=False,
        color=PALETTE['neutral_500'],
    ),
    'section': Font(
        name='Arial', size=11, bold=True,
        color=PALETTE['primary_text'],
    ),
    'metadata': Font(
        name='Arial', size=10, bold=False,
        color=PALETTE['neutral_500'],
    ),
    'link': Font(
        name='Arial', size=10, bold=False,
        color='0563C1', underline='single',
    ),
    'annotation': Font(
        name='Arial', size=9, bold=False, italic=True,
        color=PALETTE['neutral_500'],
    ),
}


def _make_font(base_name, **overrides):
    """Clone a preset font with overrides. Returns a new Font object."""
    base = FONTS.get(base_name, FONTS['body'])
    kwargs = {
        'name': base.name, 'size': base.size, 'bold': base.bold,
        'italic': base.italic, 'underline': base.underline,
        'color': base.color,
    }
    kwargs.update(overrides)
    return Font(**kwargs)


# ── Fill Presets ────────────────────────────────────────────────────
def _fill(hex_color):
    """Create a solid PatternFill from a hex color string (without alpha prefix)."""
    c = hex_color.lstrip('#')
    if len(c) == 6:
        c = 'FF' + c
    return PatternFill(start_color=c, end_color=c, fill_type='solid')


FILLS = {
    'header': _fill(PALETTE['primary_dark']),
    'header_secondary': _fill(PALETTE['secondary_dark']),
    'title': _fill(PALETTE['primary_dark']),
    'kpi': _fill(PALETTE['kpi_bg']),
    'metadata': _fill(PALETTE['metadata_bg']),
    'alt_1': _fill(PALETTE['alt_row_1']),
    'alt_2': _fill(PALETTE['alt_row_2']),
    'section': _fill(PALETTE['section_bg']),
    'subtotal': _fill(PALETTE['subtotal_bg']),
    'grand_total': _fill(PALETTE['grand_total_bg']),
    'sort_highlight': _fill(PALETTE['sort_highlight_bg']),
    'positive': _fill('E2F0D9'),
    'negative': _fill('FCE4EC'),
    'warning': _fill('FFF4E5'),
    'info': _fill('E3F2FD'),
    'green_section': _fill(PALETTE['accent_green']),
    'none': PatternFill(fill_type=None),
}

# ── Border Presets ──────────────────────────────────────────────────
SIDE_THIN = Side(style='thin', color=PALETTE['border_light'])
SIDE_MEDIUM = Side(style='medium', color=PALETTE['border_medium'])
SIDE_HEADER_BOTTOM = Side(style='medium', color=PALETTE['border_header'])
SIDE_THIN_HEADER = Side(style='thin', color=PALETTE['border_header'])
SIDE_DOUBLE = Side(style='double', color=PALETTE['neutral_700'])
SIDE_NONE = Side(style=None)

BORDERS = {
    'grid': Border(
        left=SIDE_THIN, right=SIDE_THIN,
        top=SIDE_THIN, bottom=SIDE_THIN,
    ),
    'header_row': Border(
        left=SIDE_THIN_HEADER, right=SIDE_THIN_HEADER,
        top=SIDE_MEDIUM, bottom=SIDE_MEDIUM,
    ),
    'header_bottom': Border(
        bottom=SIDE_HEADER_BOTTOM,
    ),
    'total_top': Border(
        top=SIDE_MEDIUM,
    ),
    'total_double_bottom': Border(
        top=SIDE_MEDIUM,
        bottom=SIDE_DOUBLE,
    ),
    'none': Border(
        left=SIDE_NONE, right=SIDE_NONE,
        top=SIDE_NONE, bottom=SIDE_NONE,
    ),
    'outline': Border(
        left=SIDE_MEDIUM, right=SIDE_MEDIUM,
        top=SIDE_MEDIUM, bottom=SIDE_MEDIUM,
    ),
}


def make_border(
    left=None, right=None, top=None, bottom=None,
    color=None, style='thin',
):
    """Create a custom Border. Pass a Side or None for each edge."""
    def _side(s):
        if s is None:
            return SIDE_NONE
        if isinstance(s, Side):
            return s
        c = color or PALETTE['border_light']
        return Side(style=style, color=c)

    return Border(
        left=_side(left), right=_side(right),
        top=_side(top), bottom=_side(bottom),
    )


# ── Alignment Presets ───────────────────────────────────────────────
ALIGNMENTS = {
    'center': Alignment(horizontal='center', vertical='center', wrap_text=False),
    'center_wrap': Alignment(horizontal='center', vertical='center', wrap_text=True),
    'left': Alignment(horizontal='left', vertical='center', wrap_text=False),
    'left_wrap': Alignment(horizontal='left', vertical='center', wrap_text=True),
    'right': Alignment(horizontal='right', vertical='center', wrap_text=False),
    'right_wrap': Alignment(horizontal='right', vertical='center', wrap_text=True),
    'top_left': Alignment(horizontal='left', vertical='top', wrap_text=True),
}


def make_alignment(horizontal='center', vertical='center', wrap_text=False):
    return Alignment(horizontal=horizontal, vertical=vertical, wrap_text=wrap_text)


# ── Number Format Presets ───────────────────────────────────────────
NUMBER_FORMATS = {
    'integer': '#,##0',
    'integer_zeros_dash': '#,##0;(#,##0);-',
    'decimal_1': '#,##0.0',
    'decimal_2': '#,##0.00',
    'percentage_1': '0.0%',
    'percentage_2': '0.00%',
    'currency': '#,##0',
    'currency_decimals': '#,##0.00',
    'multiple': '0.0x',
    'date': 'yyyy-mm-dd',
    'datetime': 'yyyy-mm-dd hh:mm',
    'text': '@',
    'scientific': '0.00E+00',
}

# ── Number Format Auto-Detection ────────────────────────────────────
# Keywords in column headers → number format. First match wins.
NUMBER_FORMAT_RULES = [
    # (keywords_tuple, format_key)
    (('率', '占比', '百分比', '转化', '完成', '达成', '渗透', '留存', '退费', '退款'), 'percentage_1'),
    (('金额', '净收', 'GMV', '流水', '收入', '价', '成本', '花费', '支付', '实收'), 'integer'),
    (('人数', '人头', '单数', '数量', '条数', '次数', '课时', '件数', '订单'), 'integer'),
    (('单价', '均价', '人均', '客单'), 'decimal_2'),
    (('倍数', '倍率', '系数'), 'multiple'),
    (('日期', '时间', '天'), 'date'),
]
