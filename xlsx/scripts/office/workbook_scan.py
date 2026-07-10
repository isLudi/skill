"""Shared workbook formula and cached-error scanning."""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook


EXCEL_ERRORS = (
    "#VALUE!",
    "#DIV/0!",
    "#REF!",
    "#NAME?",
    "#NULL!",
    "#NUM!",
    "#N/A",
)


def scan_workbook(filename: str | Path) -> dict[str, Any]:
    """Count formulas and cached Excel errors after a calculation backend saves."""

    path = Path(filename)
    error_details = {error: [] for error in EXCEL_ERRORS}

    values_wb = load_workbook(path, data_only=True)
    try:
        for sheet_name in values_wb.sheetnames:
            worksheet = values_wb[sheet_name]
            for row in worksheet.iter_rows():
                for cell in row:
                    if cell.value is None:
                        continue
                    value = str(cell.value) if cell.data_type == "e" else cell.value
                    if not isinstance(value, str):
                        continue
                    for error in EXCEL_ERRORS:
                        if error in value:
                            error_details[error].append(f"{sheet_name}!{cell.coordinate}")
                            break
    finally:
        values_wb.close()

    formula_count = 0
    formulas_wb = load_workbook(path, data_only=False)
    try:
        for worksheet in formulas_wb.worksheets:
            for row in worksheet.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and cell.value.startswith("="):
                        formula_count += 1
    finally:
        formulas_wb.close()

    total_errors = sum(len(locations) for locations in error_details.values())
    summary = {
        error: {"count": len(locations), "locations": locations[:20]}
        for error, locations in error_details.items()
        if locations
    }
    return {
        "status": "success" if total_errors == 0 else "errors_found",
        "total_errors": total_errors,
        "error_summary": summary,
        "total_formulas": formula_count,
    }
