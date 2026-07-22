from __future__ import annotations

import argparse
import copy
import hashlib
import json
import os
import re
import shutil
import subprocess
import sys
import time
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


SKILL_ROOT = Path(__file__).resolve().parents[1]
SKILLS_ROOT = SKILL_ROOT.parent
CODEX_ROOT = SKILLS_ROOT.parent
DEFAULT_REGISTRY = SKILL_ROOT / "references" / "workflow_registry.json"
DEFAULT_RUNTIME_ROOT = CODEX_ROOT / "runtime" / "sync-qingcheng-temp-tables"
OPERATOR_ROOT = SKILLS_ROOT / "usql-web-query-operator"
OPERATOR_SCRIPT = OPERATOR_ROOT / "scripts" / "usql_web_query.py"
OPERATOR_SCRIPTS = OPERATOR_ROOT / "scripts"
RECALC_SCRIPT = SKILLS_ROOT / "xlsx" / "scripts" / "recalc.py"
ERROR_VALUES = {"#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#NUM!", "#NULL!"}


class WorkflowError(RuntimeError):
    pass


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def canonical_json(value: Any) -> bytes:
    return json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":")).encode("utf-8")


def artifact_hash(value: dict[str, Any], hash_field: str) -> str:
    payload = dict(value)
    payload.pop(hash_field, None)
    return hashlib.sha256(canonical_json(payload)).hexdigest()


def write_artifact(path: Path, value: dict[str, Any], hash_field: str) -> str:
    value[hash_field] = artifact_hash(value, hash_field)
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(value, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    return value[hash_field]


def load_artifact(path: Path, hash_field: str, expected_hash: str | None = None) -> dict[str, Any]:
    value = json.loads(path.read_text(encoding="utf-8"))
    actual_hash = artifact_hash(value, hash_field)
    if value.get(hash_field) != actual_hash:
        raise WorkflowError(f"Artifact hash mismatch: {path}")
    if expected_hash and expected_hash != actual_hash:
        raise WorkflowError(f"Expected {hash_field}={expected_hash}, got {actual_hash}")
    return value


def load_registry(path: Path) -> dict[str, Any]:
    registry = json.loads(path.read_text(encoding="utf-8"))
    ids = [family["id"] for family in registry.get("families", [])]
    if len(ids) != len(set(ids)):
        raise WorkflowError("Workflow registry contains duplicate family ids.")
    upload_order = registry.get("upload_order") or []
    if len(upload_order) != len(set(upload_order)) or set(upload_order) != set(ids):
        raise WorkflowError("Workflow registry upload_order must contain every family id exactly once.")
    return registry


def family_map(registry: dict[str, Any]) -> dict[str, dict[str, Any]]:
    return {family["id"]: family for family in registry["families"]}


def _command_argv(executable: str, args: list[str]) -> list[str]:
    if Path(executable).suffix.casefold() in {".cmd", ".bat"}:
        return [os.environ.get("COMSPEC", "cmd.exe"), "/d", "/c", executable, *args]
    return [executable, *args]


def run_json_command(
    executable: str,
    args: list[str],
    *,
    cwd: Path | None = None,
    timeout: int = 120,
) -> dict[str, Any]:
    env = os.environ.copy()
    env["PYTHONIOENCODING"] = "utf-8"
    env["PYTHONUTF8"] = "1"
    env["LARKSUITE_CLI_NO_UPDATE_NOTIFIER"] = "1"
    env["LARKSUITE_CLI_NO_SKILLS_NOTIFIER"] = "1"
    completed = subprocess.run(
        _command_argv(executable, args),
        cwd=str(cwd) if cwd else None,
        env=env,
        capture_output=True,
        text=True,
        encoding="utf-8",
        errors="replace",
        timeout=timeout,
        check=False,
    )
    output = completed.stdout.strip() or completed.stderr.strip()
    try:
        payload = json.loads(output)
    except json.JSONDecodeError as exc:
        raise WorkflowError(
            f"Command did not return JSON (exit={completed.returncode}): {' '.join(args[:3])}"
        ) from exc
    if completed.returncode != 0 or payload.get("ok") is False:
        error = payload.get("error") or {}
        message = error.get("message") or payload.get("message") or "command failed"
        raise WorkflowError(f"{message} (exit={completed.returncode})")
    return payload


def resolve_lark_cli() -> str:
    executable = shutil.which("lark-cli.cmd") or shutil.which("lark-cli")
    if not executable:
        raise WorkflowError("lark-cli is not available on PATH.")
    # `shutil.which` can return a path relative to the process working directory
    # (for example, `.\lark-cli.cmd`). Downloads deliberately run with a separate
    # cwd, so preserve the executable location as an absolute path first.
    return str(Path(executable).resolve())


def discover_live_messages(
    registry: dict[str, Any],
    explicit_message_ids: set[str] | None = None,
) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    cli = resolve_lark_cli()
    chat_cfg = registry["chat"]
    expected_id = chat_cfg.get("expected_chat_id")
    if expected_id:
        chat = {"name": chat_cfg["name"], "chat_id": expected_id}
    else:
        search = run_json_command(
            cli,
            [
                "im",
                "+chat-search",
                "--query",
                chat_cfg["name"],
                "--as",
                "user",
                "--page-size",
                "20",
                "--format",
                "json",
            ],
            timeout=60,
        )
        chats = search.get("data", {}).get("chats", [])
        exact = [item for item in chats if item.get("name") == chat_cfg["name"]]
        if len(exact) != 1:
            raise WorkflowError(f"Expected exactly one visible chat named {chat_cfg['name']}; found {len(exact)}.")
        chat = exact[0]
    try:
        result = run_json_command(
            cli,
            [
                "im",
                "+messages-search",
                "--chat-id",
                chat["chat_id"],
                "--sender",
                chat_cfg["sender_open_id"],
                "--include-attachment-type",
                "file",
                "--page-size",
                "50",
                "--page-all",
                "--no-reactions",
                "--as",
                "user",
                "--format",
                "json",
            ],
            timeout=120,
        )
    except WorkflowError:
        if not expected_id:
            raise
        result = run_json_command(
            cli,
            [
                "im",
                "+chat-messages-list",
                "--chat-id",
                chat["chat_id"],
                "--sort",
                "desc",
                "--page-size",
                "50",
                "--page-all",
                "--no-reactions",
                "--as",
                "bot",
                "--format",
                "json",
            ],
            timeout=180,
        )
    raw_messages = list(result.get("data", {}).get("messages", []))
    explicit_message_ids = explicit_message_ids or set()
    if explicit_message_ids:
        exact = run_json_command(
            cli,
            [
                "im",
                "+messages-mget",
                "--message-ids",
                ",".join(sorted(explicit_message_ids)),
                "--no-reactions",
                "--as",
                "bot",
                "--format",
                "json",
            ],
            timeout=60,
        )
        existing_ids = {message.get("message_id") for message in raw_messages}
        raw_messages.extend(
            message
            for message in exact.get("data", {}).get("messages", [])
            if message.get("message_id") not in existing_ids
        )
    normalized = []
    pattern = re.compile(r'<file\s+key="([^"]+)"\s+name="([^"]+)"\s*/>')
    for message in raw_messages:
        match = pattern.search(str(message.get("content", "")))
        sender = message.get("sender") or {}
        if not match or message.get("deleted"):
            continue
        if message.get("chat_id") and message.get("chat_id") != chat["chat_id"]:
            continue
        if sender.get("id") != chat_cfg["sender_open_id"] or sender.get("name") != chat_cfg["sender_name"]:
            continue
        normalized.append(
            {
                "message_id": message.get("message_id"),
                "file_key": match.group(1),
                "file_name": match.group(2),
                "create_time": message.get("create_time"),
                "message_position": str(message.get("message_position") or ""),
                "message_app_link": message.get("message_app_link"),
                "sender_id": sender.get("id"),
                "sender_name": sender.get("name"),
            }
        )
    return chat, normalized


def classify_messages(
    registry: dict[str, Any], messages: list[dict[str, Any]]
) -> tuple[dict[str, list[dict[str, Any]]], list[dict[str, Any]]]:
    classified: dict[str, list[dict[str, Any]]] = defaultdict(list)
    deferred_patterns = [re.compile(pattern) for pattern in registry.get("deferred_filename_patterns", [])]
    unclassified = []
    for message in messages:
        matches = []
        for family in registry["families"]:
            if any(re.fullmatch(pattern, message["file_name"]) for pattern in family["source_filename_patterns"]):
                matches.append(family["id"])
        if len(matches) > 1:
            raise WorkflowError(f"File {message['file_name']} matches multiple families: {matches}")
        if matches:
            classified[matches[0]].append(message)
        else:
            item = dict(message)
            item["classification"] = (
                "deferred" if any(pattern.fullmatch(message["file_name"]) for pattern in deferred_patterns) else "excluded"
            )
            unclassified.append(item)
    return classified, unclassified


def message_sort_key(message: dict[str, Any]) -> tuple[str, int, str]:
    try:
        position = int(message.get("message_position") or 0)
    except (TypeError, ValueError):
        position = 0
    return str(message.get("create_time") or ""), position, str(message.get("message_id") or "")


def parse_datetime(value: str) -> datetime:
    text = str(value).strip()
    if not text:
        raise WorkflowError("Datetime value cannot be blank.")
    if re.fullmatch(r"\d+(?:\.\d+)?", text):
        timestamp = float(text)
        if timestamp > 10_000_000_000:
            timestamp /= 1000
        return datetime.fromtimestamp(timestamp, tz=timezone.utc)
    normalized = text[:-1] + "+00:00" if text.endswith("Z") else text
    try:
        parsed = datetime.fromisoformat(normalized)
    except ValueError as exc:
        raise WorkflowError(f"Invalid datetime value: {value}") from exc
    if parsed.tzinfo is None:
        parsed = parsed.astimezone()
    return parsed.astimezone(timezone.utc)


def build_selection_spec(
    registry: dict[str, Any],
    family_ids: list[str] | None = None,
    after: str | None = None,
    explicit_message_specs: list[str] | None = None,
) -> dict[str, Any]:
    known = family_map(registry)
    requested = list(family_ids or registry["upload_order"])
    if not requested:
        raise WorkflowError("At least one workbook family must be selected.")
    if len(requested) != len(set(requested)):
        raise WorkflowError(f"Workbook family selection contains duplicates: {requested}")
    unknown = [family_id for family_id in requested if family_id not in known]
    if unknown:
        raise WorkflowError(f"Unknown workbook family ids: {unknown}")
    requested_set = set(requested)
    ordered = [family_id for family_id in registry["upload_order"] if family_id in requested_set]
    explicit: dict[str, str] = {}
    for raw_spec in explicit_message_specs or []:
        family_id, separator, message_id = raw_spec.partition("=")
        family_id = family_id.strip()
        message_id = message_id.strip()
        if not separator or family_id not in known or not re.fullmatch(r"om_[A-Za-z0-9]+", message_id):
            raise WorkflowError(
                f"Invalid --message-id value {raw_spec!r}; expected <family_id>=<om_message_id>."
            )
        if family_id not in requested_set:
            raise WorkflowError(f"Explicit message family is not selected by --family: {family_id}")
        if family_id in explicit:
            raise WorkflowError(f"Only one explicit message may be bound to family {family_id}.")
        explicit[family_id] = message_id
    after_iso = parse_datetime(after).isoformat() if after else None
    return {
        "family_ids": ordered,
        "after": after_iso,
        "explicit_message_ids": explicit,
        "selection_modes": {
            family_id: "explicit_message" if family_id in explicit else "latest_matching"
            for family_id in ordered
        },
    }


def select_messages(
    registry: dict[str, Any],
    messages: list[dict[str, Any]],
    selection: dict[str, Any],
) -> tuple[dict[str, dict[str, Any]], dict[str, list[dict[str, Any]]], list[dict[str, Any]]]:
    after = parse_datetime(selection["after"]) if selection.get("after") else None
    eligible = []
    for message in messages:
        if after is not None:
            try:
                created = parse_datetime(str(message.get("create_time") or ""))
            except WorkflowError:
                continue
            if created <= after:
                continue
        eligible.append(message)
    classified, unclassified = classify_messages(registry, eligible)
    selected: dict[str, dict[str, Any]] = {}
    missing = []
    explicit = selection.get("explicit_message_ids") or {}
    for family_id in selection["family_ids"]:
        candidates = classified.get(family_id, [])
        if family_id in explicit:
            candidates = [message for message in candidates if message.get("message_id") == explicit[family_id]]
        if not candidates:
            missing.append(family_id)
        elif len(candidates) > 1 and family_id in explicit:
            raise WorkflowError(f"Explicit message id is not unique in search results: {explicit[family_id]}")
        else:
            selected[family_id] = max(candidates, key=message_sort_key)
    if missing:
        details = {family_id: explicit.get(family_id) for family_id in missing}
        raise WorkflowError(f"No matching file message found for selected families: {details}")
    return selected, classified, unclassified


def select_latest_messages(
    registry: dict[str, Any], messages: list[dict[str, Any]]
) -> tuple[dict[str, dict[str, Any]], dict[str, list[dict[str, Any]]], list[dict[str, Any]]]:
    selection = build_selection_spec(registry)
    return select_messages(registry, messages, selection)


def download_message(message: dict[str, Any], family_id: str, output_dir: Path) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    cli = resolve_lark_cli()
    output_name = f"{family_id}__{message['message_id']}.xlsx"
    payload = run_json_command(
        cli,
        [
            "im",
            "+messages-resources-download",
            "--message-id",
            message["message_id"],
            "--file-key",
            message["file_key"],
            "--type",
            "file",
            "--output",
            f".\\{output_name}",
            "--as",
            "bot",
        ],
        cwd=output_dir,
        timeout=120,
    )
    saved = Path(payload["data"]["saved_path"]).resolve()
    if saved.parent != output_dir.resolve() or not saved.exists():
        raise WorkflowError(f"Downloaded file escaped the plan directory: {saved}")
    return saved


def normalize_cell(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value


def normalized_record(record: dict[str, Any], columns: list[str]) -> tuple[Any, ...]:
    values = []
    for column in columns:
        value = normalize_cell(record.get(column))
        if isinstance(value, datetime):
            value = value.isoformat()
        values.append(value)
    return tuple(values)


def read_records(
    path: Path,
    sheet_name: str,
    target_columns: list[str],
    *,
    aliases: dict[str, str] | None = None,
    constants: dict[str, Any] | None = None,
    data_only: bool,
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    aliases = aliases or {}
    constants = constants or {}
    workbook = load_workbook(path, read_only=False, data_only=data_only, keep_links=True)
    try:
        if sheet_name not in workbook.sheetnames:
            raise WorkflowError(f"Sheet {sheet_name} not found in {path.name}; found {workbook.sheetnames}")
        sheet = workbook[sheet_name]
        raw_headers = [normalize_cell(cell.value) for cell in sheet[1]]
        while raw_headers and raw_headers[-1] in (None, ""):
            raw_headers.pop()
        headers = [aliases.get(str(header), str(header)) if header not in (None, "") else "" for header in raw_headers]
        nonblank_headers = [header for header in headers if header]
        if len(nonblank_headers) != len(set(nonblank_headers)):
            raise WorkflowError(f"Duplicate headers after alias mapping in {path.name}: {headers}")
        missing = [column for column in target_columns if column not in nonblank_headers and column not in constants]
        extras = [header for header in nonblank_headers if header not in target_columns]
        if missing or extras:
            raise WorkflowError(f"Schema mismatch in {path.name}: missing={missing}, extras={extras}, headers={headers}")
        positions = {header: index for index, header in enumerate(headers) if header}
        records = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not any(normalize_cell(value) not in (None, "") for value in row):
                continue
            record = {}
            for column in target_columns:
                if column in constants:
                    record[column] = constants[column]
                else:
                    index = positions[column]
                    record[column] = normalize_cell(row[index] if index < len(row) else None)
            records.append(record)
        metadata = {
            "sheet": sheet_name,
            "headers": headers,
            "row_count": len(records),
            "formula_count": sum(
                1
                for row in sheet.iter_rows(min_row=2)
                for cell in row
                if cell.data_type == "f"
            ),
        }
        return records, metadata
    finally:
        workbook.close()


def _text(value: Any) -> str:
    value = normalize_cell(value)
    return "" if value is None else str(value).strip()


def validate_source_records(family: dict[str, Any], records: list[dict[str, Any]]) -> list[dict[str, Any]]:
    issues: list[dict[str, Any]] = []
    if not records:
        return [{"severity": "error", "rule": "nonempty_source", "message": "Source workbook has no data rows."}]
    key_columns = family["key_columns"]
    groups: dict[tuple[str, ...], list[int]] = defaultdict(list)
    for row_number, record in enumerate(records, start=2):
        key = tuple(_text(record.get(column)) for column in key_columns)
        if any(not value for value in key):
            issues.append(
                {
                    "severity": "error",
                    "rule": "nonblank_key",
                    "message": f"Blank key value at row {row_number} for {key_columns}.",
                    "row": row_number,
                }
            )
        groups[key].append(row_number)
    duplicates = [{"key": key, "rows": rows} for key, rows in groups.items() if all(key) and len(rows) > 1]
    if duplicates:
        issues.append(
            {
                "severity": "error",
                "rule": "unique_key",
                "message": f"Source has duplicate keys for {key_columns}.",
                "count": len(duplicates),
                "examples": duplicates[:20],
            }
        )
    for rule in family.get("validation_rules", []):
        rule_type = rule["type"]
        if rule_type == "slice_format":
            invalid = [
                {"row": index, "value": _text(record.get(rule["column"]))}
                for index, record in enumerate(records, start=2)
                if not re.fullmatch(rule["pattern"], _text(record.get(rule["column"])))
            ]
        elif rule_type == "month_format":
            invalid = [
                {"row": index, "value": _text(record.get(rule["column"]))}
                for index, record in enumerate(records, start=2)
                if not re.fullmatch(r"\d{6}", _text(record.get(rule["column"])))
            ]
        elif rule_type == "lowercase_ascii_prefix":
            invalid = []
            for index, record in enumerate(records, start=2):
                for column in rule["columns"]:
                    value = _text(record.get(column))
                    first = next((char for char in value if char.isascii() and char.isalpha()), "")
                    if first and first != first.lower():
                        invalid.append({"row": index, "column": column, "value": value})
        elif rule_type == "disallow_values":
            disallowed = {_text(value) for value in rule["values"]}
            invalid = [
                {"row": index, "value": _text(record.get(rule["column"]))}
                for index, record in enumerate(records, start=2)
                if _text(record.get(rule["column"])) in disallowed
            ]
        elif rule_type == "required_value":
            invalid = [
                {"row": index, "value": _text(record.get(rule["column"]))}
                for index, record in enumerate(records, start=2)
                if _text(record.get(rule["column"])) != _text(rule["value"])
            ]
        else:
            invalid = [{"rule": rule_type}]
        if invalid:
            issues.append(
                {
                    "severity": "error",
                    "rule": rule_type,
                    "message": f"Source validation failed: {rule_type}.",
                    "count": len(invalid),
                    "examples": invalid[:20],
                }
            )
    return issues


def slice_sort_key(value: Any) -> tuple[int, str]:
    text = _text(value)
    digits = re.sub(r"\D", "", text)
    return (int(digits) if digits else -1, text)


def scope_matches(record: dict[str, Any], family: dict[str, Any]) -> bool:
    scope = family.get("target_scope")
    if not scope:
        return True
    return _text(record.get(scope["column"])) == _text(scope["equals"])


def records_equal(left: list[dict[str, Any]], right: list[dict[str, Any]], columns: list[str]) -> bool:
    return [normalized_record(record, columns) for record in left] == [
        normalized_record(record, columns) for record in right
    ]


def merge_records(
    family: dict[str, Any],
    target_write: list[dict[str, Any]],
    target_effective: list[dict[str, Any]],
    source_write: list[dict[str, Any]],
    source_effective: list[dict[str, Any]],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], dict[str, Any]]:
    if len(target_write) != len(target_effective) or len(source_write) != len(source_effective):
        raise WorkflowError("Formula and effective record streams are not aligned.")
    slice_column = family["slice_column"]
    columns = family["target_columns"]
    target_by_slice: dict[str, list[tuple[dict[str, Any], dict[str, Any]]]] = defaultdict(list)
    source_by_slice: dict[str, list[tuple[dict[str, Any], dict[str, Any]]]] = defaultdict(list)
    for write_record, effective_record in zip(target_write, target_effective):
        target_by_slice[_text(effective_record.get(slice_column))].append((write_record, effective_record))
    for write_record, effective_record in zip(source_write, source_effective):
        source_by_slice[_text(effective_record.get(slice_column))].append((write_record, effective_record))
    source_slices = set(source_by_slice)
    target_slices = set(target_by_slice)
    all_slices = sorted(
        target_slices | source_slices,
        key=slice_sort_key,
        reverse=family.get("slice_order") == "desc",
    )
    merged_write = []
    merged_effective = []
    new_slices = []
    replaced_slices = []
    unchanged_slices = []
    removed_rows = 0
    for slice_value in all_slices:
        target_pairs = target_by_slice.get(slice_value, [])
        source_pairs = source_by_slice.get(slice_value)
        if source_pairs is None:
            output_pairs = target_pairs
        else:
            scoped_target = [pair for pair in target_pairs if scope_matches(pair[1], family)]
            preserved_target = [pair for pair in target_pairs if not scope_matches(pair[1], family)]
            removed_rows += len(scoped_target)
            source_scope_effective = [pair[1] for pair in source_pairs]
            target_scope_effective = [pair[1] for pair in scoped_target]
            if slice_value not in target_slices:
                new_slices.append(slice_value)
            elif Counter(normalized_record(record, columns) for record in source_scope_effective) == Counter(
                normalized_record(record, columns) for record in target_scope_effective
            ):
                unchanged_slices.append(slice_value)
            else:
                replaced_slices.append(slice_value)
            output_pairs = [*source_pairs, *preserved_target]
        merged_write.extend(pair[0] for pair in output_pairs)
        merged_effective.extend(pair[1] for pair in output_pairs)
    changed = not records_equal(merged_effective, target_effective, columns)
    diff = {
        "changed": changed,
        "slice_column": slice_column,
        "source_slices": sorted(source_slices, key=slice_sort_key),
        "new_slices": new_slices,
        "replaced_slices": replaced_slices,
        "unchanged_slices": unchanged_slices,
        "target_rows_before": len(target_effective),
        "source_rows": len(source_effective),
        "scoped_rows_removed": removed_rows,
        "target_rows_after": len(merged_effective),
    }
    return merged_write, merged_effective, diff


def count_formula_values(records: Iterable[dict[str, Any]]) -> int:
    return sum(isinstance(value, str) and value.startswith("=") for record in records for value in record.values())


def rebuild_workbook(
    target_path: Path,
    stage_path: Path,
    family: dict[str, Any],
    records: list[dict[str, Any]],
) -> dict[str, Any]:
    shutil.copy2(target_path, stage_path)
    workbook = load_workbook(stage_path, data_only=False, keep_links=True)
    sheet = workbook[family["target_sheet"]]
    columns = family["target_columns"]
    template_cells = [copy.copy(sheet.cell(row=2, column=index)._style) for index in range(1, len(columns) + 1)]
    template_height = sheet.row_dimensions[2].height
    existing_filter = sheet.auto_filter.ref
    if sheet.max_row > 1:
        sheet.delete_rows(2, sheet.max_row - 1)
    for record in records:
        sheet.append([record.get(column) for column in columns])
        row_number = sheet.max_row
        if template_height is not None:
            sheet.row_dimensions[row_number].height = template_height
        for index, style in enumerate(template_cells, start=1):
            sheet.cell(row=row_number, column=index)._style = copy.copy(style)
    if existing_filter:
        sheet.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{max(sheet.max_row, 1)}"
    calculation = getattr(workbook, "calculation", None)
    if calculation is not None:
        calculation.fullCalcOnLoad = True
        calculation.forceFullCalc = True
        calculation.calcMode = "auto"
    workbook.save(stage_path)
    workbook.close()
    formula_count = count_formula_values(records)
    recalc = None
    if formula_count:
        recalc = recalculate_workbook(stage_path)
    return {"formula_count": formula_count, "recalculation": recalc}


def recalculate_workbook(path: Path) -> dict[str, Any]:
    if not RECALC_SCRIPT.exists():
        raise WorkflowError(f"Spreadsheet recalculation script not found: {RECALC_SCRIPT}")
    payload = run_json_command(sys.executable, [str(RECALC_SCRIPT), str(path), "60"], timeout=120)
    if payload.get("status") not in {"success", "ok"} or payload.get("total_errors", 0):
        raise WorkflowError(f"Excel recalculation failed for {path}: {payload}")
    return payload


def operator_validation(file_path: Path, target_path: Path) -> dict[str, Any]:
    if str(OPERATOR_SCRIPTS) not in sys.path:
        sys.path.insert(0, str(OPERATOR_SCRIPTS))
    from usql_web_query.manual_table_registry import ManualTableRegistry  # noqa: PLC0415
    from usql_web_query.manual_table_validation import validate_manual_table  # noqa: PLC0415

    registry = ManualTableRegistry.load()
    entry = registry.resolve_file(target_path)
    if entry is None:
        raise WorkflowError(f"Operator manual-table registry does not match target: {target_path}")
    return validate_manual_table(file_path, entry)


def validation_regressions(before: dict[str, Any], after: dict[str, Any]) -> list[dict[str, Any]]:
    baseline = {}
    for issue in before.get("issues", []):
        if issue.get("severity") != "error":
            continue
        signature = (issue.get("rule"), issue.get("column"), issue.get("message"))
        baseline[signature] = int(issue.get("count", 1))
    regressions = []
    for issue in after.get("issues", []):
        if issue.get("severity") != "error":
            continue
        signature = (issue.get("rule"), issue.get("column"), issue.get("message"))
        after_count = int(issue.get("count", 1))
        if after_count > baseline.get(signature, 0):
            regressions.append(
                {
                    "rule": issue.get("rule"),
                    "column": issue.get("column"),
                    "before_count": baseline.get(signature, 0),
                    "after_count": after_count,
                    "message": issue.get("message"),
                }
            )
    return regressions


def plan_sync(args: argparse.Namespace) -> int:
    registry_path = args.registry.resolve()
    registry = load_registry(registry_path)
    selection = build_selection_spec(
        registry,
        family_ids=args.family,
        after=args.after,
        explicit_message_specs=args.message_id,
    )
    explicit_ids = set(selection.get("explicit_message_ids", {}).values())
    chat, messages = discover_live_messages(registry, explicit_ids)
    selected, classified, unclassified = select_messages(registry, messages, selection)
    run_id = datetime.now().strftime("%Y%m%d-%H%M%S") + f"-{os.getpid()}"
    run_dir = args.runtime_root.resolve() / run_id
    downloads_dir = run_dir / "downloads"
    stages_dir = run_dir / "staged"
    stages_dir.mkdir(parents=True, exist_ok=True)
    blockers = []
    tables = []
    families = family_map(registry)
    for family_id in selection["family_ids"]:
        family = families[family_id]
        message = dict(selected[family_id])
        try:
            source_path = download_message(message, family_id, downloads_dir)
            message["download_path"] = str(source_path)
            message["source_sha256"] = sha256_file(source_path)
            target_path = Path(family["target_workbook"]).resolve()
            if not target_path.exists():
                raise WorkflowError(f"Target workbook does not exist: {target_path}")
            source_write, source_meta = read_records(
                source_path,
                family["source_sheet"],
                family["target_columns"],
                aliases=family.get("column_aliases"),
                constants=family.get("constant_columns"),
                data_only=False,
            )
            source_effective, _ = read_records(
                source_path,
                family["source_sheet"],
                family["target_columns"],
                aliases=family.get("column_aliases"),
                constants=family.get("constant_columns"),
                data_only=True,
            )
            source_issues = validate_source_records(family, source_effective)
            if source_issues:
                raise WorkflowError(f"Source validation failed: {source_issues}")
            target_write, target_meta = read_records(
                target_path,
                family["target_sheet"],
                family["target_columns"],
                data_only=False,
            )
            target_effective, _ = read_records(
                target_path,
                family["target_sheet"],
                family["target_columns"],
                data_only=True,
            )
            merged_write, merged_effective, diff = merge_records(
                family, target_write, target_effective, source_write, source_effective
            )
            before_validation = operator_validation(target_path, target_path)
            stage_path = None
            stage_meta = None
            candidate_path = target_path
            if diff["changed"]:
                stage_path = stages_dir / target_path.name
                stage_meta = rebuild_workbook(target_path, stage_path, family, merged_write)
                staged_effective, _ = read_records(
                    stage_path,
                    family["target_sheet"],
                    family["target_columns"],
                    data_only=True,
                )
                if not records_equal(staged_effective, merged_effective, family["target_columns"]):
                    raise WorkflowError(f"Staged workbook effective values differ from planned values: {family_id}")
                candidate_path = stage_path
            after_validation = operator_validation(candidate_path, target_path)
            regressions = validation_regressions(before_validation, after_validation)
            if family.get("allow_baseline_target_errors"):
                if regressions:
                    raise WorkflowError(f"Target validation regressed: {regressions}")
            elif after_validation.get("error_count", 0):
                raise WorkflowError(f"Target validation failed: {after_validation.get('issues', [])}")
            tables.append(
                {
                    "family_id": family_id,
                    "business_name": family["business_name"],
                    "source_message": message,
                    "source_metadata": source_meta,
                    "source_validation": {"ok": True, "issues": []},
                    "target_path": str(target_path),
                    "target_sheet": family["target_sheet"],
                    "platform_temp_table": family["platform_temp_table"],
                    "target_before_sha256": sha256_file(target_path),
                    "target_after_sha256": sha256_file(candidate_path),
                    "stage_path": str(stage_path) if stage_path else None,
                    "stage_metadata": stage_meta,
                    "diff": diff,
                    "validation_before": before_validation,
                    "validation_after": after_validation,
                    "validation_regressions": regressions,
                    "allow_baseline_target_errors": bool(family.get("allow_baseline_target_errors")),
                }
            )
        except Exception as exc:  # noqa: BLE001
            blockers.append({"family_id": family_id, "message": str(exc), "error_type": type(exc).__name__})
    plan = {
        "schema_version": "1.0.0",
        "artifact_type": "QingchengTempTableSyncPlan",
        "generated_at": datetime.now().astimezone().isoformat(timespec="seconds"),
        "status": "ready" if not blockers and len(tables) == len(selection["family_ids"]) else "blocked",
        "registry_path": str(registry_path),
        "registry_sha256": sha256_file(registry_path),
        "runtime_dir": str(run_dir),
        "chat": {
            "name": chat.get("name"),
            "chat_id": chat.get("chat_id"),
            "sender_name": registry["chat"]["sender_name"],
            "sender_open_id": registry["chat"]["sender_open_id"],
            "attachment_message_count": len(messages),
        },
        "selection": selection,
        "selected_message_ids": {family_id: message["message_id"] for family_id, message in selected.items()},
        "history_counts": {family_id: len(items) for family_id, items in classified.items()},
        "unclassified_files": unclassified,
        "tables": tables,
        "blockers": blockers,
        "production_upload_authorized": False,
    }
    plan_path = run_dir / "sync_plan.json"
    plan_sha = write_artifact(plan_path, plan, "plan_sha256")
    summary = {
        "ok": plan["status"] == "ready",
        "status": plan["status"],
        "plan_path": str(plan_path),
        "plan_sha256": plan_sha,
        "table_count": len(tables),
        "blockers": blockers,
        "tables": [
            {
                "family_id": table["family_id"],
                "source_file": table["source_message"]["file_name"],
                "source_time": table["source_message"]["create_time"],
                "target_path": table["target_path"],
                "platform_temp_table": table["platform_temp_table"],
                "diff": table["diff"],
            }
            for table in tables
        ],
    }
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    return 0 if summary["ok"] else 1


def apply_local(args: argparse.Namespace) -> int:
    if not args.confirm_local_write:
        raise WorkflowError("apply-local requires --confirm-local-write.")
    plan = load_artifact(args.plan.resolve(), "plan_sha256", args.expected_plan_sha256)
    if plan.get("status") != "ready":
        raise WorkflowError("Only a ready plan can be applied locally.")
    registry_path = Path(plan["registry_path"])
    if not registry_path.exists() or sha256_file(registry_path) != plan["registry_sha256"]:
        raise WorkflowError("Workflow registry drifted after planning.")
    registry = load_registry(registry_path)
    selection = plan.get("selection") or build_selection_spec(registry)
    explicit_ids = set(selection.get("explicit_message_ids", {}).values())
    _, current_messages = discover_live_messages(registry, explicit_ids)
    current_selected, _, _ = select_messages(registry, current_messages, selection)
    current_ids = {family_id: message["message_id"] for family_id, message in current_selected.items()}
    if current_ids != plan["selected_message_ids"]:
        raise WorkflowError("Newer matching Feishu files appeared after planning; create a fresh plan.")
    for table in plan["tables"]:
        target = Path(table["target_path"])
        source = Path(table["source_message"]["download_path"])
        if sha256_file(target) != table["target_before_sha256"]:
            raise WorkflowError(f"Target workbook drifted after planning: {target}")
        if sha256_file(source) != table["source_message"]["source_sha256"]:
            raise WorkflowError(f"Downloaded source drifted after planning: {source}")
        if table["stage_path"]:
            stage = Path(table["stage_path"])
            if sha256_file(stage) != table["target_after_sha256"]:
                raise WorkflowError(f"Staged workbook drifted after planning: {stage}")
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backups = []
    changed_tables = [table for table in plan["tables"] if table["diff"]["changed"]]
    try:
        for table in changed_tables:
            target = Path(table["target_path"])
            backup = target.with_name(f"{target.stem}.backup_{timestamp}{target.suffix}")
            if backup.exists():
                raise WorkflowError(f"Backup path already exists: {backup}")
            shutil.copy2(target, backup)
            backups.append({"family_id": table["family_id"], "target": str(target), "backup": str(backup)})
        for table in changed_tables:
            target = Path(table["target_path"])
            stage = Path(table["stage_path"])
            temp_target = target.with_name(f".{target.name}.{os.getpid()}.tmp")
            shutil.copy2(stage, temp_target)
            os.replace(temp_target, target)
            if sha256_file(target) != table["target_after_sha256"]:
                raise WorkflowError(f"Applied target hash mismatch: {target}")
            validation = operator_validation(target, target)
            if table["allow_baseline_target_errors"]:
                regressions = validation_regressions(table["validation_before"], validation)
                if regressions:
                    raise WorkflowError(f"Applied validation regressed for {target}: {regressions}")
            elif validation.get("error_count", 0):
                raise WorkflowError(f"Applied validation failed for {target}: {validation.get('issues', [])}")
    except Exception:
        for item in reversed(backups):
            shutil.copy2(Path(item["backup"]), Path(item["target"]))
        raise
    receipt = {
        "schema_version": "1.0.0",
        "artifact_type": "QingchengTempTableLocalApplyReceipt",
        "created_at": datetime.now().astimezone().isoformat(timespec="seconds"),
        "status": "success" if changed_tables else "success_no_local_changes",
        "plan_path": str(args.plan.resolve()),
        "plan_sha256": plan["plan_sha256"],
        "backups": backups,
        "tables": [
            {
                "family_id": table["family_id"],
                "target_path": table["target_path"],
                "changed": table["diff"]["changed"],
                "current_sha256": sha256_file(Path(table["target_path"])),
                "expected_sha256": table["target_after_sha256"],
            }
            for table in plan["tables"]
        ],
    }
    receipt_path = Path(plan["runtime_dir"]) / "local_apply_receipt.json"
    receipt_sha = write_artifact(receipt_path, receipt, "receipt_sha256")
    print(
        json.dumps(
            {
                "ok": True,
                "status": receipt["status"],
                "receipt_path": str(receipt_path),
                "receipt_sha256": receipt_sha,
                "changed_table_count": len(changed_tables),
                "backups": backups,
            },
            ensure_ascii=False,
            indent=2,
        )
    )
    return 0


def upload_production(args: argparse.Namespace) -> int:
    if not args.confirm_production_upload:
        raise WorkflowError("upload requires --confirm-production-upload.")
    receipt = load_artifact(args.local_receipt.resolve(), "receipt_sha256", args.expected_receipt_sha256)
    if not str(receipt.get("status", "")).startswith("success"):
        raise WorkflowError("Local apply receipt is not successful.")
    plan_path = Path(receipt["plan_path"])
    plan = load_artifact(plan_path, "plan_sha256", receipt["plan_sha256"])
    registry_path = Path(plan["registry_path"])
    if sha256_file(registry_path) != plan["registry_sha256"]:
        raise WorkflowError("Workflow registry drifted after planning.")
    registry = load_registry(registry_path)
    selection = plan.get("selection") or build_selection_spec(registry)
    explicit_ids = set(selection.get("explicit_message_ids", {}).values())
    _, current_messages = discover_live_messages(registry, explicit_ids)
    current_selected, _, _ = select_messages(registry, current_messages, selection)
    current_ids = {family_id: message["message_id"] for family_id, message in current_selected.items()}
    if current_ids != plan["selected_message_ids"]:
        raise WorkflowError("Newer matching Feishu files appeared after planning; create a fresh plan.")
    tables_by_id = {table["family_id"]: table for table in plan["tables"]}
    for item in receipt["tables"]:
        target = Path(item["target_path"])
        if sha256_file(target) != item["expected_sha256"]:
            raise WorkflowError(f"Local target drifted after apply: {target}")
    uploads = []
    failed = None
    selected_order = selection["family_ids"]
    for family_id in selected_order:
        table = tables_by_id[family_id]
        command = [
            str(OPERATOR_SCRIPT),
            "upload-temp-table",
            "--file",
            table["target_path"],
            "--target-table",
            table["platform_temp_table"],
            "--target-mode",
            "reuse",
            "--import-mode",
            "overwrite",
        ]
        if args.headed:
            command.append("--headed")
        if not table["allow_baseline_target_errors"]:
            command.append("--strict-validation")
        try:
            result = run_json_command(sys.executable, command, cwd=OPERATOR_ROOT, timeout=args.timeout_seconds)
            uploads.append(
                {
                    "family_id": family_id,
                    "target_path": table["target_path"],
                    "platform_temp_table": table["platform_temp_table"],
                    "ok": bool(result.get("ok")),
                    "status": result.get("status"),
                    "import_history_row": result.get("import_history_row"),
                    "validation_result": result.get("validation_result"),
                    "elapsed_seconds": result.get("elapsed_seconds"),
                }
            )
        except Exception as exc:  # noqa: BLE001
            failed = {"family_id": family_id, "message": str(exc), "error_type": type(exc).__name__}
            break
    upload_receipt = {
        "schema_version": "1.0.0",
        "artifact_type": "QingchengTempTableUploadReceipt",
        "created_at": datetime.now().astimezone().isoformat(timespec="seconds"),
        "status": "success" if failed is None and len(uploads) == len(selected_order) else "partial_failure",
        "local_receipt_path": str(args.local_receipt.resolve()),
        "local_receipt_sha256": receipt["receipt_sha256"],
        "uploads": uploads,
        "failure": failed,
        "pending_families": selected_order[len(uploads) :] if failed else [],
    }
    upload_receipt_path = Path(plan["runtime_dir"]) / "upload_receipt.json"
    upload_sha = write_artifact(upload_receipt_path, upload_receipt, "receipt_sha256")
    summary = {
        "ok": upload_receipt["status"] == "success",
        "status": upload_receipt["status"],
        "receipt_path": str(upload_receipt_path),
        "receipt_sha256": upload_sha,
        "uploads": uploads,
        "failure": failed,
        "pending_families": upload_receipt["pending_families"],
    }
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    return 0 if summary["ok"] else 1


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Plan and execute the Qingcheng Feishu-to-temp-table workflow.")
    subparsers = parser.add_subparsers(dest="command", required=True)

    plan = subparsers.add_parser("plan", help="Discover and download current group files, then build a local dry-run plan.")
    plan.add_argument("--registry", type=Path, default=DEFAULT_REGISTRY)
    plan.add_argument("--runtime-root", type=Path, default=DEFAULT_RUNTIME_ROOT)
    plan.add_argument(
        "--family",
        action="append",
        help="Limit the plan to a registered workbook family id; repeat for multiple families.",
    )
    plan.add_argument(
        "--after",
        help="Only consider messages strictly after this ISO datetime or Unix timestamp.",
    )
    plan.add_argument(
        "--message-id",
        action="append",
        help="Bind one selected family to an exact Feishu message: <family_id>=<om_message_id>.",
    )
    plan.set_defaults(func=plan_sync)

    local = subparsers.add_parser("apply-local", help="Apply a reviewed plan to local maintenance workbooks only.")
    local.add_argument("--plan", type=Path, required=True)
    local.add_argument("--expected-plan-sha256", required=True)
    local.add_argument("--confirm-local-write", action="store_true")
    local.set_defaults(func=apply_local)

    upload = subparsers.add_parser("upload", help="Upload verified local maintenance workbooks to existing temp tables.")
    upload.add_argument("--local-receipt", type=Path, required=True)
    upload.add_argument("--expected-receipt-sha256", required=True)
    upload.add_argument("--confirm-production-upload", action="store_true")
    upload.add_argument("--headed", action="store_true")
    upload.add_argument("--timeout-seconds", type=int, default=600)
    upload.set_defaults(func=upload_production)
    return parser


def main(argv: list[str] | None = None) -> int:
    parser = build_parser()
    args = parser.parse_args(argv)
    try:
        return int(args.func(args))
    except WorkflowError as exc:
        print(json.dumps({"ok": False, "error": {"type": "workflow", "message": str(exc)}}, ensure_ascii=False, indent=2))
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
